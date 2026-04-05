import argparse
import math
import re
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook

from plot_alpha_detection_quadrants import plot_alpha_detection_quadrants


RESULT_HEADER_MAP = [
    ("秒數", "second"),
    ("正確", "true_positive"),
    ("誤抓", "false_positive"),
    ("漏抓", "miss_positive"),
    ("alpha_minus_beta", "alpha_minus_beta"),
    ("alpha_minus_theta", "alpha_minus_theta"),
]
RESULT_HEADERS = [header for header, _ in RESULT_HEADER_MAP]
SECOND_HEADERS = {"second", "seconds", "sec", "time", "秒數", "秒"}
THETA_POWER_HEADERS = {"theta_power", "thetapower"}
ALPHA_POWER_HEADERS = {"alpha_power", "alphapower"}
BETA_POWER_HEADERS = {"beta_power", "betapower"}
ALPHA_BETA_HEADERS = {"alpha_beta", "alphabeta"}
ALPHA_THETA_HEADERS = {"alpha_theta", "alphatheta"}
ALPHA_RATIO_HEADERS = {"alpha_ratio", "alpharatio", "alpha_total", "alphatotal"}
ALPHA_MINUS_BETA_HEADERS = {"alpha_minus_beta", "alphaminusbeta"}
ALPHA_MINUS_THETA_HEADERS = {"alpha_minus_theta", "alphaminustheta"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Compare alpha seconds from *_alpha.dat against *_raw_FP2.xlsx "
            "using the alpha_beta > 1 and alpha_theta > 1 rule."
        )
    )
    parser.add_argument(
        "input_path",
        nargs="?",
        help='Path to the .set folder or .set file, for example "E:\\專題\\data\\s01_061102n.set".',
    )
    parser.add_argument(
        "--output",
        help="Optional output xlsx path. Defaults to <prefix>_alpha_compare_result.xlsx in the same folder.",
    )
    parser.add_argument(
        "--plot-output",
        help="Optional output PNG path. Defaults to <prefix>_alpha_detection_quadrants.png in the same folder.",
    )
    return parser.parse_args()


def normalize_user_path(raw_path: str) -> Path:
    cleaned = raw_path.strip().strip('"').strip("'")
    return Path(cleaned)


def resolve_folder_and_prefix(input_path: Path) -> Tuple[Path, str]:
    if input_path.exists() and input_path.is_dir():
        folder = input_path
        name = input_path.name
    else:
        folder = input_path.parent if input_path.parent != Path("") else Path(".")
        if input_path.suffix.lower() == ".set":
            name = input_path.stem
        else:
            name = input_path.name

    if name.lower().endswith(".set"):
        prefix = name[:-4]
    else:
        prefix = name

    return folder.resolve(), prefix


def build_input_paths(folder: Path, prefix: str) -> Tuple[Path, Path]:
    dat_path = folder / f"{prefix}_alpha.dat"
    xlsx_path = folder / f"{prefix}_raw_FP2.xlsx"
    return dat_path, xlsx_path


def build_eye_dat_path(folder: Path, prefix: str) -> Path:
    return folder / f"{prefix}_arousal info.dat"


def parse_numeric_tokens(text: str) -> List[int]:
    matches = re.findall(r"[-+]?\d+(?:\.\d+)?", text)
    return [int(round(float(token))) for token in matches]


def load_dat_seconds(dat_path: Path) -> List[int]:
    if not dat_path.exists():
        raise FileNotFoundError(f"DAT file not found: {dat_path}")

    content = dat_path.read_text(encoding="utf-8").strip()
    if not content:
        return []

    numbers = parse_numeric_tokens(content)
    if not numbers:
        return []

    # The first value is ignored as requested.
    seconds = numbers[1:]
    seconds.sort()
    return seconds


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "")


def find_column_index(header_row: Sequence[object], aliases: set[str]) -> int:
    for idx, cell_value in enumerate(header_row):
        if normalize_header(cell_value) in aliases:
            return idx
    return -1


def find_required_columns(
    header_row: Sequence[object],
) -> Tuple[int, int, int, int, int, int, int, int, int]:
    second_col = find_column_index(header_row, SECOND_HEADERS)
    theta_power_col = find_column_index(header_row, THETA_POWER_HEADERS)
    alpha_power_col = find_column_index(header_row, ALPHA_POWER_HEADERS)
    beta_power_col = find_column_index(header_row, BETA_POWER_HEADERS)
    alpha_beta_col = find_column_index(header_row, ALPHA_BETA_HEADERS)
    alpha_theta_col = find_column_index(header_row, ALPHA_THETA_HEADERS)
    alpha_ratio_col = find_column_index(header_row, ALPHA_RATIO_HEADERS)
    alpha_minus_beta_col = find_column_index(header_row, ALPHA_MINUS_BETA_HEADERS)
    alpha_minus_theta_col = find_column_index(header_row, ALPHA_MINUS_THETA_HEADERS)

    if second_col == -1:
        raise ValueError("XLSX is missing a second column.")
    if alpha_beta_col == -1:
        raise ValueError("XLSX is missing an alpha_beta column.")
    if alpha_theta_col == -1:
        raise ValueError("XLSX is missing an alpha_theta column.")
    if alpha_ratio_col == -1:
        raise ValueError("XLSX is missing an alpha_ratio/alpha_total column.")
    if alpha_minus_beta_col == -1 and (alpha_power_col == -1 or beta_power_col == -1):
        raise ValueError("XLSX is missing an alpha_minus_beta column and cannot derive it.")
    if alpha_minus_theta_col == -1 and (alpha_power_col == -1 or theta_power_col == -1):
        raise ValueError("XLSX is missing an alpha_minus_theta column and cannot derive it.")

    return (
        second_col,
        theta_power_col,
        alpha_power_col,
        beta_power_col,
        alpha_beta_col,
        alpha_theta_col,
        alpha_ratio_col,
        alpha_minus_beta_col,
        alpha_minus_theta_col,
    )


def to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number):
        return None
    return number


def exceeds_threshold(value: float | None, threshold: float) -> bool:
    return value is not None and value > threshold


def load_second_records_from_xlsx(xlsx_path: Path) -> dict[int, dict[str, int | float | None]]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return {}

        (
            second_col,
            theta_power_col,
            alpha_power_col,
            beta_power_col,
            alpha_beta_col,
            alpha_theta_col,
            alpha_ratio_col,
            alpha_minus_beta_col,
            alpha_minus_theta_col,
        ) = find_required_columns(header_row)
        second_records: dict[int, dict[str, int | float | None]] = {}

        for row in rows:
            second_value = to_float(row[second_col] if second_col < len(row) else None)
            theta_power_value = to_float(row[theta_power_col] if theta_power_col != -1 and theta_power_col < len(row) else None)
            alpha_power_value = to_float(row[alpha_power_col] if alpha_power_col != -1 and alpha_power_col < len(row) else None)
            beta_power_value = to_float(row[beta_power_col] if beta_power_col != -1 and beta_power_col < len(row) else None)
            alpha_beta_value = to_float(row[alpha_beta_col] if alpha_beta_col < len(row) else None)
            alpha_theta_value = to_float(row[alpha_theta_col] if alpha_theta_col < len(row) else None)
            alpha_ratio_value = to_float(row[alpha_ratio_col] if alpha_ratio_col < len(row) else None)
            alpha_minus_beta_value = to_float(
                row[alpha_minus_beta_col] if alpha_minus_beta_col != -1 and alpha_minus_beta_col < len(row) else None
            )
            alpha_minus_theta_value = to_float(
                row[alpha_minus_theta_col] if alpha_minus_theta_col != -1 and alpha_minus_theta_col < len(row) else None
            )

            if alpha_minus_beta_value is None and alpha_power_value is not None and beta_power_value is not None:
                alpha_minus_beta_value = alpha_power_value - beta_power_value
            if alpha_minus_theta_value is None and alpha_power_value is not None and theta_power_value is not None:
                alpha_minus_theta_value = alpha_power_value - theta_power_value

            if second_value is None:
                continue

            second = int(round(second_value))
            second_records[second] = {
                "second": second,
                "alpha_beta": alpha_beta_value,
                "alpha_theta": alpha_theta_value,
                "alpha_ratio": alpha_ratio_value,
                "alpha_minus_beta": alpha_minus_beta_value,
                "alpha_minus_theta": alpha_minus_theta_value,
            }
    finally:
        workbook.close()

    return second_records


def is_positive_record(
    record: dict[str, int | float | None],
    beta: float = 0,
    theta: float = 0,
) -> bool:
    alpha_beta_value = record.get("alpha_beta")
    alpha_theta_value = record.get("alpha_theta")
    alpha_minus_beta_value = record.get("alpha_minus_beta")
    alpha_minus_theta_value = record.get("alpha_minus_theta")

    if not isinstance(alpha_beta_value, (int, float)) or not isinstance(alpha_theta_value, (int, float)):
        return False

    return (
        alpha_beta_value > 1
        and alpha_theta_value > 1
        and (
            exceeds_threshold(float(alpha_minus_beta_value) if isinstance(alpha_minus_beta_value, (int, float)) else None, beta)
            or exceeds_threshold(float(alpha_minus_theta_value) if isinstance(alpha_minus_theta_value, (int, float)) else None, theta)
        )
    )


def select_positive_records(
    second_records: Iterable[dict[str, int | float | None]],
    beta: float = 0,
    theta: float = 0,
) -> List[dict[str, int | float | None]]:
    positive_records = [dict(record) for record in second_records if is_positive_record(record, beta, theta)]
    positive_records.sort(key=lambda row: int(row["second"]))
    return positive_records


def load_positive_records_from_xlsx(xlsx_path: Path, beta: float = 0, theta: float = 0) -> List[dict[str, int | float | None]]:
    second_records = load_second_records_from_xlsx(xlsx_path)
    return select_positive_records(second_records.values(), beta, theta)


def build_result_row(
    second: int,
    marker_key: str,
    source_record: dict[str, int | float | None] | None = None,
) -> dict[str, int | float | None]:
    return {
        "second": second,
        "true_positive": 1 if marker_key == "true_positive" else None,
        "false_positive": 1 if marker_key == "false_positive" else None,
        "miss_positive": 1 if marker_key == "miss_positive" else None,
        "alpha_beta": source_record.get("alpha_beta") if source_record else None,
        "alpha_theta": source_record.get("alpha_theta") if source_record else None,
        "alpha_ratio": source_record.get("alpha_ratio") if source_record else None,
        "alpha_minus_beta": source_record.get("alpha_minus_beta") if source_record else None,
        "alpha_minus_theta": source_record.get("alpha_minus_theta") if source_record else None,
    }


def compare_seconds(
    dat_seconds: Sequence[int],
    xlsx_positive_records: Iterable[dict[str, int | float | None]],
    eye_dat_seconds: Sequence[int] | None = None,
    all_second_records: dict[int, dict[str, int | float | None]] | None = None,
) -> Tuple[int, int, int, List[dict[str, int | float | None]]]:
    dat_list = list(dat_seconds)
    xlsx_list = list(xlsx_positive_records)
    eye_dat_second_set = set(eye_dat_seconds or [])
    all_second_records = all_second_records or {}

    pointer = 0
    true_positive = 0
    false_positive = 0
    miss_positive = 0
    result_rows: List[dict[str, int | float | None]] = []

    for record in xlsx_list:
        second = int(record["second"])
        alpha_ratio = record.get("alpha_ratio")
        alpha_minus_beta = record.get("alpha_minus_beta")
        alpha_minus_theta = record.get("alpha_minus_theta")

        while pointer < len(dat_list) and dat_list[pointer] < second:
            miss_positive += 1
            miss_second = dat_list[pointer]
            result_rows.append(
                build_result_row(
                    miss_second,
                    "miss_positive",
                    all_second_records.get(miss_second),
                )
            )
            pointer += 1

        if pointer < len(dat_list) and dat_list[pointer] == second:
            true_positive += 1
            result_rows.append(
                build_result_row(
                    second,
                    "true_positive",
                    {
                        "alpha_beta": record.get("alpha_beta"),
                        "alpha_theta": record.get("alpha_theta"),
                        "alpha_ratio": alpha_ratio,
                        "alpha_minus_beta": alpha_minus_beta,
                        "alpha_minus_theta": alpha_minus_theta,
                    },
                )
            )
            pointer += 1
        else:
            if second in eye_dat_second_set:
                continue
            false_positive += 1
            result_rows.append(
                build_result_row(
                    second,
                    "false_positive",
                    {
                        "alpha_beta": record.get("alpha_beta"),
                        "alpha_theta": record.get("alpha_theta"),
                        "alpha_ratio": alpha_ratio,
                        "alpha_minus_beta": alpha_minus_beta,
                        "alpha_minus_theta": alpha_minus_theta,
                    },
                )
            )

    while pointer < len(dat_list):
        miss_positive += 1
        miss_second = dat_list[pointer]
        result_rows.append(
            build_result_row(
                miss_second,
                "miss_positive",
                all_second_records.get(miss_second),
            )
        )
        pointer += 1

    return true_positive, false_positive, miss_positive, result_rows


def safe_divide_float(numerator: int, denominator: int) -> float | None:
    if denominator == 0:
        return None
    return numerator / denominator


def format_float(value: float | None) -> str:
    if value is None:
        return "N/A"
    return f"{value:.6f}"


def safe_divide(numerator: int, denominator: int) -> str:
    return format_float(safe_divide_float(numerator, denominator))


def calculate_percentage_threshold(sorted_values: Sequence[float], percentage: int = 50) -> float:
    if not sorted_values:
        raise ValueError("Cannot calculate percentage threshold from empty values.")
    if percentage < 1 or percentage > 100:
        raise ValueError(f"Percentage {percentage} is out of range.")

    position = max(1, math.ceil(len(sorted_values) * percentage / 100))
    return float(sorted_values[position - 1])


def calculate_percentage_threshold_or_none(
    sorted_values: Sequence[float],
    percentage: int = 50,
) -> float | None:
    if not sorted_values:
        return None
    return calculate_percentage_threshold(sorted_values, percentage)


def collect_metric_values(
    result_rows: Sequence[dict[str, int | float | None]],
    marker_key: str,
    metric_key: str,
) -> List[float]:
    values: List[float] = []
    for row in result_rows:
        if row.get(marker_key) != 1:
            continue
        metric_value = row.get(metric_key)
        if isinstance(metric_value, (int, float)):
            values.append(float(metric_value))
    return values


def collect_metric_pairs(
    result_rows: Sequence[dict[str, int | float | None]],
    marker_key: str,
    x_key: str,
    y_key: str,
) -> Tuple[List[float], List[float]]:
    x_values: List[float] = []
    y_values: List[float] = []

    for row in result_rows:
        if row.get(marker_key) != 1:
            continue

        x_value = row.get(x_key)
        y_value = row.get(y_key)
        if not isinstance(x_value, (int, float)) or not isinstance(y_value, (int, float)):
            continue

        x_values.append(float(x_value))
        y_values.append(float(y_value))

    return x_values, y_values


def summarize_counts(
    true_positive: int,
    false_positive: int,
    miss_positive: int,
) -> dict[str, int | float | None]:
    total = true_positive + false_positive + miss_positive
    false_positive_rate = safe_divide_float(false_positive, true_positive + false_positive)
    miss_positive_rate = safe_divide_float(miss_positive, true_positive + miss_positive)

    ratio_value: float | None = None
    if false_positive_rate is not None and miss_positive_rate is not None:
        denominator = false_positive_rate + miss_positive_rate
        ratio_value = 0.0 if denominator == 0 else (
            2 * false_positive_rate * miss_positive_rate / denominator
        )

    return {
        "true_positive": true_positive,
        "false_positive": false_positive,
        "miss_positive": miss_positive,
        "total": total,
        "accuracy": safe_divide_float(true_positive, total),
        "false_positive_rate": false_positive_rate,
        "miss_positive_rate": miss_positive_rate,
        "ratio": ratio_value,
    }


def save_result_xlsx(output_path: Path, result_rows: Sequence[dict[str, int | float | None]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "result"
    worksheet.append(RESULT_HEADERS)

    for row in result_rows:
        worksheet.append([row.get(key) for _, key in RESULT_HEADER_MAP])

    workbook.save(output_path)
    workbook.close()


def default_output_path(folder: Path, prefix: str) -> Path:
    return folder / f"{prefix}_alpha_compare_result.xlsx"


def default_plot_output_path(folder: Path, prefix: str) -> Path:
    return folder / f"{prefix}_alpha_detection_quadrants.png"


def main() -> int:
    args = parse_args()
    raw_input_path = args.input_path
    if not raw_input_path:
        raw_input_path = input("請輸入 .set 資料夾或檔案路徑: ")

    input_path = normalize_user_path(raw_input_path)
    folder, prefix = resolve_folder_and_prefix(input_path)
    dat_path, xlsx_path = build_input_paths(folder, prefix)
    eye_dat_path = build_eye_dat_path(folder, prefix)

    output_path = normalize_user_path(args.output) if args.output else default_output_path(folder, prefix)
    plot_output_path = (
        normalize_user_path(args.plot_output)
        if args.plot_output
        else default_plot_output_path(folder, prefix)
    )

    dat_seconds = load_dat_seconds(dat_path)
    eye_dat_seconds = load_dat_seconds(eye_dat_path) if eye_dat_path.exists() else []
    all_second_records = load_second_records_from_xlsx(xlsx_path)
    xlsx_positive_records = select_positive_records(all_second_records.values())

    true_positive, false_positive, miss_positive, result_rows = compare_seconds(
        dat_seconds,
        xlsx_positive_records,
        eye_dat_seconds,
        all_second_records,
    )
    summary = summarize_counts(true_positive, false_positive, miss_positive)

    tp_alpha_minus_beta = sorted(collect_metric_values(result_rows, "true_positive", "alpha_minus_beta"))
    tp_alpha_minus_theta = sorted(collect_metric_values(result_rows, "true_positive", "alpha_minus_theta"))
    tp_alpha_minus_beta_median = calculate_percentage_threshold_or_none(
        tp_alpha_minus_beta
    )
    tp_alpha_minus_theta_median = calculate_percentage_threshold_or_none(
        tp_alpha_minus_theta
    )

    fp_alpha_minus_beta = sorted(collect_metric_values(result_rows, "false_positive", "alpha_minus_beta"))
    fp_alpha_minus_theta = sorted(collect_metric_values(result_rows, "false_positive", "alpha_minus_theta"))
    fp_alpha_minus_beta_median = calculate_percentage_threshold_or_none(
        fp_alpha_minus_beta
    )
    fp_alpha_minus_theta_median = calculate_percentage_threshold_or_none(
        fp_alpha_minus_theta
    )

    tp_alpha_beta, tp_alpha_theta = collect_metric_pairs(
        result_rows,
        "true_positive",
        "alpha_beta",
        "alpha_theta",
    )
    fp_alpha_beta, fp_alpha_theta = collect_metric_pairs(
        result_rows,
        "false_positive",
        "alpha_beta",
        "alpha_theta",
    )
    miss_alpha_beta, miss_alpha_theta = collect_metric_pairs(
        result_rows,
        "miss_positive",
        "alpha_beta",
        "alpha_theta",
    )

    save_result_xlsx(output_path, result_rows)
    plot_alpha_detection_quadrants(
        tp_alpha_beta,
        tp_alpha_theta,
        fp_alpha_beta,
        fp_alpha_theta,
        miss_alpha_beta,
        miss_alpha_theta,
        title=f"{prefix} Alpha Detection Quadrants",
        output_path=plot_output_path,
    )

    print(f"資料夾: {folder}")
    print(f"前綴: {prefix}")
    print(f"DAT: {dat_path}")
    print(f"眼動DAT: {eye_dat_path}")
    print(f"XLSX: {xlsx_path}")
    print(f"輸出: {output_path}")
    print(f"四象限圖: {plot_output_path}")
    print(f"true_positive: {true_positive}")
    print(f"false_positive: {false_positive}")
    print(f"miss_positive: {miss_positive}")
    print(f"total: {summary['total']}")
    print(f"正確率 (true_positive / total): {format_float(summary['accuracy'])}")
    print(
        "誤抓率 (false_positive / (true_positive + false_positive)): "
        f"{format_float(summary['false_positive_rate'])}"
    )
    print(
        "漏抓率 (miss_positive / (true_positive + miss_positive)): "
        f"{format_float(summary['miss_positive_rate'])}"
    )
    print(f"正確 alpha_minus_beta 中位數: {format_float(tp_alpha_minus_beta_median)}")
    print(f"正確 alpha_minus_theta 中位數: {format_float(tp_alpha_minus_theta_median)}")
    print(f"誤抓 alpha_minus_beta 中位數: {format_float(fp_alpha_minus_beta_median)}")
    print(f"誤抓 alpha_minus_theta 中位數: {format_float(fp_alpha_minus_theta_median)}")

    tp_fp_beta = sorted(tp_alpha_minus_beta + fp_alpha_minus_beta)
    tp_fp_theta = sorted(tp_alpha_minus_theta + fp_alpha_minus_theta)
    if tp_fp_beta and tp_fp_theta:
        min_index = 30
        min_value: float | None = None
        for i in range(30,81):
            beta = calculate_percentage_threshold(tp_fp_beta, i)
            theta = calculate_percentage_threshold(tp_fp_theta, i)
            xlsx_positive_records = select_positive_records(all_second_records.values(), beta, theta)
            true_positive, false_positive, miss_positive, result_rows = compare_seconds(
                dat_seconds,
                xlsx_positive_records,
                eye_dat_seconds,
                all_second_records,
            )
            print(f"\n第{i}%: ")
            print(f"true_positve: {true_positive}" )
            print(f"false_positive: {false_positive}" )
            print(f"miss_positive: {miss_positive}" )
            print(f"total: {true_positive + false_positive + miss_positive}" )
            print("誤抓率: " + safe_divide(false_positive, true_positive + false_positive))
            print("漏抓率: " + safe_divide(miss_positive, true_positive + miss_positive))
            false_rate = safe_divide_float(false_positive, true_positive + false_positive)
            miss_rate = safe_divide_float(miss_positive, true_positive + miss_positive)
            ratio = 2 * false_rate * miss_rate / (false_rate + miss_rate) if (false_rate is not None and miss_rate is not None and (false_rate + miss_rate) > 0) else None
            print(f"ratio: {ratio if ratio is not None else 'N/A'}" )
            if ratio is not None and (min_value is None or ratio < min_value):
                min_value = ratio
                min_index = i

        print(f"\n最佳解: {min_index}% -> ratio={format_float(min_value)}")
    else:
        print("\n最佳解: N/A (沒有足夠的 alpha_minus 資料可做門檻搜尋)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
