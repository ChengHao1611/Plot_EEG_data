import argparse
import os
from typing import Dict, List, Tuple

import numpy as np
import pyedflib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

RED_FONT = Font(color="FFFF0000")

XLSX_FIELDNAMES = [
    "second",
    "theta_power",
    "alpha_power",
    "beta_power",
    "alpha_beta",
    "alpha_theta",
    "alpha_total",
    "alpha_peak",
    "alpha_minus_beta",
    "alpha_minus_theta",
    "eyeblinking_count",
    "react_time",
]

def load_eyeblinkning(data_path: str) -> List[int]:
    if not os.path.exists(data_path):
        print(f"Warning: arousal data file not found: {data_path}")
        return []
    try:
        with open(data_path, "r", encoding="utf-8") as f:
            content = f.read().strip()
            if not content:
                return []
            parts = content.split(",")
            blinks_seconds = [int(float(p)) for p in parts[1:]]  # Skip the first part (count)
            return blinks_seconds
    except Exception as e:
        print(f"Error reading arousal data from {data_path}: {e}")
        return []

def find_channel_index(labels: List[str], target: str) -> int | None:
    target_lower = target.lower()
    for i, label in enumerate(labels):
        if target_lower in label.lower():
            return i
    return None

def load_channel_signal(edf_path: str, channel_name: str) -> Tuple[np.ndarray, float]:
    if not os.path.exists(edf_path):
        raise FileNotFoundError(f"EDF not found: {edf_path}")

    f = pyedflib.EdfReader(edf_path)
    try:
        labels = f.getSignalLabels()
        ch_idx = find_channel_index(labels, channel_name)
        if ch_idx is None:
            raise ValueError(f"Channel '{channel_name}' not found in {edf_path}")

        signal = f.readSignal(ch_idx).astype(float)
        fs = float(f.getSampleFrequency(ch_idx))
    finally:
        f.close()

    return signal, fs


def compute_band_powers_and_ratios_fft(
    signal: np.ndarray,
    fs: float,
    *,
    theta_low: float,
    theta_high: float,
    alpha_low: float,
    alpha_high: float,
    beta_low: float,
    beta_high: float,
    use_psd: bool = False,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    if signal.size == 0 or fs <= 0:
        empty = np.array([], dtype=float)
        return empty, empty, empty, empty, empty, empty, empty, empty, empty

    win = int(round(fs))
    if win <= 0:
        empty = np.array([], dtype=float)
        return empty, empty, empty, empty, empty, empty, empty, empty, empty

    n_secs = int(signal.size // win)
    if n_secs == 0:
        empty = np.array([], dtype=float)
        return empty, empty, empty, empty, empty, empty, empty, empty, empty

    freqs = np.fft.rfftfreq(win, d=1.0 / fs)
    theta_mask = (freqs >= theta_low) & (freqs < theta_high)
    alpha_mask = (freqs >= alpha_low) & (freqs < alpha_high)
    beta_mask = (freqs >= beta_low) & (freqs <= beta_high)
    total_mask = (freqs >= 1) & (freqs <= 40)

    if not np.any(theta_mask) or not np.any(alpha_mask) or not np.any(beta_mask):
        empty = np.array([], dtype=float)
        return empty, empty, empty, empty, empty, empty, empty, empty, empty

    window = np.hanning(win)
    window_power = float(np.sum(window ** 2))
    df = float(freqs[1] - freqs[0]) if freqs.size > 1 else 0.0
    peak_mask = (freqs >= 2.0) & (freqs <= 30.0)
    theta_power = np.full(n_secs, np.nan, dtype=float)
    alpha_power = np.full(n_secs, np.nan, dtype=float)
    beta_power = np.full(n_secs, np.nan, dtype=float)
    alpha_beta = np.full(n_secs, np.nan, dtype=float)
    alpha_theta = np.full(n_secs, np.nan, dtype=float)
    alpha_total = np.full(n_secs, np.nan, dtype=float)
    alpha_peak = np.full(n_secs, np.nan, dtype=float)
    alpha_minus_beta = np.full(n_secs, np.nan, dtype=float)
    alpha_minus_theta = np.full(n_secs, np.nan, dtype=float)


    for s in range(n_secs):
        start = s * win
        end = start + win
        seg = signal[start:end]
        if seg.size < win:
            continue
        if np.isnan(seg).any():
            continue
        seg = seg - float(np.mean(seg))
        seg = seg * window
        spec = np.fft.rfft(seg)
        power = np.abs(spec) ** 2
        if use_psd:
            if fs <= 0.0 or window_power <= 0.0:
                continue
            power = power / (float(fs) * window_power)
            # One-sided PSD scaling for real signals:
            # double all bins except DC (0 Hz) and Nyquist (fs/2, when N even).
            if win % 2 == 0:
                if power.size > 2:
                    power[1:-1] *= 2.0
            else:
                if power.size > 1:
                    power[1:] *= 2.0

        if np.any(peak_mask):
            peak_rel_idx = int(np.argmax(power[peak_mask]))
            peak_abs_idx = int(np.flatnonzero(peak_mask)[peak_rel_idx])
            peak_freq = float(freqs[peak_abs_idx])
            alpha_peak[s] = 1.0 if (peak_freq >= float(alpha_low) and peak_freq < float(alpha_high)) else 0.0

        p_theta = float(np.sum(power[theta_mask]))
        p_alpha = float(np.sum(power[alpha_mask]))
        p_beta = float(np.sum(power[beta_mask]))
        p_total = float(np.sum(power[total_mask]))
        if use_psd and df > 0.0:
            # Convert PSD (power/Hz) to band power by integrating over frequency.
            p_theta *= df
            p_alpha *= df
            p_beta *= df
            p_total *= df

        theta_power[s] = p_theta
        alpha_power[s] = p_alpha
        beta_power[s] = p_beta
        alpha_minus_beta[s] = p_alpha - p_beta
        alpha_minus_theta[s] = p_alpha - p_theta

        if p_beta > 0.0:
            alpha_beta[s] = p_alpha / p_beta
        if p_theta > 0.0:
            alpha_theta[s] = p_alpha / p_theta
        
        if p_total > 0.0:
            alpha_total[s] = p_alpha / p_total

    return theta_power, alpha_power, beta_power, alpha_beta, alpha_theta, alpha_total, alpha_minus_beta, alpha_minus_theta, alpha_peak


def extract_react_times(edf_path: str) -> Dict[int, float]:
    """
    使用 stage 1->2 的規則，react_time = sec_253 - sec_251。
    t = sec，sec 為 1-based 秒數。
    回傳 {second_key: react_time}，second_key = round(sec_251 * 10) 的整數 (1 位小數)。
    """
    f = pyedflib.EdfReader(edf_path)
    try:
        labels = f.getSignalLabels()
        status_idx = find_channel_index(labels, "status")
        if status_idx is None:
            raise ValueError("Status channel not found in EDF.")

        status_signal = f.readSignal(status_idx)
        fs = float(f.getSampleFrequency(status_idx))
        if fs <= 0:
            raise ValueError("Invalid sample rate.")

        total_samples = len(status_signal)
        total_seconds = int(total_samples // fs)

        stage = 1
        sec_251 = None
        events: Dict[int, float] = {}

        for sec in range(1, total_seconds + 1):
            start = int((sec - 1) * fs)
            end = int(sec * fs)
            segment = status_signal[start:end]

            for i in range(len(segment)):
                if segment[i] > 1:
                    t = float(sec)
                    if stage == 1:
                        sec_251 = t + 0.002 * i
                        stage = 2
                    elif stage == 2:
                        sec_253 = t + 0.002 * i
                        stage = 3
                    elif stage == 3:
                        if sec_251 is not None and sec_253 is not None:
                            sec_251_round = int(round(sec_251, 0))
                            react_time = round(sec_253 - sec_251, 1)
                            key = int(round(sec_251_round * 10))
                            #print(sec_251, sec_251_round,key, react_time)
                            if key not in events:
                                events[key] = react_time
                        stage = 1
    finally:
        f.close()

    return events


def apply_ratio_highlights(ws, fieldnames: List[str]) -> None:
    highlight_columns = [
        fieldnames.index(name) + 1
        for name in ("alpha_beta", "alpha_theta")
        if name in fieldnames
    ]

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in highlight_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if float(cell.value) > 1.0:
                    cell.font = RED_FONT
            except (TypeError, ValueError):
                continue


def merge_react_time_into_xlsx(xlsx_path: str, events: Dict[int, float]) -> None:
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    wb = load_workbook(xlsx_path)
    ws = wb.active
    fieldnames = [cell.value for cell in ws[1] if cell.value is not None]

    if "second" not in fieldnames:
        raise ValueError("XLSX must contain 'second' column.")

    for name in XLSX_FIELDNAMES:
        if name not in fieldnames:
            fieldnames.append(name)

    rows: List[Dict[str, float | int | str | None]] = []
    for row_idx in range(2, ws.max_row + 1):
        row: Dict[str, float | int | str | None] = {}
        for col_idx, name in enumerate(fieldnames, start=1):
            row[name] = ws.cell(row=row_idx, column=col_idx).value
        rows.append(row)

    def second_to_key(val: float | int | str | None) -> int | None:
        try:
            return int(round(float(val) * 10))
        except Exception:
            return None

    existing_keys = set()
    for row in rows:
        key = second_to_key(row.get("second"))
        if key is None:
            continue
        existing_keys.add(key)
        if key in events:
            row["react_time"] = round(events[key], 1)

    for key, react_time in events.items():
        if key in existing_keys:
            continue
        sec_value = key / 10.0
        new_row = {name: None for name in fieldnames}
        new_row["second"] = sec_value
        new_row["react_time"] = round(react_time, 1)
        rows.append(new_row)

    def sort_key(row: Dict[str, float | int | str | None]) -> float:
        try:
            return float(row.get("second", ""))
        except Exception:
            return float("inf")

    rows.sort(key=sort_key)

    ws.delete_rows(1, ws.max_row)
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(name) for name in fieldnames])
    apply_ratio_highlights(ws, fieldnames)
    wb.save(xlsx_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Compute per-second theta/alpha/beta FFT band power (or PSD-based band power) and "
            "alpha/beta, alpha/theta ratios for FP2 channel."
        )
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--edf", help="EDF file path")

    parser.add_argument("--channel", default="FP2", help="Channel name to match (default: FP2)")

    parser.add_argument("--theta-low", type=float, default=4.0, help="Theta band low cutoff (Hz)")
    parser.add_argument("--theta-high", type=float, default=8.0, help="Theta band high cutoff (Hz)")
    parser.add_argument("--alpha-low", type=float, default=8.0, help="Alpha band low cutoff (Hz)")
    parser.add_argument("--alpha-high", type=float, default=13.0, help="Alpha band high cutoff (Hz)")
    parser.add_argument("--beta-low", type=float, default=13.0, help="Beta band low cutoff (Hz)")
    parser.add_argument("--beta-high", type=float, default=30.0, help="Beta band high cutoff (Hz)")
    parser.add_argument(
        "--psd",
        action="store_true",
        help=(
            "Compute band powers by integrating one-sided PSD (power/Hz) instead of raw |FFT|^2 bin sums. "
            "Enables window-power normalization and one-sided scaling."
        ),
    )

    return parser.parse_args()


def sanitize_filename_part(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in value)


def default_xlsx_path(edf_path: str, channel_name: str) -> str:
    root, _ = os.path.splitext(edf_path)
    safe_channel = sanitize_filename_part(channel_name)
    return f"{root}_{safe_channel}.xlsx"

def save_xlsx(
    output_path: str,
    rows: List[Dict[str, float | int | str | None]],
) -> None:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "result"
    ws.append(XLSX_FIELDNAMES)
    for row in rows:
        ws.append([row.get(name) for name in XLSX_FIELDNAMES])
    apply_ratio_highlights(ws, XLSX_FIELDNAMES)
    wb.save(output_path)

def main() -> int:
    args = parse_args()
    if not args.edf:
        print("No EDF file to process.")
        return 1
    edf_path = os.path.abspath(args.edf)

    total_secs = 0
    total_valid = 0
    total_skipped = 0
    output_rows: List[Dict[str, float | int | str | None]] = []
    dat_root, _ = os.path.splitext(edf_path)
    dat_path = dat_root + "_arousal info.dat"
    blink_seconds = load_eyeblinkning(dat_path)
    signal, fs = load_channel_signal(edf_path, args.channel)
    theta_power, alpha_power, beta_power, alpha_beta, alpha_theta, alpha_total, alpha_minus_beta, alpha_minus_theta, alpha_peak = compute_band_powers_and_ratios_fft(
        signal,
        fs,
        theta_low=args.theta_low,
        theta_high=args.theta_high,
        alpha_low=args.alpha_low,
        alpha_high=args.alpha_high,
        beta_low=args.beta_low,
        beta_high=args.beta_high,
        use_psd=bool(args.psd),
    )

    n_secs = int(theta_power.size)
    valid_mask = ~(np.isnan(theta_power) | np.isnan(alpha_power) | np.isnan(beta_power))
    valid_secs = int(np.sum(valid_mask))
    skipped = int(n_secs - valid_secs)

    total_secs += n_secs
    total_valid += valid_secs
    total_skipped += skipped

    print(
        f"{os.path.basename(edf_path)}: seconds={n_secs}, "
        f"valid={valid_secs}, skipped={skipped}"
    )

    for sec_idx in range(n_secs):
        t = sec_idx + 1
        start_range = max(0, t - 30)
        count = sum(1 for blink in blink_seconds if start_range <= blink <= t)
        t_val = float(theta_power[sec_idx]) if not np.isnan(theta_power[sec_idx]) else float("nan")
        a_val = float(alpha_power[sec_idx]) if not np.isnan(alpha_power[sec_idx]) else float("nan")
        b_val = float(beta_power[sec_idx]) if not np.isnan(beta_power[sec_idx]) else float("nan")
        ab_val = float(alpha_beta[sec_idx]) if not np.isnan(alpha_beta[sec_idx]) else float("nan")
        at_val = float(alpha_theta[sec_idx]) if not np.isnan(alpha_theta[sec_idx]) else float("nan")
        atotal_val = float(alpha_total[sec_idx]) if not np.isnan(alpha_total[sec_idx]) else float("nan")
        apeak_raw = float(alpha_peak[sec_idx]) if not np.isnan(alpha_peak[sec_idx]) else float("nan")
        apeak_val = int(round(apeak_raw)) if np.isfinite(apeak_raw) else None
        a_minus_b_val = float(alpha_minus_beta[sec_idx]) if not np.isnan(alpha_minus_beta[sec_idx]) else float("nan")
        a_minus_t_val = float(alpha_minus_theta[sec_idx]) if not np.isnan(alpha_minus_theta[sec_idx]) else float("nan")

        output_rows.append({
            "second": t,
            "theta_power": t_val,
            "alpha_power": a_val,
            "beta_power": b_val,
            "alpha_beta": ab_val,
            "alpha_theta": at_val,
            "alpha_total": atotal_val,
            "alpha_peak": apeak_val,
            "alpha_minus_beta": a_minus_b_val,
            "alpha_minus_theta": a_minus_t_val,
            "eyeblinking_count": count,
            "react_time": None,
        })

    print("--- Summary ---")
    print(f"Total seconds: {total_secs}")
    print(f"Valid seconds: {total_valid}")
    print(f"Skipped seconds (invalid): {total_skipped}")

    xlsx_output_path = default_xlsx_path(edf_path, args.channel)
    save_xlsx(xlsx_output_path, output_rows)
    events = extract_react_times(edf_path)
    merge_react_time_into_xlsx(xlsx_output_path, events)
    print(f"Saved per-second values to: {xlsx_output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
