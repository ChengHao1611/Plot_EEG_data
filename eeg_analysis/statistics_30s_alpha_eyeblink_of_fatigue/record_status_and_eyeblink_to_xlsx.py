import math
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pyedflib
import numpy as np
from openpyxl import Workbook  # 新增
import os  # 新增
from openpyxl.utils import get_column_letter


DEVIATION_START_CODES = frozenset({251, 252})
CORRECTION_START_CODE = 253
CORRECTION_END_CODE = 254


def round_reaction_time(value: float) -> float:
    """Round Reaction Time to one decimal using conventional half-up rounding."""
    if not math.isfinite(value):
        raise ValueError(f"Reaction Time 必須是有限數值：{value}")
    return float(
        Decimal(str(value)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
    )


@dataclass
class ReactionTimeEvent:
    """One lane-deviation event reconstructed from the EDF Status channel."""

    event_index: int
    deviation_status: int
    deviation_time: float
    correction_start_time: float
    event_second: int
    reaction_time: float
    correction_end_time: float | None = None

    @property
    def return_to_lane_time(self) -> float | None:
        if self.correction_end_time is None:
            return None
        return self.correction_end_time - self.correction_start_time


def _find_status_channel_index(labels) -> int:
    for index, label in enumerate(labels):
        if "status" in str(label).casefold():
            return index
    raise ValueError("未找到 Status 通道，請確認 EDF 通道標籤。")


def parse_reaction_time_events(
    status_signal, sample_rate: float
) -> list[ReactionTimeEvent]:
    """Parse a Status array containing 251/252 -> 253 -> 254 sequences."""
    if sample_rate <= 0:
        raise ValueError(f"Status 通道取樣率無效：{sample_rate}")

    rounded_status = np.rint(status_signal).astype(int)
    if rounded_status.size == 0:
        return []
    change_indices = np.flatnonzero(
        np.r_[True, rounded_status[1:] != rounded_status[:-1]]
    )

    events: list[ReactionTimeEvent] = []
    deviation_status: int | None = None
    deviation_time: float | None = None
    deviation_sample_index: int | None = None
    pending_completion: ReactionTimeEvent | None = None

    for sample_index in change_indices:
        status_code = int(rounded_status[sample_index])
        event_time = float(sample_index) / sample_rate

        if status_code in DEVIATION_START_CODES:
            deviation_status = status_code
            deviation_time = event_time
            deviation_sample_index = int(sample_index)
            pending_completion = None
            continue

        if status_code == CORRECTION_START_CODE:
            if (
                deviation_status is None
                or deviation_time is None
                or deviation_sample_index is None
            ):
                continue
            raw_reaction_time = (
                float(int(sample_index) - deviation_sample_index) / sample_rate
            )
            if raw_reaction_time < 0:
                deviation_status = None
                deviation_time = None
                deviation_sample_index = None
                continue
            reaction_time = round_reaction_time(raw_reaction_time)
            event = ReactionTimeEvent(
                event_index=len(events) + 1,
                deviation_status=deviation_status,
                deviation_time=deviation_time,
                correction_start_time=event_time,
                event_second=int(math.ceil(deviation_time)),
                reaction_time=reaction_time,
            )
            events.append(event)
            pending_completion = event
            deviation_status = None
            deviation_time = None
            deviation_sample_index = None
            continue

        if status_code == CORRECTION_END_CODE and pending_completion is not None:
            pending_completion.correction_end_time = event_time
            pending_completion = None

    return events


def extract_reaction_time_events_with_duration(
    edf_path,
) -> tuple[list[ReactionTimeEvent], int]:
    """Extract RT events and the complete recording end second from an EDF.

    Status 251 and 252 both mean that the vehicle starts deviating.  Status 253
    marks the driver's correction start and 254 marks correction completion.
    The event second is the upward-rounded deviation timestamp. Reaction Time
    is conventionally rounded to one decimal place before downstream use.
    """
    path = Path(edf_path)
    if not path.is_file():
        raise FileNotFoundError(f"EDF not found: {path}")

    reader = pyedflib.EdfReader(str(path))
    try:
        status_index = _find_status_channel_index(reader.getSignalLabels())
        status_signal = reader.readSignal(status_index)
        sample_rate = float(reader.getSampleFrequency(status_index))
    finally:
        reader.close()
    recording_end_second = int(len(status_signal) // sample_rate)
    return (
        parse_reaction_time_events(status_signal, sample_rate),
        recording_end_second,
    )


def extract_reaction_time_events(edf_path) -> list[ReactionTimeEvent]:
    """Extract 251/252 -> 253 reaction-time events from an EDF."""
    events, _ = extract_reaction_time_events_with_duration(edf_path)
    return events


def write_reaction_time_events_xlsx(
    events: list[ReactionTimeEvent],
    output_path,
    *,
    recording_end_second: int | None = None,
) -> Path:
    """Write all RT events and optional recording-duration metadata."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "reaction_time_events"
    worksheet.append(
        [
            "事件編號",
            "偏移Status",
            "偏移開始時間_秒",
            "導正開始時間_秒",
            "向上取整事件秒數",
            "Reaction Time_秒",
            "完成導正時間_秒",
            "導回車道用時_秒",
        ]
    )
    for event in events:
        worksheet.append(
            [
                event.event_index,
                event.deviation_status,
                event.deviation_time,
                event.correction_start_time,
                event.event_second,
                event.reaction_time,
                event.correction_end_time,
                event.return_to_lane_time,
            ]
        )
    for column in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 22
    for cell in worksheet["F"][1:]:
        cell.number_format = "0.0"
    if recording_end_second is not None:
        metadata = workbook.create_sheet("recording_metadata")
        metadata.append(["Item", "Value"])
        metadata.append(["Recording End Second", int(recording_end_second)])
        metadata.column_dimensions["A"].width = 28
        metadata.column_dimensions["B"].width = 18
    workbook.save(output)
    return output

def process_alpha_data(ws, base_path, total_seconds):
    """
    處理 α 波資料並以 10 秒滑動視窗填入 C 欄。
    例如輸入 abc_raw.EDF 時，讀取同資料夾的 abc_alpha.dat。
    """
    dat_dir = os.path.dirname(base_path)
    edf_stem = os.path.basename(base_path)
    record_id = edf_stem[:-4] if edf_stem.lower().endswith("_raw") else edf_stem
    alpha_path = os.path.join(dat_dir, f"{record_id}_alpha.dat")

    numbers = []
    if not os.path.exists(alpha_path):
        print(f"警告：找不到 α 波資料檔案 {alpha_path}")
    else:
        with open(alpha_path, "r", encoding="utf-8") as f:
            content = f.read().strip()

        if not content:
            print(f"警告：{alpha_path} 檔案為空")
        else:
            try:
                numbers = [int(x.strip()) for x in content.split(",") if x.strip()]
            except ValueError:
                print(f"警告：{alpha_path} 資料格式不正確")
                numbers = []

    last_second = int(total_seconds)

    # 第一個數字為 α 波秒數的總數量，後面才是實際秒數。
    alpha_counts = {}
    for second in numbers[1:]:
        if 0 <= second <= last_second:
            alpha_counts[second] = alpha_counts.get(second, 0) + 1
        else:
            print(f"警告：忽略超出 EDF 時間範圍的 α 波秒數 {second}")

    # 每秒統計當前秒與前 9 秒，等同 rolling(window=10) 的結果。
    rolling_alpha_counts = {}
    rolling_count = 0
    for second in range(last_second + 1):
        rolling_count += alpha_counts.get(second, 0)
        if second >= 10:
            rolling_count -= alpha_counts.get(second - 10, 0)
        rolling_alpha_counts[second] = rolling_count

    # 保留原本的 Status 資料列，再補上尚不存在的每一整數秒資料列。
    rows = []
    existing_integer_seconds = set()
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=column).value for column in range(1, 7)]
        a_value = values[0]
        if a_value is None:
            continue

        try:
            second_value = float(a_value)
        except (TypeError, ValueError):
            continue

        if second_value.is_integer() and 0 <= second_value <= last_second:
            integer_second = int(second_value)
            values[2] = rolling_alpha_counts[integer_second]
            existing_integer_seconds.add(integer_second)

        rows.append((second_value, row, values))

    for second, alpha_count in rolling_alpha_counts.items():
        if second not in existing_integer_seconds:
            rows.append((float(second), -1, [float(second), None, alpha_count, None, None, None]))

    rows.sort(key=lambda item: (item[0], item[1]))

    # 重新寫回排序後的列，讓每一整數秒都有 α 波滑動視窗結果。
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for row, (_, _, values) in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            ws.cell(row=row, column=column, value=value)

def process_eye_blink_data(ws, base_path, total_seconds):
    """
    處理眼動資料並以 30 秒滑動視窗填入 F 欄。
    base_path: EDF 檔案的路徑（不含副檔名）
    """
    # 尋找對應的 .dat 檔案
    dat_dir = os.path.dirname(base_path)
    edf_stem = os.path.basename(base_path)
    record_id = edf_stem[:-4] if edf_stem.lower().endswith("_raw") else edf_stem
    dat_filename = record_id + "_raw_arousal info.dat"
    dat_path = os.path.join(dat_dir, dat_filename)
    
    numbers = []
    if not os.path.exists(dat_path):
        print(f"警告：找不到眼動資料檔案 {dat_path}")
    else:
        with open(dat_path, "r", encoding="utf-8") as f:
            content = f.read().strip()

        if not content:
            print(f"警告：{dat_path} 檔案為空")
        else:
            try:
                numbers = [int(x.strip()) for x in content.split(",") if x.strip()]
            except ValueError:
                print(f"警告：{dat_path} 資料格式不正確")
                numbers = []

    last_second = int(total_seconds)
    eye_counts = {}
    for second in numbers[1:]:
        if 0 <= second <= last_second:
            eye_counts[second] = eye_counts.get(second, 0) + 1
        else:
            print(f"警告：忽略超出 EDF 時間範圍的眼動秒數 {second}")

    # 每秒統計當前秒與前 29 秒，等同 rolling(window=30) 的結果。
    rolling_eye_counts = {}
    rolling_count = 0
    for second in range(last_second + 1):
        rolling_count += eye_counts.get(second, 0)
        if second >= 30:
            rolling_count -= eye_counts.get(second - 30, 0)
        rolling_eye_counts[second] = rolling_count

    # process_alpha_data 已建立完整逐秒時間軸；只需將結果填入每個整數秒。
    for row in range(2, ws.max_row + 1):
        a_value = ws.cell(row=row, column=1).value
        if a_value is None:
            continue

        try:
            second_value = float(a_value)
        except (TypeError, ValueError):
            continue

        if second_value.is_integer() and 0 <= second_value <= last_second:
            ws.cell(row=row, column=6, value=rolling_eye_counts[int(second_value)])

def check_status_253(edf_path, tolerance=0.05):
    """Legacy workbook export backed by the explicit Status-code parser."""
    del tolerance  # Kept in the public signature for backward compatibility.
    base_path, _ = os.path.splitext(edf_path)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, "data")
    os.makedirs(output_dir, exist_ok=True)
    xlsx_filename = os.path.basename(base_path) + ".xlsx"
    xlsx_path = os.path.join(output_dir, xlsx_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "result"

    # A1~F1 標題
    ws["A1"] = "秒數"
    ws["B1"] = "事件反應時間"
    ws["C1"] = "α波次數（10秒滑動）"
    ws["D1"] = "導回車道用時"
    ws["E1"] = "睡著"
    ws["F1"] = "眼動次數（30秒滑動）"

    events = extract_reaction_time_events(edf_path)
    for next_row, event in enumerate(events, start=2):
        ws.cell(row=next_row, column=1, value=event.event_second)
        ws.cell(row=next_row, column=2, value=round(event.reaction_time, 1))
        ws.cell(row=next_row, column=2).number_format = "0.0"
        ws.cell(
            row=next_row,
            column=4,
            value=(
                round(event.return_to_lane_time, 3)
                if event.return_to_lane_time is not None
                else None
            ),
        )

    reader = pyedflib.EdfReader(edf_path)
    try:
        status_index = _find_status_channel_index(reader.getSignalLabels())
        sample_rate = float(reader.getSampleFrequency(status_index))
        total_seconds = int(len(reader.readSignal(status_index)) // sample_rate)
    finally:
        reader.close()

    print(f"總長度: {total_seconds} 秒，Reaction Time事件數：{len(events)}")

    # --- 處理 α 波資料並填入 C 欄 ---
    process_alpha_data(ws, base_path, total_seconds)

    # --- 處理眼動資料並填入 F 欄 ---
    process_eye_blink_data(ws, base_path, total_seconds)
    
    # --- 新增：儲存 xlsx 檔案 ---
    wb.save(xlsx_path)
    return xlsx_path

if __name__ == "__main__":
    edf_file = input("請輸入 EDF 檔案路徑: ")
    edf_file = edf_file.replace('"', '')
    check_status_253(edf_file)
