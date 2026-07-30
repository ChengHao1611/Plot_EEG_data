"""Plot Alpha and eye distributions at one second before each fatigue segment.

The input is ``fatigue_segment_pre30_summary.xlsx`` produced by
``plot_fatigue_segment_pre30_distributions.py``.  The default target is relative
second -15: exactly 15 seconds before a fatigue segment starts.  One Alpha and
one eye value are extracted from every fatigue segment across all nine source
recordings.
"""

from __future__ import annotations

import argparse
import statistics
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SOURCE_COLUMN = "來源檔案"
SEGMENT_COLUMN = "疲勞區段編號"
START_SECOND_COLUMN = "區段開始秒數"
END_SECOND_COLUMN = "區段結束秒數"
OFFSET_COLUMN = "相對秒數"
ABSOLUTE_SECOND_COLUMN = "絕對秒數"
ALPHA_COLUMN = "α波次數（10秒滑動）"
EYE_COLUMN = "眼動次數（30秒滑動）"
DEFAULT_SHEET_NAME = "前30秒逐區段"
DEFAULT_RELATIVE_SECOND = -15


@dataclass(frozen=True)
class SegmentPoint:
    """The Alpha and eye values at one relative second for a fatigue segment."""

    source_file: str
    segment_index: int
    start_second: int
    end_second: int | None
    relative_second: int
    absolute_second: int
    alpha_value: float
    eye_value: float


def to_finite_float(value: object) -> float | None:
    """Return a finite float from a worksheet cell, otherwise None."""
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if np.isfinite(number) else None


def to_int(value: object) -> int | None:
    """Return an integer from a worksheet cell, otherwise None."""
    number = to_finite_float(value)
    return int(number) if number is not None else None


def read_segment_points(
    input_path: Path,
    *,
    sheet_name: str,
    relative_second: int,
) -> tuple[list[SegmentPoint], list[list[object]], set[str]]:
    """Read one requested relative-second value from every fatigue segment."""
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(f"找不到工作表「{sheet_name}」。可用工作表：{available}")
        worksheet = workbook[sheet_name]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            raise ValueError("輸入工作表缺少標題列。")

        column_indexes = {
            str(name): index for index, name in enumerate(header) if name is not None
        }
        required_columns = {
            SOURCE_COLUMN,
            SEGMENT_COLUMN,
            START_SECOND_COLUMN,
            END_SECOND_COLUMN,
            OFFSET_COLUMN,
            ABSOLUTE_SECOND_COLUMN,
            ALPHA_COLUMN,
            EYE_COLUMN,
        }
        missing = required_columns - set(column_indexes)
        if missing:
            raise ValueError(f"輸入工作表缺少欄位：{', '.join(sorted(missing))}")

        segment_metadata: dict[tuple[str, int], tuple[int, int | None]] = {}
        points: dict[tuple[str, int], SegmentPoint] = {}
        source_files: set[str] = set()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            source_raw = row[column_indexes[SOURCE_COLUMN]]
            source_file = str(source_raw).strip() if source_raw is not None else ""
            segment_index = to_int(row[column_indexes[SEGMENT_COLUMN]])
            start_second = to_int(row[column_indexes[START_SECOND_COLUMN]])
            end_second = to_int(row[column_indexes[END_SECOND_COLUMN]])
            row_relative_second = to_int(row[column_indexes[OFFSET_COLUMN]])
            absolute_second = to_int(row[column_indexes[ABSOLUTE_SECOND_COLUMN]])
            alpha_value = to_finite_float(row[column_indexes[ALPHA_COLUMN]])
            eye_value = to_finite_float(row[column_indexes[EYE_COLUMN]])

            if not source_file or segment_index is None or start_second is None:
                continue
            key = (source_file, segment_index)
            source_files.add(source_file)
            segment_metadata[key] = (start_second, end_second)

            if row_relative_second != relative_second:
                continue
            if absolute_second is None or alpha_value is None or eye_value is None:
                continue

            points[key] = SegmentPoint(
                source_file=source_file,
                segment_index=segment_index,
                start_second=start_second,
                end_second=end_second,
                relative_second=row_relative_second,
                absolute_second=absolute_second,
                alpha_value=alpha_value,
                eye_value=eye_value,
            )
    finally:
        workbook.close()

    selected_points = [points[key] for key in sorted(points)]
    skipped_rows = [
        [
            source_file,
            segment_index,
            start_second,
            end_second,
            relative_second,
            f"缺少相對秒數 {relative_second} 的 Alpha 或眼動資料",
        ]
        for (source_file, segment_index), (start_second, end_second) in sorted(segment_metadata.items())
        if (source_file, segment_index) not in points
    ]
    return selected_points, skipped_rows, source_files


def save_distribution_plot(
    values: Iterable[float],
    *,
    title: str,
    x_label: str,
    output_path: Path,
    included_segments: int,
    relative_second: int,
    y_max: float | None = None,
) -> None:
    """Save a one-dimensional histogram with mean and median reference lines."""
    numeric_values = list(values)
    if not numeric_values:
        raise ValueError("沒有可繪製的資料。")

    lower = min(numeric_values)
    upper = max(numeric_values)
    if lower == upper:
        bins = np.array([lower - 0.5, lower + 0.5])
    else:
        bins = np.arange(np.floor(lower) - 0.5, np.ceil(upper) + 1.5, 1.0)

    plt.rcParams["font.sans-serif"] = [
        "Microsoft JhengHei",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False

    figure, axis = plt.subplots(figsize=(11, 7))
    axis.hist(numeric_values, bins=bins, color="#7db7d5", edgecolor="#3b82a0", alpha=0.9)
    axis.axvline(
        statistics.mean(numeric_values),
        color="#d1495b",
        linewidth=2,
        label=f"平均值 = {statistics.mean(numeric_values):.2f}",
    )
    axis.axvline(
        statistics.median(numeric_values),
        color="#54478c",
        linewidth=2,
        linestyle="--",
        label=f"中位數 = {statistics.median(numeric_values):.2f}",
    )
    axis.set_title(
        f"{title}\n疲勞區段開始前 {abs(relative_second)} 秒；"
        f"有效區段數 = {included_segments}；9 份資料合併"
    )
    axis.set_xlabel(x_label)
    axis.set_ylabel("疲勞區段數")
    if y_max is not None:
        axis.set_ylim(top=float(y_max))
    axis.grid(axis="y", linestyle="--", alpha=0.35)
    axis.legend()
    figure.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(figure)


def point_rows(points: Iterable[SegmentPoint]) -> list[list[object]]:
    """Return one Excel row for each fatigue segment's selected point."""
    return [
        [
            point.source_file,
            point.segment_index,
            point.start_second,
            point.end_second,
            point.relative_second,
            point.absolute_second,
            point.alpha_value,
            point.eye_value,
        ]
        for point in points
    ]


def statistics_rows(points: Iterable[SegmentPoint]) -> list[list[object]]:
    """Calculate Alpha and eye descriptive statistics for the selected second."""
    point_list = list(points)
    alpha_values = [point.alpha_value for point in point_list]
    eye_values = [point.eye_value for point in point_list]
    return [
        [
            point_list[0].relative_second,
            len(point_list),
            statistics.mean(alpha_values),
            statistics.median(alpha_values),
            min(alpha_values),
            max(alpha_values),
            len(point_list),
            statistics.mean(eye_values),
            statistics.median(eye_values),
            min(eye_values),
            max(eye_values),
        ]
    ]


def file_statistics_rows(
    points: Iterable[SegmentPoint],
    source_files: set[str],
) -> list[list[object]]:
    """Return the count of selected points per source workbook."""
    counts = Counter(point.source_file for point in points)
    return [[source_file, counts[source_file]] for source_file in sorted(source_files)]


def style_worksheet(worksheet) -> None:
    """Apply basic formatting to an output worksheet."""
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for column_cells in worksheet.columns:
        width = max(len(str(cell.value if cell.value is not None else "")) for cell in column_cells) + 2
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width, 40)


def append_rows(worksheet, headers: list[str], rows: Iterable[Iterable[object]]) -> None:
    """Append headers and rows, then apply worksheet formatting."""
    worksheet.append(headers)
    for row in rows:
        worksheet.append(list(row))
    style_worksheet(worksheet)


def write_summary_workbook(
    output_path: Path,
    *,
    points: list[SegmentPoint],
    skipped_rows: list[list[object]],
    source_files: set[str],
    input_path: Path,
    input_sheet: str,
    relative_second: int,
) -> None:
    """Write selected points, statistics, file counts, and definitions to XLSX."""
    workbook = Workbook()
    second_label = abs(relative_second)

    point_sheet = workbook.active
    point_sheet.title = f"第{second_label}秒逐區段"
    append_rows(
        point_sheet,
        [
            SOURCE_COLUMN,
            SEGMENT_COLUMN,
            START_SECOND_COLUMN,
            END_SECOND_COLUMN,
            OFFSET_COLUMN,
            ABSOLUTE_SECOND_COLUMN,
            ALPHA_COLUMN,
            EYE_COLUMN,
        ],
        point_rows(points),
    )

    statistics_sheet = workbook.create_sheet(f"第{second_label}秒統計")
    append_rows(
        statistics_sheet,
        [
            OFFSET_COLUMN,
            "Alpha樣本數",
            "Alpha平均",
            "Alpha中位數",
            "Alpha最小",
            "Alpha最大",
            "眼動樣本數",
            "眼動平均",
            "眼動中位數",
            "眼動最小",
            "眼動最大",
        ],
        statistics_rows(points),
    )

    file_sheet = workbook.create_sheet("檔案統計")
    append_rows(
        file_sheet,
        [SOURCE_COLUMN, f"納入第{second_label}秒區段數"],
        file_statistics_rows(points, source_files),
    )

    skipped_sheet = workbook.create_sheet("未納入區段")
    append_rows(
        skipped_sheet,
        [
            SOURCE_COLUMN,
            SEGMENT_COLUMN,
            START_SECOND_COLUMN,
            END_SECOND_COLUMN,
            OFFSET_COLUMN,
            "原因",
        ],
        skipped_rows,
    )

    notes_sheet = workbook.create_sheet("說明")
    notes = [
        ("輸入 Excel", str(input_path.resolve())),
        ("輸入工作表", input_sheet),
        ("擷取時間點", f"相對秒數 {relative_second}，即疲勞區段開始前 {abs(relative_second)} 秒。"),
        ("納入條件", "該疲勞區段在指定相對秒數必須同時有 Alpha 與眼動值。"),
        ("分布圖", "將 9 份資料的所有納入疲勞區段，分別繪製 Alpha 與眼動值直方圖。"),
    ]
    append_rows(notes_sheet, ["項目", "定義"], notes)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="從疲勞區段前30秒 Excel 擷取指定單秒的 Alpha／眼動資料並繪製分布。"
    )
    parser.add_argument(
        "--input-xlsx",
        type=Path,
        default=(
            script_dir
            / "data"
            / "fatigue_segment_pre30_output"
            / "fatigue_segment_pre30_summary.xlsx"
        ),
        help="含有前30秒逐區段工作表的疲勞區段 Excel。",
    )
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_SHEET_NAME,
        help=f"要讀取的工作表名稱（預設 {DEFAULT_SHEET_NAME}）。",
    )
    parser.add_argument(
        "--relative-second",
        type=int,
        default=DEFAULT_RELATIVE_SECOND,
        help="要擷取的相對秒數；預設 -15，即疲勞區段開始前第15秒。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=script_dir / "data" / "fatigue_segment_pre15_output",
        help="輸出 XLSX 與 PNG 的資料夾；預設為 data/fatigue_segment_pre15_output。",
    )
    parser.add_argument(
        "--xlsx-name",
        help="輸出 Excel 檔名；預設會依相對秒數命名，例如 -15 為 fatigue_segment_pre15_summary.xlsx。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.relative_second > 0:
        raise ValueError("relative-second 必須小於或等於 0。")

    input_path = args.input_xlsx.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"找不到輸入 Excel：{input_path}")

    output_dir = args.output_dir.resolve()
    points, skipped_rows, source_files = read_segment_points(
        input_path,
        sheet_name=args.sheet_name,
        relative_second=int(args.relative_second),
    )
    if not points:
        raise ValueError("找不到指定相對秒數的完整 Alpha／眼動資料。")

    second_label = abs(int(args.relative_second))
    xlsx_name = args.xlsx_name or f"fatigue_segment_pre{second_label}_summary.xlsx"
    workbook_output = output_dir / xlsx_name
    eye_output = output_dir / f"fatigue_segment_pre{second_label}_eye_distribution.png"
    alpha_output = output_dir / f"fatigue_segment_pre{second_label}_alpha_distribution.png"
    write_summary_workbook(
        workbook_output,
        points=points,
        skipped_rows=skipped_rows,
        source_files=source_files,
        input_path=input_path,
        input_sheet=args.sheet_name,
        relative_second=int(args.relative_second),
    )
    save_distribution_plot(
        (point.eye_value for point in points),
        title=f"疲勞區段開始前第{second_label}秒的眼動分布",
        x_label=EYE_COLUMN,
        output_path=eye_output,
        included_segments=len(points),
        relative_second=int(args.relative_second),
        y_max=30.0,
    )
    save_distribution_plot(
        (point.alpha_value for point in points),
        title=f"疲勞區段開始前第{second_label}秒的 Alpha 分布",
        x_label=ALPHA_COLUMN,
        output_path=alpha_output,
        included_segments=len(points),
        relative_second=int(args.relative_second),
    )

    print(f"輸入 Excel：{input_path}")
    print(f"擷取相對秒數：{args.relative_second}")
    print(f"納入疲勞區段：{len(points)}")
    print(f"未納入區段：{len(skipped_rows)}")
    print(f"Excel：{workbook_output}")
    print(f"眼動圖：{eye_output}")
    print(f"Alpha 圖：{alpha_output}")


if __name__ == "__main__":
    main()
