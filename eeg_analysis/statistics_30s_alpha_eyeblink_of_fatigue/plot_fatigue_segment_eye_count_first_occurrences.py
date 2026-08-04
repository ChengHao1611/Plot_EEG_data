"""Plot when 5--15 eye-movement counts first occur before fatigue starts.

For each complete fatigue-segment window, this module examines relative seconds
``-29`` through ``0`` in the ``眼動次數（30秒滑動）`` column.  For a target count
such as 5, multiple matching seconds in the same window are reduced to the
earliest one (the smallest relative second).  Segments with no matching second
are deliberately omitted from that target count's data.
"""

from __future__ import annotations

import argparse
import math
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:  # Supports both ``python -m`` and running this file directly.
    from .plot_fatigue_segment_pre30_distributions import (
        DEFAULT_RECOVERY_DURATION_SECONDS,
        DEFAULT_RECOVERY_THRESHOLD,
        DEFAULT_START_THRESHOLD,
        EYE_COLUMN,
        SegmentWindow,
        WorkbookData,
        build_segment_window,
        find_fatigue_segments,
        read_workbook,
    )
except ImportError:  # pragma: no cover - exercised only for direct execution.
    from plot_fatigue_segment_pre30_distributions import (
        DEFAULT_RECOVERY_DURATION_SECONDS,
        DEFAULT_RECOVERY_THRESHOLD,
        DEFAULT_START_THRESHOLD,
        EYE_COLUMN,
        SegmentWindow,
        WorkbookData,
        build_segment_window,
        find_fatigue_segments,
        read_workbook,
    )


DEFAULT_LOOKBACK_SECONDS = 29
DEFAULT_TARGET_COUNTS = tuple(range(5, 16))
DEFAULT_XLSX_NAME = "fatigue_segment_eye_count_first_occurrences.xlsx"


@dataclass(frozen=True)
class FirstOccurrence:
    """The earliest relative second at which one eye-count target occurred."""

    source_file: str
    segment_index: int
    start_second: int
    end_second: int | None
    target_count: int
    relative_second: int
    absolute_second: int
    matching_relative_seconds: tuple[int, ...]


@dataclass(frozen=True)
class ExcludedSegment:
    """A fatigue segment whose required pre-start eye window was unavailable."""

    source_file: str
    segment_index: int
    start_second: int
    end_second: int | None
    reason: str


def is_target_count(value: float | None, target_count: int) -> bool:
    """Return whether a numeric eye-count cell exactly represents ``target_count``."""
    return value is not None and math.isclose(
        float(value), float(target_count), rel_tol=0.0, abs_tol=1e-9
    )


def select_first_occurrences(
    windows: Iterable[SegmentWindow],
    *,
    target_counts: Sequence[int],
    lookback_seconds: int,
) -> list[FirstOccurrence]:
    """Select the earliest matching second for every target in every window.

    A non-occurrence intentionally yields no ``FirstOccurrence`` object.  This
    enforces the requested rule that it is not represented by a zero, blank, or
    any other replacement value.
    """
    occurrences: list[FirstOccurrence] = []
    offsets = range(-lookback_seconds, 1)

    for window in windows:
        segment = window.segment
        for target_count in target_counts:
            matches = tuple(
                offset
                for offset in offsets
                if is_target_count(window.values[offset][1], target_count)
            )
            if not matches:
                continue

            first_offset = matches[0]
            occurrences.append(
                FirstOccurrence(
                    source_file=segment.source_file,
                    segment_index=segment.segment_index,
                    start_second=segment.start_event.second,
                    end_second=segment.end_second,
                    target_count=target_count,
                    relative_second=first_offset,
                    absolute_second=segment.start_event.second + first_offset,
                    matching_relative_seconds=matches,
                )
            )

    return occurrences


def has_complete_eye_window(window: SegmentWindow, *, lookback_seconds: int) -> bool:
    """Return whether a closed segment has every required eye-count value.

    This analysis is intentionally independent of Alpha values: an absent
    Alpha value must not discard a segment whose eye-movement data is complete.
    """
    return window.segment.is_closed and all(
        window.values[offset][1] is not None
        for offset in range(-lookback_seconds, 1)
    )


def collect_complete_windows(
    workbook_data: Iterable[WorkbookData],
    *,
    start_threshold: float,
    recovery_threshold: float,
    recovery_duration_seconds: int,
    lookback_seconds: int,
) -> tuple[list[SegmentWindow], list[ExcludedSegment], int]:
    """Find fatigue segments and retain closed ones with complete eye windows."""
    complete_windows: list[SegmentWindow] = []
    excluded_segments: list[ExcludedSegment] = []
    total_segments = 0
    window_seconds = lookback_seconds + 1

    for data in workbook_data:
        segments = find_fatigue_segments(
            data,
            start_threshold=start_threshold,
            recovery_threshold=recovery_threshold,
            recovery_duration_seconds=recovery_duration_seconds,
        )
        total_segments += len(segments)

        for segment in segments:
            window = build_segment_window(
                data, segment, window_seconds=window_seconds
            )
            if has_complete_eye_window(window, lookback_seconds=lookback_seconds):
                complete_windows.append(window)
                continue

            if not segment.is_closed:
                reason = "疲勞區段未完成恢復判定"
            else:
                reason = "開始前視窗缺少眼動逐秒資料"
            excluded_segments.append(
                ExcludedSegment(
                    source_file=segment.source_file,
                    segment_index=segment.segment_index,
                    start_second=segment.start_event.second,
                    end_second=segment.end_second,
                    reason=reason,
                )
            )

    return complete_windows, excluded_segments, total_segments


def save_distribution_plot(
    occurrences: Sequence[FirstOccurrence],
    *,
    target_count: int,
    lookback_seconds: int,
    included_segments: int,
    output_path: Path,
) -> None:
    """Save the requested relative-second count distribution for one target."""
    offsets = list(range(-lookback_seconds, 1))
    distribution = Counter(item.relative_second for item in occurrences)
    frequencies = [distribution[offset] for offset in offsets]

    plt.rcParams["font.sans-serif"] = [
        "Microsoft JhengHei",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False

    figure, axis = plt.subplots(figsize=(14, 7))
    axis.bar(offsets, frequencies, width=0.85, color="#4c78a8", edgecolor="white")
    axis.axvline(0, color="#555555", linestyle="--", linewidth=1, label="疲勞區段開始")
    axis.set_title(
        f"第 {target_count} 次眼動最早出現時間分佈\n"
        f"範圍：疲勞區段開始前 {lookback_seconds} 至 0 秒；"
        f"有效紀錄 = {len(occurrences)}／{included_segments} 區段"
    )
    axis.set_xlabel("相對於疲勞區段開始的時間（秒）")
    axis.set_ylabel("次數（疲勞區段數）")
    axis.set_xticks(offsets)
    axis.set_xlim(-lookback_seconds - 0.6, 0.6)
    axis.set_ylim(bottom=0)
    axis.grid(axis="y", linestyle="--", alpha=0.35)
    axis.legend(loc="upper right")
    figure.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(figure)


def save_valid_record_count_distribution_plot(
    occurrences: Iterable[FirstOccurrence],
    *,
    target_counts: Sequence[int],
    included_segments: int,
    output_path: Path,
) -> None:
    """Save one overview chart of valid-record counts for targets 5 through 15."""
    valid_record_counts = Counter(item.target_count for item in occurrences)
    frequencies = [valid_record_counts[target_count] for target_count in target_counts]

    plt.rcParams["font.sans-serif"] = [
        "Microsoft JhengHei",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False

    figure, axis = plt.subplots(figsize=(11, 7))
    bars = axis.bar(
        target_counts,
        frequencies,
        width=0.72,
        color="#59a14f",
        edgecolor="white",
    )
    for bar, frequency in zip(bars, frequencies):
        axis.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            str(frequency),
            ha="center",
            va="bottom",
            fontsize=10,
        )

    axis.set_title(
        "第 5～15 次眼動的有效紀錄分佈\n"
        f"有效紀錄：在疲勞區段開始前視窗內至少出現一次；可納入區段數 = {included_segments}"
    )
    axis.set_xlabel("眼動次數目標")
    axis.set_ylabel("有效紀錄數（疲勞區段數）")
    axis.set_xticks(target_counts)
    axis.set_ylim(bottom=0)
    axis.grid(axis="y", linestyle="--", alpha=0.35)
    figure.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(figure)


def append_rows(worksheet, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    """Append a bold header, rows, and readable column widths to a worksheet."""
    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(list(row))

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)
    worksheet.freeze_panes = "A2"


def occurrence_rows(occurrences: Iterable[FirstOccurrence]) -> list[list[object]]:
    """Build audit rows containing only segments where the target occurred."""
    return [
        [
            item.source_file,
            item.segment_index,
            item.start_second,
            item.end_second,
            item.target_count,
            item.relative_second,
            item.absolute_second,
            ", ".join(str(offset) for offset in item.matching_relative_seconds),
        ]
        for item in occurrences
    ]


def distribution_rows(
    occurrences: Iterable[FirstOccurrence],
    *,
    target_counts: Sequence[int],
    lookback_seconds: int,
) -> list[list[object]]:
    """Build one frequency table row per target-count / relative-second pair."""
    distributions: dict[int, Counter[int]] = {
        count: Counter() for count in target_counts
    }
    for item in occurrences:
        distributions[item.target_count][item.relative_second] += 1

    return [
        [target_count, offset, distributions[target_count][offset]]
        for target_count in target_counts
        for offset in range(-lookback_seconds, 1)
    ]


def write_summary_workbook(
    output_path: Path,
    *,
    occurrences: Sequence[FirstOccurrence],
    excluded_segments: Sequence[ExcludedSegment],
    target_counts: Sequence[int],
    lookback_seconds: int,
    total_segments: int,
    included_segments: int,
    start_threshold: float,
    recovery_threshold: float,
    recovery_duration_seconds: int,
) -> None:
    """Write audit details, frequency data, and analysis settings to XLSX."""
    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "最早出現明細"
    append_rows(
        detail_sheet,
        [
            "來源檔案",
            "疲勞區段編號",
            "區段開始秒數",
            "區段結束秒數",
            "眼動次數目標",
            "最早相對秒數",
            "最早絕對秒數",
            "此區段所有符合秒數",
        ],
        occurrence_rows(occurrences),
    )

    distribution_sheet = workbook.create_sheet("分佈統計")
    append_rows(
        distribution_sheet,
        ["眼動次數目標", "相對秒數", "最早出現的疲勞區段數"],
        distribution_rows(
            occurrences,
            target_counts=target_counts,
            lookback_seconds=lookback_seconds,
        ),
    )

    included_by_target = Counter(item.target_count for item in occurrences)
    target_summary_sheet = workbook.create_sheet("目標次數統計")
    append_rows(
        target_summary_sheet,
        ["眼動次數目標", "有最早出現紀錄的區段數", "未出現而未紀錄的區段數"],
        [
            [
                target_count,
                included_by_target[target_count],
                included_segments - included_by_target[target_count],
            ]
            for target_count in target_counts
        ],
    )

    excluded_sheet = workbook.create_sheet("未納入視窗")
    append_rows(
        excluded_sheet,
        ["來源檔案", "疲勞區段編號", "區段開始秒數", "區段結束秒數", "未納入原因"],
        [
            [
                item.source_file,
                item.segment_index,
                item.start_second,
                item.end_second,
                item.reason,
            ]
            for item in excluded_segments
        ],
    )

    notes_sheet = workbook.create_sheet("說明")
    append_rows(
        notes_sheet,
        ["項目", "內容"],
        [
            ("目標眼動次數", ", ".join(str(value) for value in target_counts)),
            (
                "分析範圍",
                f"以疲勞區段開始為 0 秒，收集 -{lookback_seconds} 至 0 秒（含 0 秒）。",
            ),
            (
                "最早出現規則",
                "同一區段同一目標次數若出現在多個秒數，只取最小的相對秒數。",
            ),
            (
                "未出現處理",
                "目標次數在該區段視窗未出現時，不建立最早出現明細，也不計入圖表長條。",
            ),
            ("疲勞區段開始", f"第一個反應時間 >= {start_threshold:g} 秒的 event。"),
            (
                "疲勞區段結束",
                f"恢復反應時間需 < {recovery_threshold:g} 秒，並持續 {recovery_duration_seconds} 秒。",
            ),
            ("偵測到的疲勞區段數", total_segments),
            ("納入完整視窗的疲勞區段數", included_segments),
        ],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="統計疲勞區段開始前，眼動次數 5 至 15 的最早出現秒數分佈。"
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=script_dir / "data",
        help="包含逐秒統計 XLSX 的資料夾。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=script_dir / "data" / "fatigue_segment_eye_count_first_occurrences_output",
        help="輸出 XLSX 與 11 張 PNG 的資料夾。",
    )
    parser.add_argument(
        "--lookback-seconds",
        type=int,
        default=DEFAULT_LOOKBACK_SECONDS,
        help="往疲勞區段開始前回看的秒數（預設 29，範圍為 -29 至 0）。",
    )
    parser.add_argument(
        "--target-counts",
        type=int,
        nargs="+",
        default=DEFAULT_TARGET_COUNTS,
        help="要繪製的眼動次數目標（預設為 5 至 15）。",
    )
    parser.add_argument(
        "--start-threshold",
        type=float,
        default=DEFAULT_START_THRESHOLD,
        help="反應時間達此秒數以上時開始疲勞區段（預設 1.6）。",
    )
    parser.add_argument(
        "--recovery-threshold",
        type=float,
        default=DEFAULT_RECOVERY_THRESHOLD,
        help="反應時間必須嚴格低於此秒數才算恢復（預設 1.6）。",
    )
    parser.add_argument(
        "--recovery-duration-seconds",
        type=int,
        default=DEFAULT_RECOVERY_DURATION_SECONDS,
        help="恢復開始後需連續維持的秒數（預設 60）。",
    )
    parser.add_argument(
        "--xlsx-name",
        default=DEFAULT_XLSX_NAME,
        help=f"輸出 Excel 檔名（預設 {DEFAULT_XLSX_NAME}）。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.lookback_seconds < 0:
        raise ValueError("lookback-seconds 不可小於 0。")
    if args.recovery_duration_seconds <= 0:
        raise ValueError("recovery-duration-seconds 必須大於 0。")

    target_counts = tuple(sorted(set(int(value) for value in args.target_counts)))
    if not target_counts or target_counts[0] < 0:
        raise ValueError("target-counts 至少須包含一個非負整數。")

    input_dir = args.input_dir.resolve()
    output_dir = args.output_dir.resolve()
    workbook_output = output_dir / args.xlsx_name
    workbook_paths = [
        path
        for path in sorted(input_dir.glob("*.xlsx"))
        if not path.name.startswith("~$") and path.resolve() != workbook_output.resolve()
    ]
    if not workbook_paths:
        raise FileNotFoundError(f"找不到 XLSX 檔案：{input_dir}")

    workbook_data: list[WorkbookData] = []
    skipped_files: list[tuple[str, str]] = []
    for workbook_path in workbook_paths:
        try:
            workbook_data.append(read_workbook(workbook_path))
        except ValueError as error:
            skipped_files.append((workbook_path.name, str(error)))
    if not workbook_data:
        raise ValueError("找不到含有疲勞區段與眼動欄位的 XLSX 檔案。")

    complete_windows, excluded_segments, total_segments = collect_complete_windows(
        workbook_data,
        start_threshold=float(args.start_threshold),
        recovery_threshold=float(args.recovery_threshold),
        recovery_duration_seconds=int(args.recovery_duration_seconds),
        lookback_seconds=int(args.lookback_seconds),
    )
    if not complete_windows:
        raise ValueError("找不到正常結束且具有完整前置視窗的疲勞區段。")

    occurrences = select_first_occurrences(
        complete_windows,
        target_counts=target_counts,
        lookback_seconds=int(args.lookback_seconds),
    )
    write_summary_workbook(
        workbook_output,
        occurrences=occurrences,
        excluded_segments=excluded_segments,
        target_counts=target_counts,
        lookback_seconds=int(args.lookback_seconds),
        total_segments=total_segments,
        included_segments=len(complete_windows),
        start_threshold=float(args.start_threshold),
        recovery_threshold=float(args.recovery_threshold),
        recovery_duration_seconds=int(args.recovery_duration_seconds),
    )

    for target_count in target_counts:
        target_occurrences = [
            item for item in occurrences if item.target_count == target_count
        ]
        output_path = output_dir / f"eye_count_{target_count:02d}_first_occurrence_distribution.png"
        save_distribution_plot(
            target_occurrences,
            target_count=target_count,
            lookback_seconds=int(args.lookback_seconds),
            included_segments=len(complete_windows),
            output_path=output_path,
        )

    valid_record_output = output_dir / "eye_count_05_to_15_valid_record_distribution.png"
    save_valid_record_count_distribution_plot(
        occurrences,
        target_counts=target_counts,
        included_segments=len(complete_windows),
        output_path=valid_record_output,
    )

    print(f"讀取有效 XLSX：{len(workbook_data)} 份")
    print(f"偵測到的疲勞區段數：{total_segments}")
    print(f"納入完整前置視窗的區段數：{len(complete_windows)}")
    for target_count in target_counts:
        count = sum(item.target_count == target_count for item in occurrences)
        print(f"第 {target_count} 次眼動最早出現紀錄：{count}")
    if skipped_files:
        print("略過不相容 XLSX：")
        for file_name, reason in skipped_files:
            print(f"- {file_name}: {reason}")
    print(f"Excel：{workbook_output}")
    print(f"有效紀錄分佈圖：{valid_record_output}")
    print(f"圖表資料夾：{output_dir}")


if __name__ == "__main__":
    main()
