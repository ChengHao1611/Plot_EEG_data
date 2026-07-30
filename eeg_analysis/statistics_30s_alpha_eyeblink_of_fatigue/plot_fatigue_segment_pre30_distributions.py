"""Plot 30-second Alpha and eye windows before each fatigue segment.

A fatigue segment starts at the first event with reaction time >= 1.6 seconds.
It ends only after three consecutive reaction-time events <= 1.5 seconds.
Every complete segment contributes the inclusive 30-second window from its
start second - 29 through its start second.
"""

from __future__ import annotations

import argparse
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SECOND_COLUMN = "秒數"
REACTION_TIME_COLUMN = "事件反應時間"
ALPHA_COLUMN = "α波次數（10秒滑動）"
EYE_COLUMN = "眼動次數（30秒滑動）"
DEFAULT_START_THRESHOLD = 1.6
DEFAULT_RECOVERY_THRESHOLD = 1.5
DEFAULT_RECOVERY_EVENT_COUNT = 3
DEFAULT_WINDOW_SECONDS = 30
DEFAULT_XLSX_NAME = "fatigue_segment_pre30_summary.xlsx"


@dataclass(frozen=True)
class ReactionEvent:
    """One reaction-time event, numbered within its source workbook."""

    index: int
    second: int
    reaction_time: float


@dataclass(frozen=True)
class WorkbookData:
    """The per-second values and reaction-time events from one input workbook."""

    name: str
    per_second: dict[int, tuple[float | None, float | None]]
    events: list[ReactionEvent]


@dataclass(frozen=True)
class FatigueSegment:
    """A fatigue segment, optionally unclosed when the recording ends first."""

    source_file: str
    segment_index: int
    start_event: ReactionEvent
    end_event: ReactionEvent | None
    recovery_event_count: int

    @property
    def is_closed(self) -> bool:
        return self.end_event is not None


@dataclass(frozen=True)
class SegmentWindow:
    """One segment's 30-second pre-start window and its completeness state."""

    segment: FatigueSegment
    values: dict[int, tuple[float | None, float | None]]
    missing_seconds: list[int]

    @property
    def is_complete(self) -> bool:
        return self.segment.is_closed and not self.missing_seconds


def to_float(value: object) -> float | None:
    """Return a finite numeric worksheet value, or None for blanks/non-numbers."""
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if np.isfinite(number) else None


def read_workbook(path: Path) -> WorkbookData:
    """Read one Chinese-column statistics workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            raise ValueError("缺少標題列")

        column_indexes = {
            str(name): index for index, name in enumerate(header) if name is not None
        }
        required_columns = {
            SECOND_COLUMN,
            REACTION_TIME_COLUMN,
            ALPHA_COLUMN,
            EYE_COLUMN,
        }
        missing_columns = required_columns - set(column_indexes)
        if missing_columns:
            raise ValueError(f"缺少欄位：{', '.join(sorted(missing_columns))}")

        second_index = column_indexes[SECOND_COLUMN]
        reaction_index = column_indexes[REACTION_TIME_COLUMN]
        alpha_index = column_indexes[ALPHA_COLUMN]
        eye_index = column_indexes[EYE_COLUMN]
        per_second: dict[int, list[float | None]] = {}
        events: list[ReactionEvent] = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            second_value = to_float(row[second_index] if len(row) > second_index else None)
            if second_value is None:
                continue
            second = int(second_value)

            alpha_value = to_float(row[alpha_index] if len(row) > alpha_index else None)
            eye_value = to_float(row[eye_index] if len(row) > eye_index else None)
            stored_values = per_second.setdefault(second, [None, None])
            if alpha_value is not None:
                stored_values[0] = alpha_value
            if eye_value is not None:
                stored_values[1] = eye_value

            reaction_time = to_float(
                row[reaction_index] if len(row) > reaction_index else None
            )
            if reaction_time is not None:
                events.append(
                    ReactionEvent(
                        index=len(events) + 1,
                        second=second,
                        reaction_time=reaction_time,
                    )
                )
    finally:
        workbook.close()

    return WorkbookData(
        name=path.name,
        per_second={second: (values[0], values[1]) for second, values in per_second.items()},
        events=events,
    )


def find_fatigue_segments(
    data: WorkbookData,
    *,
    start_threshold: float,
    recovery_threshold: float,
    recovery_event_count: int,
) -> list[FatigueSegment]:
    """Detect segments from one high reaction through three consecutive recoveries."""
    segments: list[FatigueSegment] = []
    active_start: ReactionEvent | None = None
    recovery_streak = 0

    for event in data.events:
        if active_start is None:
            if event.reaction_time >= start_threshold:
                active_start = event
                recovery_streak = 0
            continue

        if event.reaction_time <= recovery_threshold:
            recovery_streak += 1
            if recovery_streak == recovery_event_count:
                segments.append(
                    FatigueSegment(
                        source_file=data.name,
                        segment_index=len(segments) + 1,
                        start_event=active_start,
                        end_event=event,
                        recovery_event_count=recovery_streak,
                    )
                )
                active_start = None
                recovery_streak = 0
        else:
            recovery_streak = 0

    if active_start is not None:
        segments.append(
            FatigueSegment(
                source_file=data.name,
                segment_index=len(segments) + 1,
                start_event=active_start,
                end_event=None,
                recovery_event_count=recovery_streak,
            )
        )

    return segments


def build_segment_window(
    data: WorkbookData,
    segment: FatigueSegment,
    *,
    window_seconds: int,
) -> SegmentWindow:
    """Collect the inclusive [start - window + 1, start] Alpha/eye window."""
    values: dict[int, tuple[float | None, float | None]] = {}
    missing_seconds: list[int] = []
    for offset in range(-(window_seconds - 1), 1):
        absolute_second = segment.start_event.second + offset
        alpha_value, eye_value = data.per_second.get(absolute_second, (None, None))
        values[offset] = (alpha_value, eye_value)
        if alpha_value is None or eye_value is None:
            missing_seconds.append(absolute_second)

    return SegmentWindow(
        segment=segment,
        values=values,
        missing_seconds=missing_seconds,
    )


def collect_complete_windows(
    workbook_data: Iterable[WorkbookData],
    *,
    start_threshold: float,
    recovery_threshold: float,
    recovery_event_count: int,
    window_seconds: int,
) -> tuple[list[FatigueSegment], list[SegmentWindow], dict[str, list[FatigueSegment]]]:
    """Find all segments and retain the ones with complete pre-start windows."""
    all_segments: list[FatigueSegment] = []
    windows: list[SegmentWindow] = []
    segments_by_file: dict[str, list[FatigueSegment]] = {}

    for data in workbook_data:
        segments = find_fatigue_segments(
            data,
            start_threshold=start_threshold,
            recovery_threshold=recovery_threshold,
            recovery_event_count=recovery_event_count,
        )
        segments_by_file[data.name] = segments
        all_segments.extend(segments)
        for segment in segments:
            windows.append(build_segment_window(data, segment, window_seconds=window_seconds))

    complete_windows = [window for window in windows if window.is_complete]
    return all_segments, complete_windows, segments_by_file


def collect_values_by_offset(
    windows: Iterable[SegmentWindow],
    *,
    window_seconds: int,
) -> tuple[dict[int, list[float]], dict[int, list[float]]]:
    """Pool Alpha and eye values by each relative second across complete windows."""
    alpha_by_offset: dict[int, list[float]] = defaultdict(list)
    eye_by_offset: dict[int, list[float]] = defaultdict(list)
    offsets = range(-(window_seconds - 1), 1)

    for window in windows:
        for offset in offsets:
            alpha_value, eye_value = window.values[offset]
            if alpha_value is None or eye_value is None:
                raise ValueError("完整 window 不應包含空的 Alpha 或眼動值。")
            alpha_by_offset[offset].append(alpha_value)
            eye_by_offset[offset].append(eye_value)

    return alpha_by_offset, eye_by_offset


def save_distribution_plot(
    values_by_offset: dict[int, list[float]],
    *,
    title: str,
    y_label: str,
    output_path: Path,
    included_segments: int,
    start_threshold: float,
    recovery_threshold: float,
    recovery_event_count: int,
    window_seconds: int,
) -> None:
    """Save one boxplot-and-mean distribution over the full pre-start window."""
    relative_seconds = list(range(-(window_seconds - 1), 1))
    distributions = [values_by_offset[offset] for offset in relative_seconds]
    means = [float(np.mean(values)) for values in distributions]

    plt.rcParams["font.sans-serif"] = [
        "Microsoft JhengHei",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False

    figure, axis = plt.subplots(figsize=(16, 7))
    boxplot = axis.boxplot(
        distributions,
        positions=relative_seconds,
        widths=0.58,
        patch_artist=True,
        showfliers=False,
        medianprops={"color": "#202020", "linewidth": 1.4},
        whiskerprops={"color": "#52616b"},
        capprops={"color": "#52616b"},
    )
    for box in boxplot["boxes"]:
        box.set(facecolor="#a7d8f0", edgecolor="#3b82a0", alpha=0.85)

    axis.plot(relative_seconds, means, color="#d1495b", marker="o", linewidth=2, label="平均值")
    axis.axvline(0, color="#555555", linestyle="--", linewidth=1, label="疲勞區段開始")
    axis.set_xticks(relative_seconds)
    axis.set_xlabel("相對於疲勞區段開始的時間（秒）")
    axis.set_ylabel(y_label)
    axis.set_title(
        f"{title}\n開始反應時間 >= {start_threshold:.1f} 秒；"
        f"連續 {recovery_event_count} 個反應時間 <= {recovery_threshold:.1f} 秒結束；"
        f"有效區段數 = {included_segments}"
    )
    axis.grid(axis="y", linestyle="--", alpha=0.35)
    axis.legend(loc="upper left")
    figure.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(figure)


def format_second_list(seconds: Iterable[int]) -> str:
    """Format missing window seconds compactly for an Excel cell."""
    ordered = list(seconds)
    return ", ".join(str(second) for second in ordered) if ordered else ""


def segment_detail_rows(
    segments: Iterable[FatigueSegment],
    complete_windows: Iterable[SegmentWindow],
) -> list[list[object]]:
    """Create the per-segment Excel detail rows."""
    window_by_key = {
        (window.segment.source_file, window.segment.segment_index): window
        for window in complete_windows
    }
    rows: list[list[object]] = []
    for segment in segments:
        end_event = segment.end_event
        window = window_by_key.get((segment.source_file, segment.segment_index))
        rows.append(
            [
                segment.source_file,
                segment.segment_index,
                segment.start_event.index,
                segment.start_event.second,
                segment.start_event.reaction_time,
                end_event.index if end_event else None,
                end_event.second if end_event else None,
                end_event.reaction_time if end_event else None,
                end_event.second - segment.start_event.second if end_event else None,
                segment.recovery_event_count,
                "連續恢復 event 達標" if end_event else "檔案結束前未達連續恢復條件",
                "是" if window is not None else "否",
                "" if window is not None else "未關閉或前30秒缺少 Alpha／眼動值",
            ]
        )
    return rows


def window_detail_rows(
    windows: Iterable[SegmentWindow],
    *,
    window_seconds: int,
) -> list[list[object]]:
    """Create one Excel row per relative second for every complete segment window."""
    rows: list[list[object]] = []
    offsets = range(-(window_seconds - 1), 1)
    for window in windows:
        segment = window.segment
        for offset in offsets:
            alpha_value, eye_value = window.values[offset]
            rows.append(
                [
                    segment.source_file,
                    segment.segment_index,
                    segment.start_event.second,
                    segment.end_event.second if segment.end_event else None,
                    offset,
                    segment.start_event.second + offset,
                    alpha_value,
                    eye_value,
                ]
            )
    return rows


def offset_summary_rows(
    alpha_by_offset: dict[int, list[float]],
    eye_by_offset: dict[int, list[float]],
    *,
    window_seconds: int,
) -> list[list[object]]:
    """Create per-relative-second Alpha and eye summary statistics."""
    rows: list[list[object]] = []
    for offset in range(-(window_seconds - 1), 1):
        alpha_values = alpha_by_offset[offset]
        eye_values = eye_by_offset[offset]
        rows.append(
            [
                offset,
                len(alpha_values),
                statistics.mean(alpha_values),
                statistics.median(alpha_values),
                min(alpha_values),
                max(alpha_values),
                len(eye_values),
                statistics.mean(eye_values),
                statistics.median(eye_values),
                min(eye_values),
                max(eye_values),
            ]
        )
    return rows


def file_summary_rows(
    workbook_data: Iterable[WorkbookData],
    segments_by_file: dict[str, list[FatigueSegment]],
    complete_windows: Iterable[SegmentWindow],
) -> list[list[object]]:
    """Create one summary row per source workbook."""
    completed_by_file: Counter[str] = Counter(
        window.segment.source_file for window in complete_windows
    )
    rows: list[list[object]] = []
    for data in sorted(workbook_data, key=lambda item: item.name):
        segments = segments_by_file[data.name]
        rows.append(
            [
                data.name,
                len(data.events),
                len(segments),
                sum(segment.is_closed for segment in segments),
                sum(not segment.is_closed for segment in segments),
                completed_by_file[data.name],
            ]
        )
    return rows


def style_worksheet(worksheet) -> None:
    """Apply basic formatting to an output worksheet."""
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for column_cells in worksheet.columns:
        width = max(len(str(cell.value if cell.value is not None else "")) for cell in column_cells) + 2
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width, 40)


def append_rows(worksheet, headers: list[str], rows: Iterable[Iterable[object]]) -> None:
    """Write headers and data rows, then style the worksheet."""
    worksheet.append(headers)
    for row in rows:
        worksheet.append(list(row))
    style_worksheet(worksheet)


def write_summary_workbook(
    output_path: Path,
    *,
    segments: list[FatigueSegment],
    complete_windows: list[SegmentWindow],
    alpha_by_offset: dict[int, list[float]],
    eye_by_offset: dict[int, list[float]],
    workbook_data: list[WorkbookData],
    segments_by_file: dict[str, list[FatigueSegment]],
    start_threshold: float,
    recovery_threshold: float,
    recovery_event_count: int,
    window_seconds: int,
) -> None:
    """Write all requested fatigue-segment details and 30-second summaries."""
    workbook = Workbook()

    segment_sheet = workbook.active
    segment_sheet.title = "疲勞區段明細"
    append_rows(
        segment_sheet,
        [
            "來源檔案",
            "疲勞區段編號",
            "開始event序號",
            "開始秒數",
            "開始反應時間_秒",
            "結束event序號",
            "結束秒數",
            "結束反應時間_秒",
            "區段持續秒數",
            "結束時連續恢復event數",
            "結束狀態",
            "前30秒window完整",
            "window未納入原因",
        ],
        segment_detail_rows(segments, complete_windows),
    )

    window_sheet = workbook.create_sheet("前30秒逐區段")
    append_rows(
        window_sheet,
        [
            "來源檔案",
            "疲勞區段編號",
            "區段開始秒數",
            "區段結束秒數",
            "相對秒數",
            "絕對秒數",
            "α波次數（10秒滑動）",
            "眼動次數（30秒滑動）",
        ],
        window_detail_rows(complete_windows, window_seconds=window_seconds),
    )

    summary_sheet = workbook.create_sheet("前30秒統計")
    append_rows(
        summary_sheet,
        [
            "相對秒數",
            "Alpha樣本數",
            "Alpha平均",
            "Alpha中位數",
            "Alpha最小",
            "Alpha最大",
            "眼動樣本數",
            "眼動平均",
            "眼動中位數",
            "眼動最小",
            "眼動最大",
        ],
        offset_summary_rows(
            alpha_by_offset,
            eye_by_offset,
            window_seconds=window_seconds,
        ),
    )

    file_sheet = workbook.create_sheet("檔案統計")
    append_rows(
        file_sheet,
        [
            "來源檔案",
            "event總數",
            "疲勞區段數",
            "正常結束區段數",
            "未關閉區段數",
            "納入前30秒window區段數",
        ],
        file_summary_rows(workbook_data, segments_by_file, complete_windows),
    )

    notes_sheet = workbook.create_sheet("說明")
    notes = [
        ("疲勞區段開始", f"第一個反應時間 >= {start_threshold:g} 秒的 event。"),
        (
            "疲勞區段結束",
            f"開始後出現連續 {recovery_event_count} 個反應時間 <= {recovery_threshold:g} 秒的 event；第 {recovery_event_count} 個 event 為結束點。",
        ),
        (
            "恢復計數重設",
            f"疲勞區段中只要 event 反應時間 > {recovery_threshold:g} 秒，連續恢復計數便歸零。",
        ),
        (
            "前30秒window",
            f"以區段開始秒數為 0，收集 {-(window_seconds - 1)} 至 0 秒，共 {window_seconds} 秒；區段開始秒包含在內。",
        ),
        ("納入圖表條件", "區段正常結束，且前30秒每一秒都同時有 Alpha 與眼動數值。"),
    ]
    append_rows(notes_sheet, ["項目", "定義"], notes)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="依疲勞區段繪製開始前30秒的 Alpha 與眼動分布，並輸出 XLSX。"
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=script_dir / "data",
        help="包含中文欄位逐秒統計 XLSX 的資料夾。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=script_dir / "data" / "fatigue_segment_pre30_output",
        help="輸出 XLSX 與 PNG 的資料夾；預設為 data/fatigue_segment_pre30_output。",
    )
    parser.add_argument(
        "--start-threshold",
        type=float,
        default=DEFAULT_START_THRESHOLD,
        help="反應時間達此秒數以上時開始疲勞區段（預設 1.6）。",
    )
    parser.add_argument(
        "--recovery-threshold",
        type=float,
        default=DEFAULT_RECOVERY_THRESHOLD,
        help="反應時間低於或等於此秒數時計入恢復連續次數（預設 1.5）。",
    )
    parser.add_argument(
        "--recovery-event-count",
        type=int,
        default=DEFAULT_RECOVERY_EVENT_COUNT,
        help="結束疲勞區段所需的連續恢復 event 數（預設 3）。",
    )
    parser.add_argument(
        "--window-seconds",
        type=int,
        default=DEFAULT_WINDOW_SECONDS,
        help="區段開始前含開始秒的 window 長度（預設 30）。",
    )
    parser.add_argument(
        "--xlsx-name",
        default=DEFAULT_XLSX_NAME,
        help=f"輸出 Excel 檔名（預設 {DEFAULT_XLSX_NAME}）。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.recovery_event_count <= 0:
        raise ValueError("recovery-event-count 必須大於 0。")
    if args.window_seconds <= 0:
        raise ValueError("window-seconds 必須大於 0。")

    input_dir = args.input_dir.resolve()
    output_dir = args.output_dir.resolve()
    workbook_output = output_dir / args.xlsx_name
    workbook_paths = [
        path
        for path in sorted(input_dir.glob("*.xlsx"))
        if not path.name.startswith("~$") and path.resolve() != workbook_output.resolve()
    ]
    if not workbook_paths:
        raise FileNotFoundError(f"找不到 XLSX 檔案：{input_dir}")

    workbook_data: list[WorkbookData] = []
    skipped_files: list[tuple[str, str]] = []
    for workbook_path in workbook_paths:
        try:
            workbook_data.append(read_workbook(workbook_path))
        except ValueError as error:
            skipped_files.append((workbook_path.name, str(error)))

    if not workbook_data:
        raise ValueError("找不到含有疲勞區段所需欄位的 XLSX 檔案。")

    segments, complete_windows, segments_by_file = collect_complete_windows(
        workbook_data,
        start_threshold=float(args.start_threshold),
        recovery_threshold=float(args.recovery_threshold),
        recovery_event_count=int(args.recovery_event_count),
        window_seconds=int(args.window_seconds),
    )
    if not complete_windows:
        raise ValueError("找不到正常結束且具有完整前30秒 Alpha／眼動資料的疲勞區段。")

    alpha_by_offset, eye_by_offset = collect_values_by_offset(
        complete_windows,
        window_seconds=int(args.window_seconds),
    )
    write_summary_workbook(
        workbook_output,
        segments=segments,
        complete_windows=complete_windows,
        alpha_by_offset=alpha_by_offset,
        eye_by_offset=eye_by_offset,
        workbook_data=workbook_data,
        segments_by_file=segments_by_file,
        start_threshold=float(args.start_threshold),
        recovery_threshold=float(args.recovery_threshold),
        recovery_event_count=int(args.recovery_event_count),
        window_seconds=int(args.window_seconds),
    )

    eye_output = output_dir / "fatigue_segment_pre30_eye_distribution.png"
    alpha_output = output_dir / "fatigue_segment_pre30_alpha_distribution.png"
    save_distribution_plot(
        eye_by_offset,
        title="疲勞區段開始前30秒的眼動分布",
        y_label="眼動次數（30秒滑動）",
        output_path=eye_output,
        included_segments=len(complete_windows),
        start_threshold=float(args.start_threshold),
        recovery_threshold=float(args.recovery_threshold),
        recovery_event_count=int(args.recovery_event_count),
        window_seconds=int(args.window_seconds),
    )
    save_distribution_plot(
        alpha_by_offset,
        title="疲勞區段開始前30秒的 Alpha 分布",
        y_label="α波次數（10秒滑動）",
        output_path=alpha_output,
        included_segments=len(complete_windows),
        start_threshold=float(args.start_threshold),
        recovery_threshold=float(args.recovery_threshold),
        recovery_event_count=int(args.recovery_event_count),
        window_seconds=int(args.window_seconds),
    )

    print(f"讀取有效 XLSX：{len(workbook_data)} 份")
    print(f"疲勞區段數：{len(segments)}")
    print(f"納入完整前30秒 window 的區段數：{len(complete_windows)}")
    if skipped_files:
        print("略過不相容 XLSX：")
        for file_name, reason in skipped_files:
            print(f"- {file_name}: {reason}")
    print(f"Excel：{workbook_output}")
    print(f"眼動圖：{eye_output}")
    print(f"Alpha 圖：{alpha_output}")


if __name__ == "__main__":
    main()
