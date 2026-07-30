"""Find fatigue intervals and the event gaps between adjacent intervals.

A fatigue interval begins at the first event with reaction time >= 1.6 seconds
and ends at the first later event with reaction time < 1.5 seconds.  The
interval is half-open: [start, end).  The requested gap between two adjacent
fatigue intervals is:

    next_interval.start_event_index - previous_interval.end_event_index

For example, an interval ending at event 14 and the next one starting at event
20 has an event gap of 6.
"""

from __future__ import annotations

import argparse
import statistics
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SECOND_COLUMN = "秒數"
REACTION_TIME_COLUMN = "事件反應時間"
DEFAULT_START_THRESHOLD = 1.6
DEFAULT_END_THRESHOLD = 1.5
DEFAULT_XLSX_NAME = "fatigue_interval_gap_summary.xlsx"
DEFAULT_PLOT_NAME = "fatigue_interval_gap_distribution.png"


@dataclass(frozen=True)
class Event:
    """One row with a valid reaction time, numbered within its source workbook."""

    source_file: str
    index: int
    second: float
    reaction_time: float


@dataclass(frozen=True)
class FatigueInterval:
    """One [start, end) fatigue interval; end is None when the file ends first."""

    source_file: str
    interval_index: int
    start_event: Event
    end_event: Event | None

    @property
    def is_closed(self) -> bool:
        return self.end_event is not None


@dataclass(frozen=True)
class IntervalGap:
    """The requested event gap between two adjacent fatigue intervals."""

    source_file: str
    previous_interval: FatigueInterval
    next_interval: FatigueInterval
    event_gap: int
    time_gap_seconds: float


def to_finite_float(value: object) -> float | None:
    """Convert a worksheet value to a finite float, or return None."""
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if np.isfinite(number) else None


def display_number(value: float | int | None) -> float | int | None:
    """Keep integer-looking values clean when they are written to Excel."""
    if value is None:
        return None
    value_float = float(value)
    return int(value_float) if value_float.is_integer() else value_float


def read_events(workbook_path: Path) -> list[Event]:
    """Read reaction-time events from one Chinese-column statistics workbook."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            raise ValueError("缺少標題列")

        column_indexes = {
            str(column_name): index
            for index, column_name in enumerate(header)
            if column_name is not None
        }
        required = {SECOND_COLUMN, REACTION_TIME_COLUMN}
        missing = required - set(column_indexes)
        if missing:
            raise ValueError(f"缺少欄位：{', '.join(sorted(missing))}")

        second_index = column_indexes[SECOND_COLUMN]
        reaction_index = column_indexes[REACTION_TIME_COLUMN]
        events: list[Event] = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            second = to_finite_float(row[second_index] if len(row) > second_index else None)
            reaction_time = to_finite_float(
                row[reaction_index] if len(row) > reaction_index else None
            )
            if second is None or reaction_time is None:
                continue

            events.append(
                Event(
                    source_file=workbook_path.name,
                    index=len(events) + 1,
                    second=second,
                    reaction_time=reaction_time,
                )
            )
    finally:
        workbook.close()

    return events


def find_fatigue_intervals(
    events: Iterable[Event],
    *,
    start_threshold: float,
    end_threshold: float,
) -> list[FatigueInterval]:
    """Build half-open fatigue intervals using the approved threshold rules."""
    intervals: list[FatigueInterval] = []
    active_start: Event | None = None
    source_file: str | None = None

    for event in events:
        source_file = event.source_file
        if active_start is None:
            if event.reaction_time >= start_threshold:
                active_start = event
            continue

        if event.reaction_time < end_threshold:
            intervals.append(
                FatigueInterval(
                    source_file=event.source_file,
                    interval_index=len(intervals) + 1,
                    start_event=active_start,
                    end_event=event,
                )
            )
            active_start = None

    if active_start is not None:
        intervals.append(
            FatigueInterval(
                source_file=source_file or active_start.source_file,
                interval_index=len(intervals) + 1,
                start_event=active_start,
                end_event=None,
            )
        )

    return intervals


def find_interval_gaps(intervals: Iterable[FatigueInterval]) -> list[IntervalGap]:
    """Calculate gaps between adjacent intervals in one source workbook."""
    ordered_intervals = list(intervals)
    gaps: list[IntervalGap] = []

    for previous, following in zip(ordered_intervals, ordered_intervals[1:]):
        if previous.end_event is None:
            continue

        gaps.append(
            IntervalGap(
                source_file=previous.source_file,
                previous_interval=previous,
                next_interval=following,
                event_gap=following.start_event.index - previous.end_event.index,
                time_gap_seconds=following.start_event.second - previous.end_event.second,
            )
        )

    return gaps


def interval_rows(intervals: Iterable[FatigueInterval]) -> list[list[object]]:
    """Return Excel-ready detail rows for all fatigue intervals."""
    rows: list[list[object]] = []
    for interval in intervals:
        end_event = interval.end_event
        rows.append(
            [
                interval.source_file,
                interval.interval_index,
                interval.start_event.index,
                display_number(interval.start_event.second),
                interval.start_event.reaction_time,
                end_event.index if end_event else None,
                display_number(end_event.second) if end_event else None,
                end_event.reaction_time if end_event else None,
                (
                    display_number(end_event.second - interval.start_event.second)
                    if end_event
                    else None
                ),
                "反應時間 < 結束門檻" if end_event else "檔案結束，區間未關閉",
            ]
        )
    return rows


def gap_rows(gaps: Iterable[IntervalGap]) -> list[list[object]]:
    """Return Excel-ready rows for the gaps between adjacent intervals."""
    rows: list[list[object]] = []
    for gap in gaps:
        previous = gap.previous_interval
        following = gap.next_interval
        previous_end = previous.end_event
        if previous_end is None:
            continue
        rows.append(
            [
                gap.source_file,
                previous.interval_index,
                previous.start_event.index,
                display_number(previous.start_event.second),
                previous_end.index,
                display_number(previous_end.second),
                previous_end.reaction_time,
                following.interval_index,
                following.start_event.index,
                display_number(following.start_event.second),
                following.start_event.reaction_time,
                gap.event_gap,
                display_number(gap.time_gap_seconds),
            ]
        )
    return rows


def summary_rows(
    events_by_file: dict[str, list[Event]],
    intervals_by_file: dict[str, list[FatigueInterval]],
    gaps_by_file: dict[str, list[IntervalGap]],
) -> list[list[object]]:
    """Return one summary row for every input workbook."""
    rows: list[list[object]] = []
    for source_file in sorted(events_by_file):
        intervals = intervals_by_file[source_file]
        gaps = gaps_by_file[source_file]
        gap_values = [gap.event_gap for gap in gaps]
        rows.append(
            [
                source_file,
                len(events_by_file[source_file]),
                len(intervals),
                sum(interval.is_closed for interval in intervals),
                sum(not interval.is_closed for interval in intervals),
                len(gaps),
                min(gap_values) if gap_values else None,
                statistics.mean(gap_values) if gap_values else None,
                statistics.median(gap_values) if gap_values else None,
                max(gap_values) if gap_values else None,
            ]
        )
    return rows


def distribution_rows(gaps: Iterable[IntervalGap]) -> list[list[object]]:
    """Build the exact discrete distribution of event-gap values."""
    values = [gap.event_gap for gap in gaps]
    counts = Counter(values)
    total = len(values)
    return [
        [event_gap, count, count / total if total else 0.0]
        for event_gap, count in sorted(counts.items())
    ]


def style_worksheet(worksheet) -> None:
    """Apply lightweight, readable formatting to one output worksheet."""
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for column_cells in worksheet.columns:
        width = max(len(str(cell.value if cell.value is not None else "")) for cell in column_cells) + 2
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width, 36)


def add_rows(worksheet, headers: list[str], rows: Iterable[Iterable[object]]) -> None:
    """Append a header and row sequence, then apply shared worksheet styling."""
    worksheet.append(headers)
    for row in rows:
        worksheet.append(list(row))
    style_worksheet(worksheet)


def write_summary_workbook(
    output_path: Path,
    *,
    interval_detail_rows: list[list[object]],
    interval_gap_rows: list[list[object]],
    file_summary_rows: list[list[object]],
    distribution: list[list[object]],
    start_threshold: float,
    end_threshold: float,
) -> None:
    """Write interval details, gaps, summaries, distribution, and definitions."""
    workbook = Workbook()

    intervals_sheet = workbook.active
    intervals_sheet.title = "疲勞區間明細"
    add_rows(
        intervals_sheet,
        [
            "來源檔案",
            "疲勞區間編號",
            "開始event序號",
            "開始秒數",
            "開始反應時間_秒",
            "結束event序號",
            "結束秒數",
            "結束反應時間_秒",
            "區間持續秒數",
            "結束狀態",
        ],
        interval_detail_rows,
    )

    gaps_sheet = workbook.create_sheet("區間間隔")
    add_rows(
        gaps_sheet,
        [
            "來源檔案",
            "前疲勞區間編號",
            "前區間開始event序號",
            "前區間開始秒數",
            "前區間結束event序號",
            "前區間結束秒數",
            "前區間結束反應時間_秒",
            "下個疲勞區間編號",
            "下個區間開始event序號",
            "下個區間開始秒數",
            "下個區間開始反應時間_秒",
            "相差event數",
            "間隔秒數",
        ],
        interval_gap_rows,
    )

    summary_sheet = workbook.create_sheet("檔案統計")
    add_rows(
        summary_sheet,
        [
            "來源檔案",
            "event總數",
            "疲勞區間數",
            "正常結束區間數",
            "未關閉區間數",
            "區間間隔數",
            "最小相差event數",
            "平均相差event數",
            "中位數相差event數",
            "最大相差event數",
        ],
        file_summary_rows,
    )

    distribution_sheet = workbook.create_sheet("event差分佈")
    add_rows(
        distribution_sheet,
        ["相差event數", "疲勞區間間隔數", "百分比"],
        distribution,
    )
    for cell in distribution_sheet["C"][1:]:
        cell.number_format = "0.00%"

    notes_sheet = workbook.create_sheet("說明")
    notes = [
        ("開始門檻", f"反應時間 >= {start_threshold:g} 秒時開啟疲勞區間。"),
        ("結束門檻", f"反應時間 < {end_threshold:g} 秒時結束疲勞區間。"),
        ("區間關係", "疲勞區間採 [開始, 結束)；結束 event 不屬於前一個疲勞區間。"),
        (
            "相差event數",
            "下一疲勞區間開始 event 序號 − 前疲勞區間結束 event 序號。"
            "例如前區間於第 14 個 event 結束、下區間於第 20 個 event 開始，結果為 6。",
        ),
        ("未關閉區間", "若檔案結尾前沒有出現反應時間 < 結束門檻，區間標示為未關閉，且不會有下一區間間隔。"),
    ]
    add_rows(notes_sheet, ["項目", "定義"], notes)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def save_distribution_plot(gaps: Iterable[IntervalGap], output_path: Path) -> None:
    """Save a bar-chart distribution of the event gaps between intervals."""
    values = [gap.event_gap for gap in gaps]
    counts = Counter(values)

    plt.rcParams["font.sans-serif"] = ["Microsoft JhengHei", "Microsoft YaHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    figure, axis = plt.subplots(figsize=(13, 7))
    if counts:
        event_gaps = sorted(counts)
        frequencies = [counts[event_gap] for event_gap in event_gaps]
        axis.bar(event_gaps, frequencies, width=0.8, color="#2B6CB0", edgecolor="white")
        axis.set_xticks(event_gaps)

        mean_value = statistics.mean(values)
        median_value = statistics.median(values)
        axis.axvline(mean_value, color="#D1495B", linewidth=2, label=f"平均值 = {mean_value:.2f}")
        axis.axvline(
            median_value,
            color="#54478C",
            linewidth=2,
            linestyle="--",
            label=f"中位數 = {median_value:.2f}",
        )
        axis.legend()
    else:
        axis.text(
            0.5,
            0.5,
            "沒有可計算的相鄰疲勞區間間隔",
            transform=axis.transAxes,
            ha="center",
            va="center",
            fontsize=13,
        )

    axis.set_title("相鄰疲勞區間相差 event 數的分布")
    axis.set_xlabel("相差 event 數")
    axis.set_ylabel("疲勞區間間隔數")
    axis.grid(axis="y", alpha=0.25)
    axis.set_axisbelow(True)
    figure.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(figure)


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="判斷疲勞區間，統計相鄰疲勞區間相差的 event 數，並輸出 XLSX 與分布圖。"
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=script_dir / "data",
        help="包含中文欄位逐秒統計 XLSX 的資料夾。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=script_dir / "data",
        help="輸出 XLSX 與 PNG 的資料夾；預設與輸入資料夾相同。",
    )
    parser.add_argument(
        "--start-threshold",
        type=float,
        default=DEFAULT_START_THRESHOLD,
        help="反應時間達此秒數以上時開啟疲勞區間（預設 1.6）。",
    )
    parser.add_argument(
        "--end-threshold",
        type=float,
        default=DEFAULT_END_THRESHOLD,
        help="反應時間低於此秒數時結束疲勞區間（預設 1.5）。",
    )
    parser.add_argument(
        "--xlsx-name",
        default=DEFAULT_XLSX_NAME,
        help=f"輸出 Excel 檔名（預設 {DEFAULT_XLSX_NAME}）。",
    )
    parser.add_argument(
        "--plot-name",
        default=DEFAULT_PLOT_NAME,
        help=f"輸出 PNG 檔名（預設 {DEFAULT_PLOT_NAME}）。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.end_threshold >= args.start_threshold:
        raise ValueError("end-threshold 必須小於 start-threshold，才能保留 1.5 至 1.6 的緩衝區間。")

    input_dir = args.input_dir.resolve()
    output_dir = args.output_dir.resolve()
    workbook_output = output_dir / args.xlsx_name
    plot_output = output_dir / args.plot_name

    workbook_paths = [
        path
        for path in sorted(input_dir.glob("*.xlsx"))
        if not path.name.startswith("~$") and path.resolve() != workbook_output.resolve()
    ]
    if not workbook_paths:
        raise FileNotFoundError(f"找不到輸入 XLSX：{input_dir}")

    events_by_file: dict[str, list[Event]] = {}
    intervals_by_file: dict[str, list[FatigueInterval]] = {}
    gaps_by_file: dict[str, list[IntervalGap]] = {}

    for workbook_path in workbook_paths:
        events = read_events(workbook_path)
        intervals = find_fatigue_intervals(
            events,
            start_threshold=float(args.start_threshold),
            end_threshold=float(args.end_threshold),
        )
        gaps = find_interval_gaps(intervals)
        events_by_file[workbook_path.name] = events
        intervals_by_file[workbook_path.name] = intervals
        gaps_by_file[workbook_path.name] = gaps

    all_intervals = [
        interval
        for source_file in sorted(intervals_by_file)
        for interval in intervals_by_file[source_file]
    ]
    all_gaps = [
        gap
        for source_file in sorted(gaps_by_file)
        for gap in gaps_by_file[source_file]
    ]
    detail_rows = interval_rows(all_intervals)
    gap_detail_rows = gap_rows(all_gaps)
    summary = summary_rows(events_by_file, intervals_by_file, gaps_by_file)
    distribution = distribution_rows(all_gaps)

    write_summary_workbook(
        workbook_output,
        interval_detail_rows=detail_rows,
        interval_gap_rows=gap_detail_rows,
        file_summary_rows=summary,
        distribution=distribution,
        start_threshold=float(args.start_threshold),
        end_threshold=float(args.end_threshold),
    )
    save_distribution_plot(all_gaps, plot_output)

    print(f"讀取檔案：{len(workbook_paths)} 份")
    print(f"event 總數：{sum(len(events) for events in events_by_file.values())}")
    print(f"疲勞區間數：{len(all_intervals)}")
    print(f"相鄰疲勞區間間隔數：{len(all_gaps)}")
    print(f"Excel：{workbook_output}")
    print(f"分布圖：{plot_output}")


if __name__ == "__main__":
    main()
