"""Summarise manually annotated awake periods with non-overlapping eye windows.

``data/awake/awake.xlsx`` is the annotation source.  It must contain the
columns ``來源檔案``, ``清醒開始時間``, and ``清醒結束時間``.  Both endpoints are
inclusive.  Each row's end time is retained as the point 60 seconds before the
following fatigue onset; that 60-second buffer is never included in an awake
window.

The input eye feature is already a trailing 30-second count.  Therefore this
script samples it only at the end of adjacent, non-overlapping 30-second
windows that lie wholly within an annotated awake period.  It deliberately
does not pool every overlapping second as independent observations.
"""

from __future__ import annotations

import argparse
import math
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SOURCE_COLUMN = "來源檔案"
AWAKE_START_COLUMN = "清醒開始時間"
AWAKE_END_COLUMN = "清醒結束時間"
EYE_COLUMN = "眼動次數（30秒滑動）"
WINDOW_SECONDS = 30
FATIGUE_BUFFER_SECONDS = 60
DEFAULT_XLSX_NAME = "awake_30s_eye_statistics.xlsx"


@dataclass(frozen=True)
class AwakeInterval:
    """One inclusive manually annotated awake interval."""

    source_file: str
    interval_index: int
    start_second: int
    end_second: int

    @property
    def fatigue_start_second(self) -> int:
        """Return the fatigue onset implied by the supplied 60-second buffer."""
        return self.end_second + FATIGUE_BUFFER_SECONDS

    @property
    def duration_seconds(self) -> int:
        return self.end_second - self.start_second + 1


@dataclass(frozen=True)
class AwakeWindow:
    """One complete, non-overlapping 30-second awake eye-count window."""

    source_file: str
    interval_index: int
    interval_start_second: int
    interval_end_second: int
    implied_fatigue_start_second: int
    window_index: int
    start_second: int
    end_second: int
    eye_count: float


@dataclass(frozen=True)
class ExcludedWindow:
    """A planned awake window that cannot be used due to a data-quality issue."""

    source_file: str
    interval_index: int
    start_second: int
    end_second: int
    reason: str


def finite_number(value: object) -> float | None:
    """Return a finite float, or ``None`` for blank/non-numeric cells."""
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def integer_second(value: object, *, field_name: str, row_number: int) -> int:
    """Read an integral, non-negative annotation time with a helpful error."""
    number = finite_number(value)
    if number is None or not number.is_integer() or number < 0:
        raise ValueError(
            f"awake.xlsx 第 {row_number} 列的「{field_name}」必須是非負整數秒。"
        )
    return int(number)


def read_awake_intervals(path: Path) -> list[AwakeInterval]:
    """Load and validate manual awake annotations from the active worksheet."""
    if not path.is_file():
        raise FileNotFoundError(f"找不到清醒標註檔：{path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            raise ValueError("awake.xlsx 缺少標題列。")
        indexes = {
            str(value).strip(): index
            for index, value in enumerate(header)
            if value is not None
        }
        required = {SOURCE_COLUMN, AWAKE_START_COLUMN, AWAKE_END_COLUMN}
        missing = required - set(indexes)
        if missing:
            raise ValueError(f"awake.xlsx 缺少欄位：{', '.join(sorted(missing))}")

        intervals: list[AwakeInterval] = []
        source_counts: Counter[str] = Counter()
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not row or not any(value is not None for value in row):
                continue
            source_value = row[indexes[SOURCE_COLUMN]] if len(row) > indexes[SOURCE_COLUMN] else None
            if source_value is None or not str(source_value).strip():
                raise ValueError(f"awake.xlsx 第 {row_number} 列缺少「{SOURCE_COLUMN}」。")
            source_file = Path(str(source_value).strip()).name
            start_second = integer_second(
                row[indexes[AWAKE_START_COLUMN]]
                if len(row) > indexes[AWAKE_START_COLUMN]
                else None,
                field_name=AWAKE_START_COLUMN,
                row_number=row_number,
            )
            end_second = integer_second(
                row[indexes[AWAKE_END_COLUMN]]
                if len(row) > indexes[AWAKE_END_COLUMN]
                else None,
                field_name=AWAKE_END_COLUMN,
                row_number=row_number,
            )
            if end_second < start_second:
                raise ValueError(
                    f"awake.xlsx 第 {row_number} 列的結束時間不可早於開始時間。"
                )
            source_counts[source_file] += 1
            intervals.append(
                AwakeInterval(
                    source_file=source_file,
                    interval_index=source_counts[source_file],
                    start_second=start_second,
                    end_second=end_second,
                )
            )
    finally:
        workbook.close()

    if not intervals:
        raise ValueError("awake.xlsx 沒有任何清醒標註資料。")
    validate_non_overlapping_intervals(intervals)
    return intervals


def validate_non_overlapping_intervals(intervals: Iterable[AwakeInterval]) -> None:
    """Reject overlapping annotations that would otherwise duplicate eye windows."""
    by_source: dict[str, list[AwakeInterval]] = defaultdict(list)
    for interval in intervals:
        by_source[interval.source_file].append(interval)

    for source_file, source_intervals in by_source.items():
        ordered = sorted(source_intervals, key=lambda item: item.start_second)
        for previous, current in zip(ordered, ordered[1:]):
            if current.start_second <= previous.end_second:
                raise ValueError(
                    f"{source_file} 的清醒標註區段 {previous.interval_index} 與 "
                    f"{current.interval_index} 重疊。"
                )


def read_eye_values(path: Path) -> dict[int, float | None]:
    """Read the per-second trailing 30-second eye count from one raw workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            raise ValueError(f"{path.name} 缺少標題列。")
        indexes = {
            str(value).strip(): index
            for index, value in enumerate(header)
            if value is not None
        }
        if "秒數" not in indexes or EYE_COLUMN not in indexes:
            raise ValueError(f"{path.name} 缺少「秒數」或「{EYE_COLUMN}」欄位。")

        values: dict[int, float | None] = {}
        second_index = indexes["秒數"]
        eye_index = indexes[EYE_COLUMN]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            second_value = row[second_index] if len(row) > second_index else None
            second_number = finite_number(second_value)
            if second_number is None or not second_number.is_integer():
                continue
            second = int(second_number)
            if second < 0:
                continue
            eye_value = row[eye_index] if len(row) > eye_index else None
            values[second] = finite_number(eye_value)
    finally:
        workbook.close()
    return values


def planned_windows(interval: AwakeInterval) -> Iterable[tuple[int, int, int]]:
    """Yield ``(index, start, end)`` for complete 30-second windows only."""
    window_index = 1
    start_second = interval.start_second
    while start_second + WINDOW_SECONDS - 1 <= interval.end_second:
        end_second = start_second + WINDOW_SECONDS - 1
        yield window_index, start_second, end_second
        window_index += 1
        start_second += WINDOW_SECONDS


def collect_windows(
    interval: AwakeInterval, eye_values: dict[int, float | None]
) -> tuple[list[AwakeWindow], list[ExcludedWindow]]:
    """Collect usable windows and explicitly record incomplete source data."""
    included: list[AwakeWindow] = []
    excluded: list[ExcludedWindow] = []

    for window_index, start_second, end_second in planned_windows(interval):
        missing_seconds = [
            second
            for second in range(start_second, end_second + 1)
            if eye_values.get(second) is None
        ]
        if missing_seconds:
            excluded.append(
                ExcludedWindow(
                    source_file=interval.source_file,
                    interval_index=interval.interval_index,
                    start_second=start_second,
                    end_second=end_second,
                    reason=(
                        "視窗內缺少眼動逐秒資料："
                        f"{len(missing_seconds)} 秒（首筆 {missing_seconds[0]} 秒）"
                    ),
                )
            )
            continue

        # Eye_Sum at the endpoint represents exactly [start_second, end_second].
        included.append(
            AwakeWindow(
                source_file=interval.source_file,
                interval_index=interval.interval_index,
                interval_start_second=interval.start_second,
                interval_end_second=interval.end_second,
                implied_fatigue_start_second=interval.fatigue_start_second,
                window_index=window_index,
                start_second=start_second,
                end_second=end_second,
                eye_count=eye_values[end_second],  # checked above as non-missing
            )
        )

    return included, excluded


def observed_quantile(values: Sequence[float], probability: float) -> float | None:
    """Return a nearest-rank observed value without interpolating count data."""
    if not values:
        return None
    ordered = sorted(values)
    index = max(0, min(len(ordered) - 1, math.ceil(probability * len(ordered)) - 1))
    return ordered[index]


def summarise(values: Sequence[float]) -> dict[str, float | int | None]:
    """Return count-friendly descriptive statistics for one value collection."""
    if not values:
        return {
            "n": 0,
            "mean": None,
            "sd": None,
            "min": None,
            "p5": None,
            "p25": None,
            "median": None,
            "p75": None,
            "p95": None,
            "max": None,
        }
    return {
        "n": len(values),
        "mean": statistics.mean(values),
        "sd": statistics.stdev(values) if len(values) >= 2 else None,
        "min": min(values),
        "p5": observed_quantile(values, 0.05),
        "p25": observed_quantile(values, 0.25),
        "median": observed_quantile(values, 0.50),
        "p75": observed_quantile(values, 0.75),
        "p95": observed_quantile(values, 0.95),
        "max": max(values),
    }


def weighted_quantile(
    value_weights: Sequence[tuple[float, float]], probability: float
) -> float | None:
    """Return a weighted observed quantile for equal-record weighting."""
    if not value_weights:
        return None
    ordered = sorted(value_weights, key=lambda item: item[0])
    total_weight = sum(weight for _, weight in ordered)
    threshold = total_weight * probability
    cumulative = 0.0
    for value, weight in ordered:
        cumulative += weight
        if cumulative >= threshold:
            return value
    return ordered[-1][0]


def summarise_equal_record_weighted(
    values_by_source: dict[str, list[float]]
) -> dict[str, float | int | None]:
    """Summarise all windows while giving each source workbook weight one."""
    value_weights = [
        (value, 1.0 / len(values))
        for values in values_by_source.values()
        if values
        for value in values
    ]
    if not value_weights:
        return {
            "records": 0,
            "windows": 0,
            "mean": None,
            "sd": None,
            "p5": None,
            "p25": None,
            "median": None,
            "p75": None,
            "p95": None,
        }
    total_weight = sum(weight for _, weight in value_weights)
    weighted_mean = sum(value * weight for value, weight in value_weights) / total_weight
    weighted_variance = (
        sum(weight * (value - weighted_mean) ** 2 for value, weight in value_weights)
        / total_weight
    )
    return {
        "records": sum(bool(values) for values in values_by_source.values()),
        "windows": len(value_weights),
        "mean": weighted_mean,
        "sd": math.sqrt(weighted_variance),
        "p5": weighted_quantile(value_weights, 0.05),
        "p25": weighted_quantile(value_weights, 0.25),
        "median": weighted_quantile(value_weights, 0.50),
        "p75": weighted_quantile(value_weights, 0.75),
        "p95": weighted_quantile(value_weights, 0.95),
    }


def set_plot_style() -> None:
    """Configure portable Chinese text rendering for the output figures."""
    plt.rcParams["font.sans-serif"] = [
        "Microsoft JhengHei",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False


def save_source_boxplot(values_by_source: dict[str, list[float]], output_path: Path) -> None:
    """Plot complete awake windows by source without implying equal window counts."""
    sources = [source for source in sorted(values_by_source) if values_by_source[source]]
    if not sources:
        return
    set_plot_style()
    figure, axis = plt.subplots(figsize=(max(10, len(sources) * 1.5), 7))
    groups = [values_by_source[source] for source in sources]
    axis.boxplot(groups, tick_labels=[Path(source).stem.replace("_raw", "") for source in sources])
    for index, values in enumerate(groups, start=1):
        # Deterministic offsets retain visibility without adding randomness.
        offsets = [index + ((position % 7) - 3) * 0.035 for position in range(len(values))]
        axis.scatter(offsets, values, color="#4c78a8", alpha=0.65, s=22, zorder=3)
    axis.set_title("人工標註清醒區段：各資料檔的 30 秒眼動次數")
    axis.set_xlabel("來源資料檔")
    axis.set_ylabel("眼動次數（30 秒、不重疊窗口）")
    axis.grid(axis="y", linestyle="--", alpha=0.35)
    figure.autofmt_xdate(rotation=25, ha="right")
    figure.tight_layout()
    figure.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(figure)


def save_equal_record_distribution(
    values_by_source: dict[str, list[float]], output_path: Path
) -> None:
    """Plot a probability mass and ECDF with each source given equal weight."""
    nonempty = {source: values for source, values in values_by_source.items() if values}
    if not nonempty:
        return
    set_plot_style()
    mass: defaultdict[float, float] = defaultdict(float)
    for values in nonempty.values():
        weight = 1.0 / len(nonempty) / len(values)
        for value in values:
            mass[value] += weight
    x_values = sorted(mass)
    probabilities = [mass[value] for value in x_values]
    cumulative: list[float] = []
    running = 0.0
    for probability in probabilities:
        running += probability
        cumulative.append(running)

    figure, (mass_axis, cdf_axis) = plt.subplots(2, 1, figsize=(11, 9), sharex=True)
    mass_axis.bar(x_values, probabilities, width=0.8, color="#59a14f", edgecolor="white")
    mass_axis.set_title("清醒眼動分布（每份資料檔等權重）")
    mass_axis.set_ylabel("加權比例")
    mass_axis.grid(axis="y", linestyle="--", alpha=0.35)
    cdf_axis.step(x_values, cumulative, where="post", color="#e15759", linewidth=2)
    cdf_axis.set_xlabel("眼動次數（30 秒、不重疊窗口）")
    cdf_axis.set_ylabel("累積比例")
    cdf_axis.set_ylim(0, 1.03)
    cdf_axis.grid(axis="both", linestyle="--", alpha=0.35)
    figure.tight_layout()
    figure.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(figure)


def append_rows(worksheet, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    """Append and lightly format a worksheet."""
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for column in worksheet.columns:
        width = max(len(str(cell.value if cell.value is not None else "")) for cell in column) + 2
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(width, 12), 42)


def stat_row(prefix: Sequence[object], summary: dict[str, float | int | None]) -> list[object]:
    """Attach common summary fields to a leading identifier sequence."""
    return list(prefix) + [
        summary["n"],
        summary["mean"],
        summary["sd"],
        summary["min"],
        summary["p5"],
        summary["p25"],
        summary["median"],
        summary["p75"],
        summary["p95"],
        summary["max"],
    ]


def write_output_workbook(
    output_path: Path,
    *,
    intervals: Sequence[AwakeInterval],
    windows: Sequence[AwakeWindow],
    excluded: Sequence[ExcludedWindow],
    all_source_files: Sequence[str],
) -> None:
    """Write detailed windows, summaries, coverage, and method documentation."""
    windows_by_interval: dict[tuple[str, int], list[float]] = defaultdict(list)
    windows_by_source: dict[str, list[float]] = defaultdict(list)
    for window in windows:
        windows_by_interval[(window.source_file, window.interval_index)].append(window.eye_count)
        windows_by_source[window.source_file].append(window.eye_count)

    interval_rows = []
    for interval in intervals:
        values = windows_by_interval[(interval.source_file, interval.interval_index)]
        planned_count = sum(1 for _ in planned_windows(interval))
        interval_rows.append(
            stat_row(
                [
                    interval.source_file,
                    interval.interval_index,
                    interval.start_second,
                    interval.end_second,
                    interval.duration_seconds,
                    interval.fatigue_start_second,
                    planned_count,
                ],
                summarise(values),
            )
        )

    source_intervals: dict[str, list[AwakeInterval]] = defaultdict(list)
    for interval in intervals:
        source_intervals[interval.source_file].append(interval)
    source_rows = []
    for source_file in sorted(source_intervals):
        source_rows.append(
            stat_row(
                [source_file, len(source_intervals[source_file])],
                summarise(windows_by_source[source_file]),
            )
        )

    weighted = summarise_equal_record_weighted(windows_by_source)
    record_medians = [
        summarise(values)["median"]
        for values in windows_by_source.values()
        if values
    ]
    record_median_summary = summarise([float(value) for value in record_medians if value is not None])

    annotation_seconds = Counter()
    planned_counts = Counter()
    included_counts = Counter()
    excluded_counts = Counter()
    for interval in intervals:
        annotation_seconds[interval.source_file] += interval.duration_seconds
        planned_counts[interval.source_file] += sum(1 for _ in planned_windows(interval))
    for window in windows:
        included_counts[window.source_file] += 1
    for item in excluded:
        excluded_counts[item.source_file] += 1

    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "清醒30秒窗口"
    append_rows(
        detail_sheet,
        [
            "來源檔案",
            "清醒區段編號",
            "人工清醒開始秒（含）",
            "人工清醒結束秒（含）",
            "疲勞開始秒（結束+60）",
            "窗口編號",
            "窗口開始秒（含）",
            "窗口結束秒（含）",
            "眼動次數（30秒滑動，取窗口結束秒）",
        ],
        [
            [
                item.source_file,
                item.interval_index,
                item.interval_start_second,
                item.interval_end_second,
                item.implied_fatigue_start_second,
                item.window_index,
                item.start_second,
                item.end_second,
                item.eye_count,
            ]
            for item in windows
        ],
    )

    summary_headers = [
        "來源檔案",
        "清醒區段編號",
        "開始秒（含）",
        "結束秒（含）",
        "標註秒數",
        "疲勞開始秒（結束+60）",
        "規劃完整30秒窗口數",
        "納入30秒窗口數",
        "眼動平均",
        "眼動標準差",
        "眼動最小",
        "眼動P5",
        "眼動P25",
        "眼動中位數",
        "眼動P75",
        "眼動P95",
        "眼動最大",
    ]
    interval_sheet = workbook.create_sheet("清醒區段統計")
    append_rows(interval_sheet, summary_headers, interval_rows)

    source_sheet = workbook.create_sheet("資料檔清醒統計")
    append_rows(
        source_sheet,
        ["來源檔案", "清醒區段數"] + summary_headers[7:],
        source_rows,
    )

    cohort_sheet = workbook.create_sheet("群體等權重統計")
    append_rows(
        cohort_sheet,
        [
            "統計單位",
            "有納入窗口的資料檔數",
            "完整30秒窗口數",
            "平均",
            "標準差",
            "P5",
            "P25",
            "中位數",
            "P75",
            "P95",
        ],
        [
            [
                "每份資料檔等權重的30秒窗口分布",
                weighted["records"],
                weighted["windows"],
                weighted["mean"],
                weighted["sd"],
                weighted["p5"],
                weighted["p25"],
                weighted["median"],
                weighted["p75"],
                weighted["p95"],
            ],
            [
                "資料檔中位數的分布（每檔一值）",
                record_median_summary["n"],
                record_median_summary["n"],
                record_median_summary["mean"],
                record_median_summary["sd"],
                record_median_summary["p5"],
                record_median_summary["p25"],
                record_median_summary["median"],
                record_median_summary["p75"],
                record_median_summary["p95"],
            ],
        ],
    )

    coverage_sheet = workbook.create_sheet("資料覆蓋與品質")
    coverage_rows = []
    for source_file in sorted(all_source_files):
        has_annotation = source_file in source_intervals
        status = (
            "納入清醒統計"
            if included_counts[source_file] > 0
            else "有清醒標註但沒有合格窗口"
            if has_annotation
            else "未提供清醒標註"
        )
        coverage_rows.append(
            [
                source_file,
                "是" if has_annotation else "否",
                len(source_intervals[source_file]),
                annotation_seconds[source_file],
                planned_counts[source_file],
                included_counts[source_file],
                excluded_counts[source_file],
                status,
            ]
        )
    for source_file in sorted(set(source_intervals) - set(all_source_files)):
        coverage_rows.append(
            [
                source_file,
                "是",
                len(source_intervals[source_file]),
                annotation_seconds[source_file],
                planned_counts[source_file],
                0,
                0,
                "標註來源檔不存在",
            ]
        )
    append_rows(
        coverage_sheet,
        [
            "來源檔案",
            "有人工清醒標註",
            "清醒區段數",
            "人工標註清醒秒數",
            "規劃完整30秒窗口數",
            "納入窗口數",
            "因資料不完整排除窗口數",
            "結果",
        ],
        coverage_rows,
    )

    excluded_sheet = workbook.create_sheet("排除窗口")
    append_rows(
        excluded_sheet,
        ["來源檔案", "清醒區段編號", "窗口開始秒", "窗口結束秒", "排除原因"],
        [
            [item.source_file, item.interval_index, item.start_second, item.end_second, item.reason]
            for item in excluded
        ],
    )

    method_sheet = workbook.create_sheet("方法說明")
    append_rows(
        method_sheet,
        ["項目", "內容"],
        [
            ["清醒標註來源", "data/awake/awake.xlsx；開始與結束秒皆包含。"],
            ["疲勞前保留時間", "每列結束秒為疲勞開始前60秒；該60秒不納入清醒統計。"],
            ["眼動特徵", "來源資料的「眼動次數（30秒滑動）」。"],
            ["30秒窗口", "從每段開始切成互不重疊完整30秒；不足30秒尾段捨棄。"],
            ["窗口眼動值", "取窗口結束秒的30秒滑動眼動值，代表整個窗口。"],
            ["資料品質", "窗口內任一秒眼動值缺失時，整個窗口排除並列於「排除窗口」。"],
            ["群體分布", "每份有合格窗口的資料檔總權重相同，避免清醒段長者主導結果。"],
            ["未標註資料", "未提供清醒標註的資料不納入清醒基準；僅列於資料覆蓋與品質。"],
        ],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def run_analysis(input_path: Path, data_dir: Path, output_dir: Path) -> dict[str, object]:
    """Execute the awake-eye analysis and return concise run metadata."""
    intervals = read_awake_intervals(input_path)
    all_source_files = sorted(path.name for path in data_dir.glob("*_raw.xlsx"))
    intervals_by_source: dict[str, list[AwakeInterval]] = defaultdict(list)
    for interval in intervals:
        intervals_by_source[interval.source_file].append(interval)

    windows: list[AwakeWindow] = []
    excluded: list[ExcludedWindow] = []
    missing_sources: list[str] = []
    for source_file in sorted(intervals_by_source):
        source_path = data_dir / source_file
        if not source_path.is_file():
            missing_sources.append(source_file)
            continue
        eye_values = read_eye_values(source_path)
        for interval in intervals_by_source[source_file]:
            included, rejected = collect_windows(interval, eye_values)
            windows.extend(included)
            excluded.extend(rejected)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = output_dir / DEFAULT_XLSX_NAME
    write_output_workbook(
        output_xlsx,
        intervals=intervals,
        windows=windows,
        excluded=excluded,
        all_source_files=all_source_files,
    )

    values_by_source: dict[str, list[float]] = defaultdict(list)
    for window in windows:
        values_by_source[window.source_file].append(window.eye_count)
    save_source_boxplot(values_by_source, output_dir / "awake_eye_by_source_boxplot.png")
    save_equal_record_distribution(
        values_by_source, output_dir / "awake_eye_equal_record_distribution.png"
    )

    return {
        "output_xlsx": output_xlsx,
        "window_count": len(windows),
        "excluded_count": len(excluded),
        "included_sources": len({item.source_file for item in windows}),
        "missing_sources": missing_sources,
    }


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    default_data_dir = script_dir / "data"
    parser = argparse.ArgumentParser(
        description="統計人工標註清醒區段中的不重疊30秒眼動次數。"
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=default_data_dir / "awake" / "awake.xlsx",
        help="人工清醒標註xlsx檔。",
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=default_data_dir,
        help="9份原始統計xlsx所在資料夾。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=default_data_dir / "awake" / "output",
        help="統計xlsx與圖表輸出資料夾。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = run_analysis(args.input, args.data_dir, args.output_dir)
    print(f"完成：{result['output_xlsx']}")
    print(
        "納入 "
        f"{result['window_count']} 個完整30秒窗口，"
        f"來自 {result['included_sources']} 份資料檔；"
        f"因資料不完整排除 {result['excluded_count']} 個窗口。"
    )
    if result["missing_sources"]:
        print("找不到標註來源檔：" + ", ".join(result["missing_sources"]))


if __name__ == "__main__":
    main()
