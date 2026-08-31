"""Function One: initial 300-second fatigue screening and RT baseline export."""

from __future__ import annotations

import argparse
import statistics
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

# When this file is launched directly, Python adds only this subdirectory to
# sys.path.  Add the repository root so absolute ``eeg_analysis.*`` imports
# work for both:
#   python eeg_analysis/fatigue_driving_prediction_system/function_one.py
#   python -m eeg_analysis.fatigue_driving_prediction_system.function_one
if __package__ in {None, ""}:
    project_root = Path(__file__).resolve().parents[2]
    if str(project_root) not in sys.path:
        sys.path.insert(0, str(project_root))

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import mne
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from eeg_analysis.detection.record_arousal import detect_eye_movements
from eeg_analysis.detection.record_alpha import detect_alpha
from eeg_analysis.fatigue_driving_prediction_system.behavioral_fatigue import (
    BehavioralFatigueEvaluation,
    GLOBAL_RT_WINDOW_SECONDS,
    PERSONALIZED_RT_MULTIPLIER,
    PHASE_ONE_DURATION_SECONDS,
    PHASE_ONE_RT_THRESHOLD,
    calculate_personalized_rt_threshold,
    evaluate_backward_behavioral_fatigue_events,
    first_behavioral_fatigue,
)
from eeg_analysis.fatigue_driving_prediction_system.physiological_fatigue import (
    RobustBaseline,
    TRAINING_POOLED_ALPHA_SCALE,
    TRAINING_POOLED_EYE_SCALE,
    compute_robust_baseline,
    rolling_event_counts,
)
from eeg_analysis.statistics_30s_alpha_eyeblink_of_fatigue.record_status_and_eyeblink_to_xlsx import (
    ReactionTimeEvent,
    extract_reaction_time_events_with_duration,
    write_reaction_time_events_xlsx,
)


@dataclass(frozen=True)
class FunctionOneConfig:
    behavioral_start_second: int = 0
    baseline_start_second: int = 1
    baseline_end_second: int = PHASE_ONE_DURATION_SECONDS
    fatigue_reaction_threshold: float = PHASE_ONE_RT_THRESHOLD
    global_rt_window_seconds: int = GLOBAL_RT_WINDOW_SECONDS
    personalized_multiplier: float = PERSONALIZED_RT_MULTIPLIER
    physiological_window_seconds: int = 30
    pooled_alpha_scale: float = TRAINING_POOLED_ALPHA_SCALE
    pooled_eye_scale: float = TRAINING_POOLED_EYE_SCALE


DEFAULT_CONFIG = FunctionOneConfig()


@dataclass(frozen=True)
class FunctionOneResult:
    record_id: str
    decision: str
    allow_driving: bool
    events: tuple[ReactionTimeEvent, ...]
    behavioral_evaluations: tuple[BehavioralFatigueEvaluation, ...]
    trigger_evaluation: BehavioralFatigueEvaluation | None
    rt_mean: float | None
    rt_median: float | None
    personalized_rt_threshold: float | None
    eye_seconds: tuple[int, ...]
    alpha_seconds: tuple[int, ...]
    alpha_baseline: RobustBaseline | None
    eye_baseline: RobustBaseline | None
    output_dir: Path


def record_id_from_edf(edf_path: str | Path) -> str:
    stem = Path(edf_path).stem
    return stem[:-4] if stem.casefold().endswith("_raw") else stem


def select_baseline_events(
    events: Sequence[ReactionTimeEvent],
    config: FunctionOneConfig = DEFAULT_CONFIG,
) -> list[ReactionTimeEvent]:
    return [
        event
        for event in events
        if config.baseline_start_second
        <= event.event_second
        <= config.baseline_end_second
    ]


def evaluate_phase_one_behavioral_fatigue(
    events: Sequence[ReactionTimeEvent],
    config: FunctionOneConfig = DEFAULT_CONFIG,
    *,
    recording_end_second: int | None = None,
) -> tuple[BehavioralFatigueEvaluation, ...]:
    """Evaluate Phase 1 onsets using complete backward 90-second data."""
    evaluations = evaluate_backward_behavioral_fatigue_events(
        events,
        config.fatigue_reaction_threshold,
        global_window_seconds=config.global_rt_window_seconds,
        recording_start_second=config.behavioral_start_second,
    )
    return tuple(
        evaluation
        for evaluation in evaluations
        if config.behavioral_start_second
        <= evaluation.event.event_second
        <= config.baseline_end_second
    )


def _find_fp2_channel(channel_names: Sequence[str]) -> str:
    for channel_name in channel_names:
        if str(channel_name).strip().casefold() == "fp2":
            return str(channel_name)
    for channel_name in channel_names:
        if "fp2" in str(channel_name).casefold():
            return str(channel_name)
    raise ValueError("EDF中找不到FP2通道。")


def _autosize_worksheet(worksheet) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        values = [
            str(worksheet.cell(row=row, column=column_index).value or "")
            for row in range(1, worksheet.max_row + 1)
        ]
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max(map(len, values), default=0) + 2, 12), 36
        )


def write_function_one_workbook(
    result: FunctionOneResult,
    output_path: str | Path,
    config: FunctionOneConfig = DEFAULT_CONFIG,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "功能一摘要"

    summary_rows = [
        ("record_id", result.record_id),
        ("功能一結果", result.decision),
        ("允許繼續駕駛", "是" if result.allow_driving else "否"),
        ("行為篩檢開始秒", config.behavioral_start_second),
        ("Baseline開始秒", config.baseline_start_second),
        ("Baseline結束秒", config.baseline_end_second),
        ("固定疲勞RT門檻_秒", config.fatigue_reaction_threshold),
        ("Backward Global RT窗口_秒", config.global_rt_window_seconds),
        ("前300秒RT事件數", len(result.events)),
        (
            "前300秒Local與Backward Global確認疲勞事件數",
            sum(item.sustained_fatigue for item in result.behavioral_evaluations),
        ),
        ("RT平均Baseline", result.rt_mean),
        ("RT中位數Baseline", result.rt_median),
        ("個人化RT疲勞門檻", result.personalized_rt_threshold),
        ("眼動秒數", len(result.eye_seconds) if result.allow_driving else None),
        ("Alpha秒數", len(result.alpha_seconds) if result.allow_driving else None),
        ("生理特徵窗口_秒", config.physiological_window_seconds),
        (
            "Alpha Baseline Median",
            result.alpha_baseline.median if result.alpha_baseline else None,
        ),
        (
            "Alpha Baseline MAD",
            result.alpha_baseline.mad if result.alpha_baseline else None,
        ),
        (
            "Alpha Baseline IQR",
            result.alpha_baseline.iqr if result.alpha_baseline else None,
        ),
        (
            "Alpha Baseline Scale",
            result.alpha_baseline.scale if result.alpha_baseline else None,
        ),
        (
            "Alpha Scale Method",
            result.alpha_baseline.scale_method if result.alpha_baseline else None,
        ),
        (
            "Eye Baseline Median",
            result.eye_baseline.median if result.eye_baseline else None,
        ),
        (
            "Eye Baseline MAD",
            result.eye_baseline.mad if result.eye_baseline else None,
        ),
        (
            "Eye Baseline IQR",
            result.eye_baseline.iqr if result.eye_baseline else None,
        ),
        (
            "Eye Baseline Scale",
            result.eye_baseline.scale if result.eye_baseline else None,
        ),
        (
            "Eye Scale Method",
            result.eye_baseline.scale_method if result.eye_baseline else None,
        ),
        (
            "觸發行為疲勞事件_秒",
            result.trigger_evaluation.event.event_second
            if result.trigger_evaluation
            else None,
        ),
        (
            "觸發Local RT_秒",
            result.trigger_evaluation.event.reaction_time
            if result.trigger_evaluation
            else None,
        ),
        (
            "觸發Backward Global RT_秒",
            result.trigger_evaluation.global_rt if result.trigger_evaluation else None,
        ),
        (
            "行為疲勞確認秒",
            result.trigger_evaluation.confirmation_second
            if result.trigger_evaluation
            else None,
        ),
        (
            "觸發原因",
            result.trigger_evaluation.trigger_reason
            if result.trigger_evaluation
            else None,
        ),
    ]
    summary.append(["項目", "數值"])
    for row in summary_rows:
        summary.append(list(row))

    event_sheet = workbook.create_sheet("Reaction Time事件")
    event_sheet.append(
        [
            "事件編號",
            "偏移Status",
            "偏移開始時間_秒",
            "導正開始時間_秒",
            "向上取整事件秒數",
            "Local RT_秒",
            "Backward Global RT_秒",
            "Backward窗口開始秒",
            "Backward窗口結束秒",
            "Phase",
            "Active Threshold_秒",
            "具完整90秒Backward窗口",
            "先前不同秒RT事件數",
            "Local>=Threshold",
            "Backward Global>=Threshold",
            "Local+Backward Global確認疲勞",
            "Behavioral Fatigue",
            "確認秒",
            "觸發原因",
            "Phase 1排除事件",
        ]
    )
    trigger_index = (
        result.trigger_evaluation.event.event_index
        if result.trigger_evaluation
        else None
    )
    for evaluation in result.behavioral_evaluations:
        event = evaluation.event
        event_sheet.append(
            [
                event.event_index,
                event.deviation_status,
                event.deviation_time,
                event.correction_start_time,
                event.event_second,
                event.reaction_time,
                evaluation.global_rt,
                evaluation.window_start_second,
                evaluation.window_end_second,
                "Phase 1",
                evaluation.active_threshold,
                evaluation.has_full_global_window,
                evaluation.past_event_count,
                evaluation.local_exceed,
                evaluation.global_exceed,
                evaluation.sustained_fatigue,
                evaluation.behavioral_fatigue,
                evaluation.confirmation_second,
                evaluation.trigger_reason,
                event.event_index == trigger_index,
            ]
        )

    for column in ("F", "G", "K"):
        for cell in event_sheet[column][1:]:
            cell.number_format = "0.0"
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        _autosize_worksheet(worksheet)
    workbook.save(output)
    return output


def save_rt_validation_plot(
    result: FunctionOneResult,
    output_path: str | Path,
    config: FunctionOneConfig = DEFAULT_CONFIG,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    plt.rcParams["font.sans-serif"] = ["Microsoft JhengHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    figure, axis = plt.subplots(figsize=(13, 6.5))
    normal_evaluations = [
        item for item in result.behavioral_evaluations if not item.behavioral_fatigue
    ]
    triggered_evaluations = [
        item for item in result.behavioral_evaluations if item.behavioral_fatigue
    ]
    if normal_evaluations:
        axis.scatter(
            [item.event.event_second for item in normal_evaluations],
            [item.event.reaction_time for item in normal_evaluations],
            color="#4472C4",
            label="Local RT（未觸發）",
            zorder=3,
        )
    if triggered_evaluations:
        axis.scatter(
            [item.event.event_second for item in triggered_evaluations],
            [item.event.reaction_time for item in triggered_evaluations],
            color="#C00000",
            label="Local RT（行為觸發）",
            zorder=4,
        )
    if result.behavioral_evaluations:
        axis.plot(
            [item.event.event_second for item in result.behavioral_evaluations],
            [item.global_rt for item in result.behavioral_evaluations],
            color="#ED7D31",
            marker="o",
            markersize=3,
            linewidth=1.6,
            label=(
                "Backward Global RT（含當次，"
                f"回看{config.global_rt_window_seconds}秒）"
            ),
            zorder=2,
        )
    axis.axhline(
        config.fatigue_reaction_threshold,
        color="#C00000",
        linestyle="--",
        linewidth=1.5,
        label=f"Phase 1門檻 {config.fatigue_reaction_threshold:g}秒",
    )
    if result.personalized_rt_threshold is not None:
        axis.axhline(
            result.personalized_rt_threshold,
            color="#70AD47",
            linestyle=":",
            linewidth=1.8,
            label=f"個人化門檻 {result.personalized_rt_threshold:.1f}秒",
        )
    if result.trigger_evaluation is not None:
        event = result.trigger_evaluation.event
        axis.annotate(
            f"{event.event_second}s, {result.trigger_evaluation.trigger_reason}\n"
            f"Local={event.reaction_time:.1f}, Backward Global={result.trigger_evaluation.global_rt:.2f}\n"
            f"Confirmed={result.trigger_evaluation.confirmation_second}s",
            (event.event_second, event.reaction_time),
            xytext=(8, 10),
            textcoords="offset points",
            fontsize=9,
        )

    detail_lines = [
        f"結果：{result.decision}",
        f"RT事件數：{len(result.events)}",
    ]
    if result.rt_mean is not None:
        detail_lines.append(f"RT平均：{result.rt_mean:.3f}秒")
    if result.rt_median is not None:
        detail_lines.append(f"RT中位數：{result.rt_median:.3f}秒")
    axis.text(
        0.99,
        0.97,
        "\n".join(detail_lines),
        transform=axis.transAxes,
        ha="right",
        va="top",
        bbox={"boxstyle": "round", "facecolor": "white", "alpha": 0.85},
    )
    axis.set_xlim(config.behavioral_start_second, config.baseline_end_second)
    axis.set_xlabel("記錄秒數（向上取整）")
    axis.set_ylabel("Reaction Time（秒）")
    axis.set_title(f"{result.record_id}：前300秒Reaction Time驗證圖")
    axis.grid(True, linestyle="--", alpha=0.3)
    axis.legend(loc="upper left")
    figure.tight_layout()
    figure.savefig(output, dpi=300, bbox_inches="tight")
    plt.close(figure)
    return output


def analyze_function_one(
    edf_path: str | Path,
    output_dir: str | Path | None = None,
    config: FunctionOneConfig = DEFAULT_CONFIG,
) -> FunctionOneResult:
    path = Path(edf_path)
    if not path.is_file():
        raise FileNotFoundError(f"EDF not found: {path}")
    record_id = record_id_from_edf(path)
    resolved_output_dir = (
        Path(output_dir)
        if output_dir is not None
        else Path("data")
        / "derived"
        / "fatigue_driving_prediction_system"
        / record_id
    )
    resolved_output_dir.mkdir(parents=True, exist_ok=True)

    all_events, recording_end_second = extract_reaction_time_events_with_duration(path)
    events = select_baseline_events(all_events, config)
    write_reaction_time_events_xlsx(
        all_events,
        resolved_output_dir / "reaction_time_events.xlsx",
        recording_end_second=recording_end_second,
    )
    behavioral_evaluations = evaluate_phase_one_behavioral_fatigue(
        all_events,
        config,
        recording_end_second=recording_end_second,
    )
    trigger_evaluation = first_behavioral_fatigue(behavioral_evaluations)
    if not events:
        decision = "INSUFFICIENT_DATA"
        allow_driving = False
    elif trigger_evaluation is not None:
        decision = "FATIGUE"
        allow_driving = False
    else:
        decision = "NON_FATIGUE"
        allow_driving = True

    rt_mean: float | None = None
    rt_median: float | None = None
    personalized_threshold: float | None = None
    eye_seconds: tuple[int, ...] = ()
    alpha_seconds: tuple[int, ...] = ()
    alpha_baseline: RobustBaseline | None = None
    eye_baseline: RobustBaseline | None = None

    if allow_driving:
        rt_values = [event.reaction_time for event in events]
        rt_mean = statistics.mean(rt_values)
        rt_median = statistics.median(rt_values)
        personalized_threshold = calculate_personalized_rt_threshold(
            rt_values,
            multiplier=config.personalized_multiplier,
            maximum_threshold=config.fatigue_reaction_threshold,
        )

        raw = mne.io.read_raw_edf(path, preload=False, verbose=False)
        try:
            fp2_channel = _find_fp2_channel(raw.ch_names)
            eye_output = resolved_output_dir / "eyeblink.dat"
            eye_seconds = tuple(
                detect_eye_movements(
                    raw,
                    [fp2_channel],
                    eye_output,
                    start_second=config.baseline_start_second,
                    end_second=config.baseline_end_second,
                )
            )
        finally:
            raw.close()

        alpha_result = detect_alpha(
            path,
            resolved_output_dir / "Alpha_function_one.dat",
            [fp2_channel],
            eye_seconds=eye_seconds,
            start_second=config.baseline_start_second,
            end_second=config.baseline_end_second,
        )
        alpha_seconds = tuple(alpha_result.alpha_seconds)

        complete_window_start = max(
            config.baseline_start_second + config.physiological_window_seconds - 1,
            config.physiological_window_seconds,
        )
        alpha_windows = rolling_event_counts(
            alpha_seconds,
            start_second=complete_window_start,
            end_second=config.baseline_end_second,
            window_seconds=config.physiological_window_seconds,
        )
        eye_windows = rolling_event_counts(
            eye_seconds,
            start_second=complete_window_start,
            end_second=config.baseline_end_second,
            window_seconds=config.physiological_window_seconds,
        )
        alpha_baseline = compute_robust_baseline(
            list(alpha_windows.values()),
            pooled_scale=config.pooled_alpha_scale,
        )
        eye_baseline = compute_robust_baseline(
            list(eye_windows.values()),
            pooled_scale=config.pooled_eye_scale,
        )
    result = FunctionOneResult(
        record_id=record_id,
        decision=decision,
        allow_driving=allow_driving,
        events=tuple(events),
        behavioral_evaluations=behavioral_evaluations,
        trigger_evaluation=trigger_evaluation,
        rt_mean=rt_mean,
        rt_median=rt_median,
        personalized_rt_threshold=personalized_threshold,
        eye_seconds=eye_seconds,
        alpha_seconds=alpha_seconds,
        alpha_baseline=alpha_baseline,
        eye_baseline=eye_baseline,
        output_dir=resolved_output_dir,
    )
    write_function_one_workbook(
        result, resolved_output_dir / "function_one_results.xlsx", config
    )
    save_rt_validation_plot(
        result, resolved_output_dir / "rt_validation.png", config
    )
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run Function One of the fatigue-driving prediction system."
    )
    parser.add_argument("--file", required=True, help="Input EDF path.")
    parser.add_argument("--output-dir", help="Output folder for this recording.")
    parser.add_argument(
        "--pooled-alpha-scale",
        type=float,
        default=TRAINING_POOLED_ALPHA_SCALE,
        help="Training-derived Alpha fallback scale when baseline MAD/IQR are zero.",
    )
    parser.add_argument(
        "--pooled-eye-scale",
        type=float,
        default=TRAINING_POOLED_EYE_SCALE,
        help="Training-derived Eye fallback scale when baseline MAD/IQR are zero.",
    )
    parser.add_argument(
        "--function-one-only",
        action="store_true",
        help="Stop after Function One instead of continuing to Function Two.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    function_one_config = FunctionOneConfig(
        pooled_alpha_scale=args.pooled_alpha_scale,
        pooled_eye_scale=args.pooled_eye_scale,
    )
    result = analyze_function_one(args.file, args.output_dir, function_one_config)
    print(f"功能一結果：{result.decision}")
    print(f"允許繼續駕駛：{'是' if result.allow_driving else '否'}")
    if result.trigger_evaluation is not None:
        print(
            "功能一行為觸發："
            f"第{result.trigger_evaluation.event.event_second}秒"
            f"（{result.trigger_evaluation.trigger_reason}）"
        )
    if result.allow_driving and not args.function_one_only:
        from eeg_analysis.fatigue_driving_prediction_system.function_two import (
            analyze_function_two,
        )

        function_two_result = analyze_function_two(
            args.file,
            result,
            result.output_dir,
        )
        print(f"功能二狀態：{function_two_result.status}")
        if function_two_result.target_event is not None:
            assert function_two_result.target_evaluation is not None
            print(
                "第一個疲勞事件："
                f"第{function_two_result.target_event.event_second}秒"
                f"（{function_two_result.target_evaluation.trigger_reason}）"
            )
        else:
            print("第一個疲勞事件：無")
        if function_two_result.first_warning_second is not None:
            print(
                "第一次警報："
                f"第{function_two_result.first_warning_second}秒"
                f"（{function_two_result.first_warning_reason}）"
            )
        else:
            print("第一次警報：無")
    print(f"輸出資料夾：{result.output_dir}")


if __name__ == "__main__":
    main()
