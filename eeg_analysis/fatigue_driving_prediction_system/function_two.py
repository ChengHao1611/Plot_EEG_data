"""Run the two-phase fatigue workflow for train_data or one EDF."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

if __package__ in {None, ""}:
    project_root = Path(__file__).resolve().parents[2]
    if str(project_root) not in sys.path:
        sys.path.insert(0, str(project_root))

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import mne
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from eeg_analysis.detection.record_alpha import (
    AlphaDetectionResult,
    BandPowerRecord,
    detect_alpha,
    save_dat_seconds,
)
from eeg_analysis.detection.record_arousal import detect_eye_movements
from eeg_analysis.fatigue_driving_prediction_system.behavioral_fatigue import (
    BehavioralFatigueEvaluation,
    GLOBAL_RT_WINDOW_SECONDS,
    PHASE_ONE_DURATION_SECONDS,
    evaluate_behavioral_fatigue_events,
    first_behavioral_fatigue,
)
from eeg_analysis.fatigue_driving_prediction_system.function_one import (
    FunctionOneConfig,
    FunctionOneResult,
    analyze_function_one,
)
from eeg_analysis.fatigue_driving_prediction_system.physiological_fatigue import (
    RobustBaseline,
    TRAINING_POOLED_ALPHA_SCALE,
    TRAINING_POOLED_EYE_SCALE,
    classify_lead_time,
    compute_fatigue_score,
)
from eeg_analysis.fatigue_driving_prediction_system.training_baseline import (
    DEFAULT_MANIFEST_PATH,
    DEFAULT_RESULTS_ROOT,
    read_training_record_ids,
)
from eeg_analysis.statistics_30s_alpha_eyeblink_of_fatigue.record_status_and_eyeblink_to_xlsx import (
    ReactionTimeEvent,
    extract_reaction_time_events_with_duration,
)


@dataclass(frozen=True)
class FunctionTwoConfig:
    baseline_end_second: int = PHASE_ONE_DURATION_SECONDS
    global_rt_window_seconds: int = GLOBAL_RT_WINDOW_SECONDS
    eye_window_seconds: int = 30
    alpha_window_seconds: int = 30
    score_threshold: float = 0.8
    confirmation_seconds: int = 4
    minimum_lead_seconds: int = 30
    maximum_lead_seconds: int = 60
    plot_seconds_before_fatigue: int = 90


DEFAULT_CONFIG = FunctionTwoConfig()


@dataclass(frozen=True)
class FunctionTwoFeatureRecord:
    second: int
    seconds_before_fatigue: int | None
    eye_detected: bool
    eye_window_count: int
    eye_alert: bool
    z_eye: float | None
    alpha_valid: bool
    theta_power: float | None
    alpha_power: float | None
    beta_power: float | None
    alpha_qualified: bool
    alpha_window_count: int
    alpha_alert: bool
    z_alpha: float | None
    fatigue_score: float | None
    score_above_threshold: bool
    consecutive_seconds: int
    warning: bool
    warning_reason: str
    target_fatigue: bool


@dataclass(frozen=True)
class FunctionTwoResult:
    record_id: str
    status: str
    analysis_start_second: int
    analysis_end_second: int
    personalized_rt_threshold: float | None
    alpha_baseline: RobustBaseline | None
    eye_baseline: RobustBaseline | None
    target_event: ReactionTimeEvent | None
    target_evaluation: BehavioralFatigueEvaluation | None
    first_warning_second: int | None
    first_warning_reason: str | None
    prediction_success: bool | None
    lead_seconds: int | None
    features: tuple[FunctionTwoFeatureRecord, ...]
    post_baseline_events: tuple[ReactionTimeEvent, ...]
    post_baseline_evaluations: tuple[BehavioralFatigueEvaluation, ...]
    eye_seconds: tuple[int, ...]
    alpha_result: AlphaDetectionResult | None
    output_dir: Path


@dataclass(frozen=True)
class BatchRecordResult:
    """Phase 1/2 results, or a captured error, for one manifest record."""

    record_id: str
    edf_path: Path | None
    function_one_result: FunctionOneResult | None
    function_two_result: FunctionTwoResult | None
    error: str | None


@dataclass(frozen=True)
class TrainingBatchResult:
    """Summary of one complete ``train_data`` batch invocation."""

    records: tuple[BatchRecordResult, ...]
    summary_path: Path

    @property
    def completed_count(self) -> int:
        return sum(item.error is None for item in self.records)

    @property
    def error_count(self) -> int:
        return sum(item.error is not None for item in self.records)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_EDF_ROOT = PROJECT_ROOT / "data" / "raw_edf" / "eeg"
DEFAULT_BATCH_SUMMARY_NAME = "training_batch_results.xlsx"


def find_first_post_baseline_fatigue_event(
    events: Sequence[ReactionTimeEvent],
    personalized_rt_threshold: float,
    baseline_end_second: int = PHASE_ONE_DURATION_SECONDS,
    global_rt_window_seconds: int = GLOBAL_RT_WINDOW_SECONDS,
    recording_end_second: int | None = None,
) -> ReactionTimeEvent | None:
    """Return the first post-baseline forward-confirmed fatigue event."""
    evaluation = find_first_post_baseline_fatigue_evaluation(
        events,
        personalized_rt_threshold,
        baseline_end_second,
        global_rt_window_seconds,
        recording_end_second,
    )
    return evaluation.event if evaluation is not None else None


def find_first_post_baseline_fatigue_evaluation(
    events: Sequence[ReactionTimeEvent],
    personalized_rt_threshold: float,
    baseline_end_second: int = PHASE_ONE_DURATION_SECONDS,
    global_rt_window_seconds: int = GLOBAL_RT_WINDOW_SECONDS,
    recording_end_second: int | None = None,
) -> BehavioralFatigueEvaluation | None:
    """Return the first Phase 2 Local/Forward-Global trigger."""
    evaluations = evaluate_behavioral_fatigue_events(
        events,
        personalized_rt_threshold,
        global_window_seconds=global_rt_window_seconds,
        recording_end_second=recording_end_second,
    )
    return first_behavioral_fatigue(
        evaluations,
        after_second=baseline_end_second,
    )


def classify_warning(
    fatigue_score: float | None,
    consecutive_seconds: int,
    config: FunctionTwoConfig = DEFAULT_CONFIG,
) -> tuple[bool, str]:
    """Classify a real-time warning from the shared robust score rule."""
    confirmed = (
        fatigue_score is not None
        and fatigue_score >= config.score_threshold
        and consecutive_seconds >= config.confirmation_seconds
    )
    return (
        (True, "ROBUST_Z_MIN_CONFIRMED")
        if confirmed
        else (False, "NONE")
    )


def build_function_two_features(
    *,
    start_second: int,
    end_second: int,
    eye_seconds: Iterable[int],
    alpha_records: Iterable[BandPowerRecord],
    alpha_baseline: RobustBaseline,
    eye_baseline: RobustBaseline,
    target_second: int | None,
    config: FunctionTwoConfig = DEFAULT_CONFIG,
) -> list[FunctionTwoFeatureRecord]:
    """Build per-second windows, carrying pre-300 records into Function Two."""
    if start_second < 1:
        raise ValueError("start_second must be at least 1")
    if end_second < start_second:
        return []
    eye_second_set = {int(second) for second in eye_seconds}
    alpha_by_second = {record.second: record for record in alpha_records}
    qualified_alpha_seconds = {
        record.second
        for record in alpha_by_second.values()
        if record.alpha_qualified
    }

    features: list[FunctionTwoFeatureRecord] = []
    consecutive_seconds = 0
    for second in range(start_second, end_second + 1):
        eye_window_start = second - config.eye_window_seconds + 1
        alpha_window_start = second - config.alpha_window_seconds + 1
        eye_window_count = sum(
            window_second in eye_second_set
            for window_second in range(eye_window_start, second + 1)
        )
        alpha_window_count = sum(
            window_second in qualified_alpha_seconds
            for window_second in range(alpha_window_start, second + 1)
        )
        score = compute_fatigue_score(
            alpha_window_count,
            eye_window_count,
            alpha_baseline=alpha_baseline,
            eye_baseline=eye_baseline,
        )
        alpha_record = alpha_by_second.get(second)
        alpha_valid = bool(
            alpha_record is not None
            and not alpha_record.excluded_by_eye
            and alpha_record.alpha_power is not None
        )
        alpha_qualified = second in qualified_alpha_seconds
        eye_alert = score.z_eye is not None and score.z_eye >= config.score_threshold
        alpha_alert = (
            score.z_alpha is not None and score.z_alpha >= config.score_threshold
        )
        score_above_threshold = (
            score.score is not None and score.score >= config.score_threshold
        )
        consecutive_seconds = (
            consecutive_seconds + 1 if score_above_threshold else 0
        )
        warning, warning_reason = classify_warning(
            score.score,
            consecutive_seconds,
            config,
        )
        features.append(
            FunctionTwoFeatureRecord(
                second=second,
                seconds_before_fatigue=(
                    target_second - second if target_second is not None else None
                ),
                eye_detected=second in eye_second_set,
                eye_window_count=eye_window_count,
                eye_alert=eye_alert,
                z_eye=score.z_eye,
                alpha_valid=alpha_valid,
                theta_power=(alpha_record.theta_power if alpha_record else None),
                alpha_power=(alpha_record.alpha_power if alpha_record else None),
                beta_power=(alpha_record.beta_power if alpha_record else None),
                alpha_qualified=alpha_qualified,
                alpha_window_count=alpha_window_count,
                alpha_alert=alpha_alert,
                z_alpha=score.z_alpha,
                fatigue_score=score.score,
                score_above_threshold=score_above_threshold,
                consecutive_seconds=consecutive_seconds,
                warning=warning,
                warning_reason=warning_reason,
                target_fatigue=target_second == second,
            )
        )
    return features


def _find_fp2_channel(channel_names: Sequence[str]) -> str:
    for channel_name in channel_names:
        if str(channel_name).strip().casefold() == "fp2":
            return str(channel_name)
    for channel_name in channel_names:
        if "fp2" in str(channel_name).casefold():
            return str(channel_name)
    raise ValueError("EDF中找不到FP2通道。")


def _autosize_worksheet(worksheet) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        values = [
            str(worksheet.cell(row=row, column=column_index).value or "")
            for row in range(1, worksheet.max_row + 1)
        ]
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max(map(len, values), default=0) + 2, 12), 36
        )


def resolve_training_edf(
    record_id: str,
    edf_root: str | Path = DEFAULT_EDF_ROOT,
) -> Path:
    """Resolve one manifest record to exactly one case-insensitive EDF file."""
    root = Path(edf_root)
    if not root.is_dir():
        raise FileNotFoundError(f"找不到EDF資料夾：{root}")

    expected_stems = {record_id.casefold(), f"{record_id}_raw".casefold()}
    matches = sorted(
        (
            path
            for path in root.rglob("*")
            if path.is_file()
            and path.suffix.casefold() == ".edf"
            and path.stem.casefold() in expected_stems
        ),
        key=lambda path: str(path).casefold(),
    )
    if not matches:
        raise FileNotFoundError(
            f"train_data中的{record_id}找不到對應EDF：{root}"
        )
    if len(matches) > 1:
        raise ValueError(
            f"{record_id}對應到多個EDF："
            + "、".join(str(path) for path in matches)
        )
    return matches[0]


def write_training_batch_summary(
    records: Sequence[BatchRecordResult],
    output_path: str | Path,
) -> Path:
    """Write one row per manifest record with both phase outcomes."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "批次分析摘要"
    worksheet.append(
        [
            "record_id",
            "EDF路徑",
            "批次狀態",
            "錯誤",
            "Phase 1結果",
            "允許進入Phase 2",
            "Phase 1觸發秒",
            "Phase 1觸發原因",
            "Phase 1確認秒",
            "個人化RT門檻",
            "Phase 2狀態",
            "Behavioral Onset秒",
            "Behavioral Onset原因",
            "Behavioral Onset Local RT",
            "Behavioral Onset Forward Global RT",
            "Behavioral Confirmation秒",
            "第一次生理警報秒",
            "第一次生理警報原因",
            "預測成功",
            "提前秒數",
            "輸出資料夾",
        ]
    )

    for item in records:
        phase_one = item.function_one_result
        phase_two = item.function_two_result
        phase_one_trigger = phase_one.trigger_evaluation if phase_one else None
        target = phase_two.target_evaluation if phase_two else None
        output_dir = (
            phase_two.output_dir
            if phase_two is not None
            else phase_one.output_dir
            if phase_one is not None
            else None
        )
        worksheet.append(
            [
                item.record_id,
                str(item.edf_path) if item.edf_path else None,
                "COMPLETED" if item.error is None else "ERROR",
                item.error,
                phase_one.decision if phase_one else None,
                phase_one.allow_driving if phase_one else None,
                phase_one_trigger.event.event_second if phase_one_trigger else None,
                phase_one_trigger.trigger_reason if phase_one_trigger else None,
                phase_one_trigger.confirmation_second if phase_one_trigger else None,
                phase_one.personalized_rt_threshold if phase_one else None,
                phase_two.status if phase_two else None,
                target.event.event_second if target else None,
                target.trigger_reason if target else None,
                target.event.reaction_time if target else None,
                target.global_rt if target else None,
                target.confirmation_second if target else None,
                phase_two.first_warning_second if phase_two else None,
                phase_two.first_warning_reason if phase_two else None,
                phase_two.prediction_success if phase_two else None,
                phase_two.lead_seconds if phase_two else None,
                str(output_dir) if output_dir else None,
            ]
        )

    worksheet.freeze_panes = "A2"
    for column in ("J", "N", "O"):
        for cell in worksheet[column][1:]:
            cell.number_format = "0.0"
    _autosize_worksheet(worksheet)
    workbook.save(output)
    return output


def write_function_two_workbook(
    result: FunctionTwoResult,
    output_path: str | Path,
    config: FunctionTwoConfig = DEFAULT_CONFIG,
) -> Path:
    """Write the Function Two summary, per-second features, and RT ground truth."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "功能二摘要"
    summary.append(["項目", "數值"])
    summary_rows = [
        ("record_id", result.record_id),
        ("功能二狀態", result.status),
        ("分析開始秒", result.analysis_start_second),
        ("分析結束秒", result.analysis_end_second),
        ("個人化RT疲勞門檻", result.personalized_rt_threshold),
        ("Forward Global RT窗口_秒", config.global_rt_window_seconds),
        ("眼動窗口_秒", config.eye_window_seconds),
        ("Alpha窗口_秒", config.alpha_window_seconds),
        ("生理分數", "min(Z_Alpha, Z_Eye)"),
        ("生理分數門檻_h", config.score_threshold),
        ("連續確認秒數", config.confirmation_seconds),
        ("EWMA", "否"),
        (
            "Alpha Baseline Median",
            result.alpha_baseline.median if result.alpha_baseline else None,
        ),
        (
            "Alpha Baseline Scale",
            result.alpha_baseline.scale if result.alpha_baseline else None,
        ),
        (
            "Alpha Scale Method",
            result.alpha_baseline.scale_method if result.alpha_baseline else None,
        ),
        (
            "Eye Baseline Median",
            result.eye_baseline.median if result.eye_baseline else None,
        ),
        (
            "Eye Baseline Scale",
            result.eye_baseline.scale if result.eye_baseline else None,
        ),
        (
            "Eye Scale Method",
            result.eye_baseline.scale_method if result.eye_baseline else None,
        ),
        (
            "第一個疲勞事件_秒",
            result.target_event.event_second if result.target_event else None,
        ),
        (
            "第一個疲勞事件RT_秒",
            result.target_event.reaction_time if result.target_event else None,
        ),
        (
            "第一個疲勞事件Forward Global RT_秒",
            result.target_evaluation.global_rt if result.target_evaluation else None,
        ),
        (
            "第一個疲勞事件確認秒",
            result.target_evaluation.confirmation_second
            if result.target_evaluation
            else None,
        ),
        (
            "第一個疲勞事件觸發原因",
            result.target_evaluation.trigger_reason
            if result.target_evaluation
            else None,
        ),
        ("第一次警報_秒", result.first_warning_second),
        ("第一次警報原因", result.first_warning_reason),
        (
            "成功提前預測",
            "是"
            if result.prediction_success is True
            else "否"
            if result.prediction_success is False
            else None,
        ),
        ("提前秒數", result.lead_seconds),
        ("逐秒資料筆數", len(result.features)),
    ]
    for row in summary_rows:
        summary.append(list(row))

    feature_sheet = workbook.create_sheet("功能二逐秒特徵")
    feature_sheet.append(
        [
            "秒數",
            "距離疲勞事件_秒",
            "該秒有眼動",
            f"EyeWindow{config.eye_window_seconds}",
            "Z_Eye>=h",
            "Z_Eye",
            "該秒Alpha有效",
            "Theta Power",
            "Alpha Power",
            "Beta Power",
            "符合Alpha條件",
            f"AlphaWindow{config.alpha_window_seconds}",
            "Z_Alpha>=h",
            "Z_Alpha",
            "S=min(Z_Alpha,Z_Eye)",
            "S>=h",
            "連續成立秒數",
            "警報",
            "警報原因",
            "第一個疲勞事件",
        ]
    )
    for feature in result.features:
        feature_sheet.append(
            [
                feature.second,
                feature.seconds_before_fatigue,
                feature.eye_detected,
                feature.eye_window_count,
                feature.eye_alert,
                feature.z_eye,
                feature.alpha_valid,
                feature.theta_power,
                feature.alpha_power,
                feature.beta_power,
                feature.alpha_qualified,
                feature.alpha_window_count,
                feature.alpha_alert,
                feature.z_alpha,
                feature.fatigue_score,
                feature.score_above_threshold,
                feature.consecutive_seconds,
                feature.warning,
                feature.warning_reason,
                feature.target_fatigue,
            ]
        )

    rt_sheet = workbook.create_sheet("300秒後RT事件")
    rt_sheet.append(
        [
            "事件編號",
            "事件秒數",
            "Local RT_秒",
            "Forward Global RT_秒",
            "Forward窗口開始秒",
            "Forward窗口結束秒",
            "Phase",
            "Active Threshold_秒",
            "具完整90秒Forward窗口",
            "後續RT事件數",
            "Local>=Threshold",
            "Forward Global>=Threshold",
            "Local+Forward Global確認疲勞",
            "Behavioral Fatigue",
            "確認秒",
            "觸發原因",
            "第一個疲勞事件",
        ]
    )
    target_index = result.target_event.event_index if result.target_event else None
    for evaluation in result.post_baseline_evaluations:
        event = evaluation.event
        rt_sheet.append(
            [
                event.event_index,
                event.event_second,
                event.reaction_time,
                evaluation.global_rt,
                evaluation.window_start_second,
                evaluation.window_end_second,
                "Phase 2",
                evaluation.active_threshold,
                evaluation.has_full_global_window,
                evaluation.future_event_count,
                evaluation.local_exceed,
                evaluation.global_exceed,
                evaluation.sustained_fatigue,
                evaluation.behavioral_fatigue,
                evaluation.confirmation_second,
                evaluation.trigger_reason,
                event.event_index == target_index,
            ]
        )
    for column in ("C", "D", "H"):
        for cell in rt_sheet[column][1:]:
            cell.number_format = "0.0"

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        _autosize_worksheet(worksheet)
    workbook.save(output)
    return output


def save_pre_fatigue_plot(
    result: FunctionTwoResult,
    output_path: str | Path,
    config: FunctionTwoConfig = DEFAULT_CONFIG,
) -> Path | None:
    """Plot raw 30-second features and robust Z scores before the target."""
    if result.target_event is None or not result.features:
        return None

    target_second = result.target_event.event_second
    plot_start = target_second - config.plot_seconds_before_fatigue
    plot_features = [
        feature for feature in result.features if feature.second >= plot_start
    ]
    if not plot_features:
        return None

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    plt.rcParams["font.sans-serif"] = ["Microsoft JhengHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    relative_seconds = [feature.second - target_second for feature in plot_features]
    eye_values = [feature.eye_window_count for feature in plot_features]
    alpha_values = [feature.alpha_window_count for feature in plot_features]
    z_alpha_values = [feature.z_alpha for feature in plot_features]
    z_eye_values = [feature.z_eye for feature in plot_features]

    figure, (raw_axis, score_axis) = plt.subplots(
        2, 1, figsize=(13, 8.5), sharex=True
    )
    raw_axis.plot(
        relative_seconds,
        eye_values,
        color="#4472C4",
        linewidth=1.8,
        label=f"EyeWindow{config.eye_window_seconds}",
    )
    raw_axis.plot(
        relative_seconds,
        alpha_values,
        color="#70AD47",
        linewidth=1.8,
        label=f"AlphaWindow{config.alpha_window_seconds}",
    )
    raw_axis.set_ylabel("30秒事件次數")
    raw_axis.set_title("原始30秒滑動窗口")
    raw_axis.grid(True, linestyle="--", alpha=0.3)

    score_axis.plot(relative_seconds, z_alpha_values, color="#7030A0", label="Z-Alpha")
    score_axis.plot(relative_seconds, z_eye_values, color="#ED7D31", label="Z-Eye")
    score_axis.axhline(
        config.score_threshold,
        color="#C00000",
        linestyle="--",
        label=f"h={config.score_threshold:g}",
    )
    score_axis.set_ylabel("Robust Z")
    score_axis.set_xlabel("距離第一個疲勞事件的秒數（0 = 疲勞事件）")
    score_axis.set_title(
        "Z-Alpha與Z-Eye皆達門檻"
        f"（連續{config.confirmation_seconds}秒確認）"
    )
    score_axis.grid(True, linestyle="--", alpha=0.3)

    for axis in (raw_axis, score_axis):
        axis.axvline(0, color="#7030A0", linewidth=1.8, label="疲勞事件")
        if result.first_warning_second is not None:
            axis.axvline(
                result.first_warning_second - target_second,
                color="#008C95",
                linestyle="-.",
                linewidth=1.8,
                label="正式生理警報",
            )
        axis.set_xlim(min(relative_seconds), 0.5)
        axis.legend(loc="best")

    warning_text = (
        f"第一次警報：第{result.first_warning_second}秒，"
        f"提前{result.lead_seconds}秒，原因={result.first_warning_reason}"
        if result.first_warning_second is not None
        else "第一個疲勞事件前未發出警報"
    )
    figure.suptitle(
        f"{result.record_id}：第一個疲勞事件前{config.plot_seconds_before_fatigue}秒\n"
        f"疲勞事件=第{target_second}秒；{warning_text}",
        fontsize=14,
    )
    figure.tight_layout(rect=(0, 0, 1, 0.92))
    figure.savefig(output, dpi=300, bbox_inches="tight")
    plt.close(figure)
    return output


def save_behavioral_rt_debug_plot(
    record_id: str,
    evaluations: Sequence[BehavioralFatigueEvaluation],
    target_evaluation: BehavioralFatigueEvaluation | None,
    output_path: str | Path,
    config: FunctionTwoConfig = DEFAULT_CONFIG,
) -> Path | None:
    """Plot Local RT, forward Global RT, active thresholds, and onset."""
    if not evaluations:
        return None

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    ordered = sorted(
        evaluations,
        key=lambda item: (
            item.event.event_second,
            item.event.deviation_time,
            item.event.event_index,
        ),
    )
    seconds = [item.event.event_second for item in ordered]
    local_values = [item.event.reaction_time for item in ordered]
    global_values = [item.global_rt for item in ordered]
    threshold_values = [item.active_threshold for item in ordered]

    plt.rcParams["font.sans-serif"] = ["Microsoft JhengHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    figure, axis = plt.subplots(figsize=(14, 7))
    axis.scatter(seconds, local_values, color="#4472C4", s=30, label="Local RT")
    axis.plot(
        seconds,
        global_values,
        color="#ED7D31",
        marker="o",
        markersize=3,
        linewidth=1.7,
        label=f"Forward Global RT（含當次，往後{config.global_rt_window_seconds}秒）",
    )
    axis.step(
        seconds,
        threshold_values,
        where="post",
        color="#C00000",
        linestyle="--",
        linewidth=1.7,
        label="Active Threshold",
    )
    axis.axvline(
        config.baseline_end_second,
        color="#595959",
        linestyle=":",
        linewidth=1.7,
        label="Phase 1／2邊界",
    )

    triggered = [item for item in ordered if item.behavioral_fatigue]
    if triggered:
        axis.scatter(
            [item.event.event_second for item in triggered],
            [item.event.reaction_time for item in triggered],
            color="#C00000",
            edgecolor="white",
            linewidth=0.7,
            s=75,
            zorder=5,
            label="Behavioral Trigger",
        )
    if target_evaluation is not None:
        target_second = target_evaluation.event.event_second
        axis.axvline(
            target_second,
            color="#008C95",
            linewidth=2,
            label="Phase 2 Behavioral Onset",
        )
        axis.annotate(
            target_evaluation.trigger_reason,
            (target_second, target_evaluation.event.reaction_time),
            xytext=(8, 10),
            textcoords="offset points",
            fontsize=9,
        )

    axis.set_xlabel("事件秒數（整數）")
    axis.set_ylabel("Reaction Time（秒）")
    axis.set_title(f"{record_id}：Behavioral RT Debug")
    axis.grid(True, linestyle="--", alpha=0.3)
    axis.legend(loc="best")
    figure.tight_layout()
    figure.savefig(output, dpi=300, bbox_inches="tight")
    plt.close(figure)
    return output


def _empty_result(
    function_one_result: FunctionOneResult,
    *,
    status: str,
    output_dir: Path,
    personalized_rt_threshold: float | None,
    config: FunctionTwoConfig,
) -> FunctionTwoResult:
    result = FunctionTwoResult(
        record_id=function_one_result.record_id,
        status=status,
        analysis_start_second=config.baseline_end_second + 1,
        analysis_end_second=config.baseline_end_second,
        personalized_rt_threshold=personalized_rt_threshold,
        alpha_baseline=function_one_result.alpha_baseline,
        eye_baseline=function_one_result.eye_baseline,
        target_event=None,
        target_evaluation=None,
        first_warning_second=None,
        first_warning_reason=None,
        prediction_success=None,
        lead_seconds=None,
        features=(),
        post_baseline_events=(),
        post_baseline_evaluations=(),
        eye_seconds=(),
        alpha_result=None,
        output_dir=output_dir,
    )
    write_function_two_workbook(
        result, output_dir / "function_two_results.xlsx", config
    )
    return result


def analyze_function_two(
    edf_path: str | Path,
    function_one_result: FunctionOneResult,
    output_dir: str | Path | None = None,
    config: FunctionTwoConfig = DEFAULT_CONFIG,
) -> FunctionTwoResult:
    """Run Function Two using the RT baseline returned by Function One."""
    path = Path(edf_path)
    if not path.is_file():
        raise FileNotFoundError(f"EDF not found: {path}")
    resolved_output_dir = (
        Path(output_dir) if output_dir is not None else function_one_result.output_dir
    )
    resolved_output_dir.mkdir(parents=True, exist_ok=True)

    personalized_threshold = function_one_result.personalized_rt_threshold
    if not function_one_result.allow_driving:
        return _empty_result(
            function_one_result,
            status="SKIPPED_FUNCTION_ONE_NOT_PASSED",
            output_dir=resolved_output_dir,
            personalized_rt_threshold=personalized_threshold,
            config=config,
        )
    if personalized_threshold is None:
        return _empty_result(
            function_one_result,
            status="INSUFFICIENT_RT_BASELINE",
            output_dir=resolved_output_dir,
            personalized_rt_threshold=None,
            config=config,
        )
    if (
        function_one_result.alpha_baseline is None
        or function_one_result.eye_baseline is None
        or not function_one_result.alpha_baseline.valid
        or not function_one_result.eye_baseline.valid
    ):
        return _empty_result(
            function_one_result,
            status="INSUFFICIENT_PHYSIOLOGICAL_BASELINE",
            output_dir=resolved_output_dir,
            personalized_rt_threshold=personalized_threshold,
            config=config,
        )

    all_events, recording_end_second = extract_reaction_time_events_with_duration(path)
    behavioral_evaluations = evaluate_behavioral_fatigue_events(
        all_events,
        personalized_threshold,
        global_window_seconds=config.global_rt_window_seconds,
        recording_end_second=recording_end_second,
    )
    post_baseline_evaluations = tuple(
        evaluation
        for evaluation in behavioral_evaluations
        if evaluation.event.event_second > config.baseline_end_second
    )
    target_evaluation = first_behavioral_fatigue(
        post_baseline_evaluations,
    )
    target_event = target_evaluation.event if target_evaluation else None

    raw = mne.io.read_raw_edf(path, preload=False, verbose=False)
    try:
        fp2_channel = _find_fp2_channel(raw.ch_names)
        sample_rate = float(raw.info["sfreq"])
        samples_per_second = max(int(round(sample_rate)), 1)
        signal_recording_end_second = int(raw.n_times // samples_per_second)
        recording_end_second = min(
            recording_end_second,
            signal_recording_end_second,
        )
        requested_end_second = (
            target_event.event_second if target_event else recording_end_second
        )
        analysis_end_second = min(requested_end_second, recording_end_second)
        analysis_start_second = config.baseline_end_second + 1
        if analysis_end_second < analysis_start_second:
            return _empty_result(
                function_one_result,
                status="INSUFFICIENT_POST_BASELINE_DATA",
                output_dir=resolved_output_dir,
                personalized_rt_threshold=personalized_threshold,
                config=config,
            )

        post_eye_seconds = tuple(
            detect_eye_movements(
                raw,
                [fp2_channel],
                resolved_output_dir / "eyeblink_function_two.dat",
                start_second=analysis_start_second,
                end_second=analysis_end_second,
            )
        )
    finally:
        raw.close()

    combined_eye_seconds = tuple(
        sorted(set(function_one_result.eye_seconds) | set(post_eye_seconds))
    )
    save_dat_seconds(
        resolved_output_dir / "eyeblink_function_two.dat",
        combined_eye_seconds,
    )
    print(
        "功能二完整眼動DAT已更新為第1秒至"
        f"第{analysis_end_second}秒，共{len(combined_eye_seconds)}個眼動秒。"
    )

    alpha_result = detect_alpha(
        path,
        resolved_output_dir / "Alpha_function_two.dat",
        [fp2_channel],
        eye_seconds=combined_eye_seconds,
        start_second=1,
        end_second=analysis_end_second,
    )
    print(
        "功能二完整Alpha DAT已更新為第1秒至"
        f"第{analysis_end_second}秒，共{len(alpha_result.alpha_seconds)}個Alpha秒。"
    )
    features = build_function_two_features(
        start_second=analysis_start_second,
        end_second=analysis_end_second,
        eye_seconds=combined_eye_seconds,
        alpha_records=alpha_result.records,
        alpha_baseline=function_one_result.alpha_baseline,
        eye_baseline=function_one_result.eye_baseline,
        target_second=target_event.event_second if target_event else None,
        config=config,
    )
    first_warning = next((feature for feature in features if feature.warning), None)
    first_warning_second = first_warning.second if first_warning else None
    first_warning_reason = first_warning.warning_reason if first_warning else None

    lead_seconds: int | None = None
    prediction_success: bool | None = None
    if target_event is not None:
        target_window_start = max(
            analysis_start_second,
            target_event.event_second - config.maximum_lead_seconds,
        )
        target_window_end = target_event.event_second - config.minimum_lead_seconds
        if target_window_start > target_window_end:
            status = "NOT_EVALUABLE_FOR_REQUIRED_LEAD"
            prediction_success = None
        elif first_warning_second is not None:
            lead_seconds = target_event.event_second - first_warning_second
            status = classify_lead_time(
                lead_seconds,
                min_lead_seconds=config.minimum_lead_seconds,
                max_lead_seconds=config.maximum_lead_seconds,
            )
            prediction_success = status == "TARGET_PREDICTED_30_TO_60"
        else:
            status = "TARGET_WITHOUT_WARNING"
            prediction_success = False
    else:
        status = (
            "NO_TARGET_FATIGUE_WITH_WARNING"
            if first_warning_second is not None
            else "NO_TARGET_FATIGUE"
        )

    result = FunctionTwoResult(
        record_id=function_one_result.record_id,
        status=status,
        analysis_start_second=analysis_start_second,
        analysis_end_second=analysis_end_second,
        personalized_rt_threshold=personalized_threshold,
        alpha_baseline=function_one_result.alpha_baseline,
        eye_baseline=function_one_result.eye_baseline,
        target_event=target_event,
        target_evaluation=target_evaluation,
        first_warning_second=first_warning_second,
        first_warning_reason=first_warning_reason,
        prediction_success=prediction_success,
        lead_seconds=lead_seconds,
        features=tuple(features),
        post_baseline_events=tuple(
            evaluation.event
            for evaluation in post_baseline_evaluations
            if evaluation.event.event_second <= analysis_end_second
        ),
        post_baseline_evaluations=tuple(
            evaluation
            for evaluation in post_baseline_evaluations
            if evaluation.event.event_second <= analysis_end_second
        ),
        eye_seconds=combined_eye_seconds,
        alpha_result=alpha_result,
        output_dir=resolved_output_dir,
    )
    write_function_two_workbook(
        result, resolved_output_dir / "function_two_results.xlsx", config
    )
    save_behavioral_rt_debug_plot(
        result.record_id,
        (*function_one_result.behavioral_evaluations, *result.post_baseline_evaluations),
        result.target_evaluation,
        resolved_output_dir / "behavioral_rt_debug.png",
        config,
    )
    save_pre_fatigue_plot(
        result,
        resolved_output_dir / "function_two_pre_fatigue_90s.png",
        config,
    )
    return result


def analyze_two_phase_recording(
    edf_path: str | Path,
    output_dir: str | Path | None = None,
    function_one_config: FunctionOneConfig | None = None,
    function_two_config: FunctionTwoConfig = DEFAULT_CONFIG,
) -> tuple[FunctionOneResult, FunctionTwoResult]:
    """Run both phases and always export a Phase 2 status workbook."""
    phase_one = analyze_function_one(
        edf_path,
        output_dir,
        function_one_config or FunctionOneConfig(),
    )
    phase_two = analyze_function_two(
        edf_path,
        phase_one,
        phase_one.output_dir,
        function_two_config,
    )
    return phase_one, phase_two


def analyze_training_manifest(
    manifest_path: str | Path = DEFAULT_MANIFEST_PATH,
    edf_root: str | Path = DEFAULT_EDF_ROOT,
    results_root: str | Path = DEFAULT_RESULTS_ROOT,
    *,
    function_one_config: FunctionOneConfig | None = None,
    function_two_config: FunctionTwoConfig = DEFAULT_CONFIG,
) -> TrainingBatchResult:
    """Run Phase 1 and Phase 2 for every record listed in ``train_data``.

    Individual errors are captured so later recordings continue.  The batch
    workbook is rewritten after every recording, preserving partial progress
    during a long analysis run.
    """
    record_ids = read_training_record_ids(manifest_path)
    output_root = Path(results_root)
    output_root.mkdir(parents=True, exist_ok=True)
    summary_path = output_root / DEFAULT_BATCH_SUMMARY_NAME
    records: list[BatchRecordResult] = []
    phase_one_config = function_one_config or FunctionOneConfig()

    for index, record_id in enumerate(record_ids, start=1):
        print(f"[{index}/{len(record_ids)}] 開始分析 {record_id}")
        edf_path: Path | None = None
        phase_one: FunctionOneResult | None = None
        phase_two: FunctionTwoResult | None = None
        error: str | None = None
        try:
            edf_path = resolve_training_edf(record_id, edf_root)
            record_output_dir = output_root / record_id
            phase_one, phase_two = analyze_two_phase_recording(
                edf_path,
                record_output_dir,
                phase_one_config,
                function_two_config,
            )
            print(
                f"[{index}/{len(record_ids)}] {record_id} 完成："
                f"Phase 1={phase_one.decision}，Phase 2={phase_two.status}"
            )
        except Exception as exc:
            error = f"{type(exc).__name__}: {exc}"
            print(f"[{index}/{len(record_ids)}] {record_id} 失敗：{error}")

        records.append(
            BatchRecordResult(
                record_id=record_id,
                edf_path=edf_path,
                function_one_result=phase_one,
                function_two_result=phase_two,
                error=error,
            )
        )
        write_training_batch_summary(records, summary_path)

    return TrainingBatchResult(tuple(records), summary_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run Phase 1 and Phase 2 for every record in train_data. "
            "Use --file to analyze one EDF only."
        )
    )
    parser.add_argument("--file", help="Optional single EDF path instead of batch mode.")
    parser.add_argument(
        "--output-dir",
        help="Single-file output folder; used only together with --file.",
    )
    parser.add_argument(
        "--manifest",
        type=Path,
        default=DEFAULT_MANIFEST_PATH,
        help="Batch record manifest (default: project train_data).",
    )
    parser.add_argument(
        "--edf-dir",
        type=Path,
        default=DEFAULT_EDF_ROOT,
        help="Folder searched recursively for <record_id>_raw.EDF files.",
    )
    parser.add_argument(
        "--results-root",
        type=Path,
        default=DEFAULT_RESULTS_ROOT,
        help="Batch output root, with one subfolder per record.",
    )
    parser.add_argument(
        "--pooled-alpha-scale",
        type=float,
        default=TRAINING_POOLED_ALPHA_SCALE,
    )
    parser.add_argument(
        "--pooled-eye-scale",
        type=float,
        default=TRAINING_POOLED_EYE_SCALE,
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.output_dir and not args.file:
        raise SystemExit("--output-dir只能與--file一起使用；批次模式請使用--results-root。")

    function_one_config = FunctionOneConfig(
        pooled_alpha_scale=args.pooled_alpha_scale,
        pooled_eye_scale=args.pooled_eye_scale,
    )
    if not args.file:
        batch = analyze_training_manifest(
            args.manifest,
            args.edf_dir,
            args.results_root,
            function_one_config=function_one_config,
        )
        print(
            "批次分析完成："
            f"成功{batch.completed_count}筆，失敗{batch.error_count}筆。"
        )
        print(f"批次摘要：{batch.summary_path}")
        return

    function_one_result, result = analyze_two_phase_recording(
        args.file,
        args.output_dir,
        function_one_config,
    )
    print(f"功能一結果：{function_one_result.decision}")
    if not function_one_result.allow_driving:
        print("功能一未通過；功能二已匯出跳過狀態。")
    print(f"功能二狀態：{result.status}")
    target_text = "無"
    if result.target_event is not None:
        assert result.target_evaluation is not None
        target_text = (
            f"第{result.target_event.event_second}秒"
            f"（{result.target_evaluation.trigger_reason}）"
        )
    print(
        "第一個疲勞事件："
        + target_text
    )
    print(
        "第一次警報："
        + (
            f"第{result.first_warning_second}秒（{result.first_warning_reason}）"
            if result.first_warning_second is not None
            else "無"
        )
    )
    print(f"輸出資料夾：{result.output_dir}")


if __name__ == "__main__":
    main()


__all__ = [
    "BatchRecordResult",
    "DEFAULT_CONFIG",
    "DEFAULT_EDF_ROOT",
    "FunctionTwoConfig",
    "FunctionTwoFeatureRecord",
    "FunctionTwoResult",
    "TrainingBatchResult",
    "analyze_training_manifest",
    "analyze_two_phase_recording",
    "analyze_function_two",
    "build_function_two_features",
    "classify_warning",
    "find_first_post_baseline_fatigue_evaluation",
    "find_first_post_baseline_fatigue_event",
    "resolve_training_edf",
    "save_behavioral_rt_debug_plot",
    "save_pre_fatigue_plot",
    "write_function_two_workbook",
    "write_training_batch_summary",
]
