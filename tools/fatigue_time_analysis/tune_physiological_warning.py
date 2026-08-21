"""Batch-tune physiological warning h and confirmation duration on train_data.

The script reuses the per-second robust score already exported by Function Two;
it never reruns EDF detection or Observe.  The test set is excluded by reading
only the record IDs explicitly listed in ``train_data``.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from eeg_analysis.fatigue_driving_prediction_system.physiological_fatigue import (
    ceil_to_one_decimal,
)


DEFAULT_MANIFEST = PROJECT_ROOT / "train_data"
DEFAULT_RESULTS_ROOT = (
    PROJECT_ROOT / "data" / "derived" / "fatigue_driving_prediction_system"
)
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data" / "derived" / "fatigue_parameter_tuning"

H_VALUES = (0.80, 0.85, 0.90, 0.95, 1.00, 1.05, 1.10)
DURATION_VALUES = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10)
SCORE_COLUMN = "S=min(Z_Alpha,Z_Eye)"
SUMMARY_SHEET = "功能二摘要"
FEATURE_SHEET = "功能二逐秒特徵"
RT_SHEET = "300秒後RT事件"


@dataclass(frozen=True)
class RecordingScoreData:
    record_id: str
    subject_id: str
    workbook_path: Path
    analysis_start_second: int
    personalized_rt_threshold: float | None
    source_target_second: int | None
    target_second: int | None
    seconds: np.ndarray
    scores: np.ndarray


class DataValidationError(ValueError):
    """Raised when strict training-input validation finds any issue."""

    def __init__(self, errors: Sequence[str]):
        self.errors = tuple(errors)
        super().__init__("\n".join(self.errors))


def parse_training_manifest(path: str | Path) -> list[str]:
    """Read record IDs separated by whitespace or Chinese/ASCII commas."""
    manifest = Path(path)
    if not manifest.is_file():
        raise DataValidationError([f"找不到訓練清單：{manifest}"])
    tokens = [
        item.strip()
        for item in re.split(r"[\s,，、]+", manifest.read_text(encoding="utf-8"))
        if item.strip()
    ]
    if not tokens:
        raise DataValidationError([f"訓練清單是空的：{manifest}"])
    duplicates = sorted({item for item in tokens if tokens.count(item) > 1})
    if duplicates:
        raise DataValidationError(
            ["train_data 含重複record：" + "、".join(duplicates)]
        )
    return tokens


def subject_id_from_record(record_id: str) -> str:
    match = re.match(r"^(s\d+)", record_id.strip(), flags=re.IGNORECASE)
    return match.group(1).casefold() if match else record_id.split("_", 1)[0].casefold()


def _summary_mapping(workbook_path: Path) -> dict[str, object]:
    summary = pd.read_excel(workbook_path, sheet_name=SUMMARY_SHEET)
    if not {"項目", "數值"}.issubset(summary.columns):
        raise ValueError("功能二摘要缺少『項目／數值』欄")
    return {
        str(row["項目"]).strip(): row["數值"]
        for _, row in summary.iterrows()
        if pd.notna(row["項目"])
    }


def _target_with_rounded_rt_threshold(
    workbook_path: Path,
    *,
    analysis_start_second: int,
    threshold: float,
) -> int | None:
    events = pd.read_excel(workbook_path, sheet_name=RT_SHEET)
    if "事件秒數" not in events.columns:
        raise ValueError("RT事件表缺少欄位：事件秒數")
    event_seconds = pd.to_numeric(events["事件秒數"], errors="coerce")
    if "Behavioral Fatigue" in events.columns:
        fatigue_flags = events["Behavioral Fatigue"].isin(
            [True, 1, "TRUE", "True", "true"]
        )
        eligible = pd.DataFrame(
            {"event_second": event_seconds, "behavioral_fatigue": fatigue_flags}
        ).dropna(subset=["event_second"])
        eligible = eligible[
            eligible["event_second"].ge(analysis_start_second)
            & eligible["behavioral_fatigue"]
        ].sort_values("event_second", kind="stable")
    else:
        rt_column = next(
            (
                column
                for column in ("Local RT_秒", "Reaction Time_秒")
                if column in events.columns
            ),
            None,
        )
        if rt_column is None:
            raise ValueError(
                "RT事件表缺少Behavioral Fatigue或Local RT／Reaction Time欄位"
            )
        reaction_times = pd.to_numeric(events[rt_column], errors="coerce")
        eligible = pd.DataFrame(
            {"event_second": event_seconds, "reaction_time": reaction_times}
        ).dropna()
        eligible = eligible[
            eligible["event_second"].ge(analysis_start_second)
            & eligible["reaction_time"].ge(threshold)
        ].sort_values("event_second", kind="stable")
    if eligible.empty:
        return None
    return int(eligible.iloc[0]["event_second"])


def load_recording_score(
    record_id: str,
    results_root: str | Path,
) -> RecordingScoreData:
    workbook_path = Path(results_root) / record_id / "function_two_results.xlsx"
    if not workbook_path.is_file():
        raise FileNotFoundError(f"找不到 {record_id} 的結果：{workbook_path}")

    summary = _summary_mapping(workbook_path)
    features = pd.read_excel(workbook_path, sheet_name=FEATURE_SHEET)
    required_columns = {"秒數", SCORE_COLUMN}
    missing_columns = required_columns.difference(features.columns)
    if missing_columns:
        raise ValueError("逐秒特徵缺少欄位：" + "、".join(sorted(missing_columns)))

    seconds = pd.to_numeric(features["秒數"], errors="coerce")
    scores = pd.to_numeric(features[SCORE_COLUMN], errors="coerce")
    valid_seconds = seconds.notna()
    frame = pd.DataFrame(
        {"second": seconds[valid_seconds], "score": scores[valid_seconds]}
    ).sort_values("second", kind="stable")
    if frame.empty:
        raise ValueError("逐秒特徵沒有有效秒數")
    if frame["second"].duplicated().any():
        raise ValueError("逐秒特徵含重複秒數")

    start_value = summary.get("分析開始秒", 301)
    if pd.isna(start_value):
        start_value = 301
    analysis_start_second = int(float(start_value))
    target_value = summary.get("第一個疲勞事件_秒")
    source_target_second = (
        None if target_value is None or pd.isna(target_value) else int(float(target_value))
    )
    threshold_value = summary.get("個人化RT疲勞門檻")
    personalized_rt_threshold: float | None = None
    target_second = source_target_second
    if threshold_value is not None and not pd.isna(threshold_value):
        personalized_rt_threshold = ceil_to_one_decimal(float(threshold_value))
        if personalized_rt_threshold <= 0:
            raise ValueError("個人化RT疲勞門檻必須大於0")
        rounded_target = _target_with_rounded_rt_threshold(
            workbook_path,
            analysis_start_second=analysis_start_second,
            threshold=personalized_rt_threshold,
        )
        if rounded_target != source_target_second:
            raise ValueError(
                "RT門檻向上取一位小數後，第一個疲勞事件由"
                f"{source_target_second}變成{rounded_target}；"
                "請先重跑Function One/Two，避免逐秒分數長度與目標事件不一致"
            )
        target_second = rounded_target
    workbook_record_id = summary.get("record_id")
    if workbook_record_id is not None and str(workbook_record_id) != record_id:
        raise ValueError(
            f"摘要record_id={workbook_record_id}，與train_data的{record_id}不一致"
        )

    return RecordingScoreData(
        record_id=record_id,
        subject_id=subject_id_from_record(record_id),
        workbook_path=workbook_path,
        analysis_start_second=analysis_start_second,
        personalized_rt_threshold=personalized_rt_threshold,
        source_target_second=source_target_second,
        target_second=target_second,
        seconds=frame["second"].to_numpy(dtype=int),
        scores=frame["score"].to_numpy(dtype=float),
    )


def load_training_recordings(
    manifest_path: str | Path,
    results_root: str | Path,
    *,
    expected_records: int = 7,
) -> list[RecordingScoreData]:
    record_ids = parse_training_manifest(manifest_path)
    errors: list[str] = []
    if expected_records > 0 and len(record_ids) != expected_records:
        errors.append(
            f"train_data目前有{len(record_ids)}筆，預期必須是{expected_records}筆。"
        )

    recordings: list[RecordingScoreData] = []
    for record_id in record_ids:
        try:
            recordings.append(load_recording_score(record_id, results_root))
        except Exception as exc:
            errors.append(f"{record_id}：{exc}")
    if errors:
        raise DataValidationError(errors)
    return recordings


def first_confirmed_alarm(
    seconds: Sequence[int] | np.ndarray,
    scores: Sequence[float] | np.ndarray,
    *,
    h: float,
    duration_seconds: int,
    analysis_start_second: int = 301,
) -> int | None:
    """Return the real-time confirmation second without backdating."""
    if duration_seconds <= 0:
        raise ValueError("duration_seconds must be positive")
    run_length = 0
    previous_second: int | None = None
    for second_value, score_value in zip(seconds, scores):
        second = int(second_value)
        if second < analysis_start_second:
            continue
        if previous_second is not None and second != previous_second + 1:
            run_length = 0
        above = np.isfinite(score_value) and float(score_value) >= h
        run_length = run_length + 1 if above else 0
        previous_second = second
        if run_length >= duration_seconds:
            return second
    return None


def classify_record_result(
    recording: RecordingScoreData,
    alarm_second: int | None,
    *,
    min_lead_seconds: int = 30,
    max_lead_seconds: int = 60,
) -> tuple[str, int | None]:
    target = recording.target_second
    if target is None:
        return (
            "NO_TARGET_FALSE_ALARM" if alarm_second is not None else "NO_TARGET_NO_ALARM",
            None,
        )

    target_window_start = max(
        recording.analysis_start_second,
        target - max_lead_seconds,
    )
    target_window_end = target - min_lead_seconds
    if target_window_start > target_window_end:
        return "NOT_EVALUABLE", None
    if alarm_second is None:
        return "MISS", None

    lead = target - alarm_second
    if lead > max_lead_seconds:
        return "TOO_EARLY", lead
    if min_lead_seconds <= lead <= max_lead_seconds:
        return "SUCCESS", lead
    if 0 < lead < min_lead_seconds:
        return "TOO_LATE", lead
    return "AT_OR_AFTER_TARGET", lead


def evaluate_parameter_grid(
    recordings: Sequence[RecordingScoreData],
    *,
    h_values: Iterable[float] = H_VALUES,
    duration_values: Iterable[int] = DURATION_VALUES,
    maximum_early_false_count: int = 1,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    detail_rows: list[dict[str, object]] = []
    for h in h_values:
        for duration in duration_values:
            for recording in recordings:
                alarm = first_confirmed_alarm(
                    recording.seconds,
                    recording.scores,
                    h=float(h),
                    duration_seconds=int(duration),
                    analysis_start_second=recording.analysis_start_second,
                )
                classification, lead = classify_record_result(recording, alarm)
                detail_rows.append(
                    {
                        "record_id": recording.record_id,
                        "subject_id": recording.subject_id,
                        "h": float(h),
                        "duration_seconds": int(duration),
                        "target_second": recording.target_second,
                        "alarm_second": alarm,
                        "lead_seconds": lead,
                        "classification": classification,
                    }
                )
    details = pd.DataFrame(detail_rows)

    metric_rows: list[dict[str, object]] = []
    for (h, duration), group in details.groupby(
        ["h", "duration_seconds"], sort=True
    ):
        evaluable = group[
            ~group["classification"].isin(
                ["NOT_EVALUABLE", "NO_TARGET_NO_ALARM", "NO_TARGET_FALSE_ALARM"]
            )
        ].copy()
        success = evaluable["classification"].eq("SUCCESS")
        subject_success = (
            evaluable.assign(success=success.astype(float))
            .groupby("subject_id")["success"]
            .mean()
        )
        early_false_count = int(
            group["classification"].isin(["TOO_EARLY", "NO_TARGET_FALSE_ALARM"]).sum()
        )
        success_leads = pd.to_numeric(
            group.loc[group["classification"].eq("SUCCESS"), "lead_seconds"],
            errors="coerce",
        ).dropna()
        median_lead = float(success_leads.median()) if not success_leads.empty else math.nan
        metric_rows.append(
            {
                "h": float(h),
                "duration_seconds": int(duration),
                "evaluable_count": int(len(evaluable)),
                "success_count": int(success.sum()),
                "too_early_count": int(group["classification"].eq("TOO_EARLY").sum()),
                "too_late_count": int(group["classification"].eq("TOO_LATE").sum()),
                "miss_count": int(group["classification"].eq("MISS").sum()),
                "at_or_after_count": int(
                    group["classification"].eq("AT_OR_AFTER_TARGET").sum()
                ),
                "not_evaluable_count": int(
                    group["classification"].eq("NOT_EVALUABLE").sum()
                ),
                "no_target_false_alarm_count": int(
                    group["classification"].eq("NO_TARGET_FALSE_ALARM").sum()
                ),
                "early_false_count": early_false_count,
                "record_success_rate": (
                    float(success.mean()) if len(evaluable) else math.nan
                ),
                "subject_success_rate": (
                    float(subject_success.mean()) if not subject_success.empty else math.nan
                ),
                "median_success_lead": median_lead,
                "lead_distance_from_45": (
                    abs(median_lead - 45) if np.isfinite(median_lead) else math.inf
                ),
                "early_false_constraint_ok": (
                    early_false_count <= maximum_early_false_count
                ),
            }
        )

    metrics = pd.DataFrame(metric_rows)
    ranked = metrics.assign(
        _subject_success=metrics["subject_success_rate"].fillna(-1),
        _lead_distance=metrics["lead_distance_from_45"].replace(
            [np.inf, -np.inf], 1_000_000
        ),
    ).sort_values(
        [
            "early_false_constraint_ok",
            "_subject_success",
            "early_false_count",
            "miss_count",
            "too_late_count",
            "_lead_distance",
            "h",
            "duration_seconds",
        ],
        ascending=[False, False, True, True, True, True, False, False],
        kind="stable",
    )
    ranked = ranked.drop(columns=["_subject_success", "_lead_distance"]).reset_index(
        drop=True
    )
    ranked.insert(0, "rank", np.arange(1, len(ranked) + 1))
    return ranked, details


def _format_heatmap_axes(axis, h_values: Sequence[float], durations: Sequence[int]) -> None:
    axis.set_xticks(range(len(h_values)), [f"{value:.2f}" for value in h_values])
    axis.set_yticks(range(len(durations)), [str(value) for value in durations])
    axis.set_xlabel("h")
    axis.set_ylabel("Confirmation duration (seconds)")


def _save_heatmap(
    ranked: pd.DataFrame,
    *,
    value_column: str,
    output_path: Path,
    title: str,
    color_map: str,
    annotation,
    vmin: float | None = None,
    vmax: float | None = None,
) -> None:
    h_values = sorted(ranked["h"].unique())
    durations = sorted(ranked["duration_seconds"].unique())
    lookup = ranked.set_index(["duration_seconds", "h"])
    matrix = np.array(
        [
            [lookup.loc[(duration, h), value_column] for h in h_values]
            for duration in durations
        ],
        dtype=float,
    )
    figure, axis = plt.subplots(figsize=(8.5, 5.2))
    image = axis.imshow(matrix, aspect="auto", cmap=color_map, vmin=vmin, vmax=vmax)
    _format_heatmap_axes(axis, h_values, durations)
    axis.set_title(title)
    for row_index, duration in enumerate(durations):
        for column_index, h in enumerate(h_values):
            record = lookup.loc[(duration, h)]
            axis.text(
                column_index,
                row_index,
                annotation(record),
                ha="center",
                va="center",
                fontsize=9,
                color="black",
            )
    figure.colorbar(image, ax=axis, shrink=0.9)
    figure.tight_layout()
    figure.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(figure)


def save_heatmaps(ranked: pd.DataFrame, output_dir: Path) -> None:
    _save_heatmap(
        ranked,
        value_column="subject_success_rate",
        output_path=output_dir / "success_heatmap.png",
        title="Subject-balanced success rate (lead 30–60 s)",
        color_map="YlGn",
        annotation=lambda row: (
            f"{int(row.success_count)}/{int(row.evaluable_count)}\n"
            f"subj={float(row.subject_success_rate):.2f}"
        ),
        vmin=0,
        vmax=1,
    )
    _save_heatmap(
        ranked,
        value_column="early_false_count",
        output_path=output_dir / "too_early_heatmap.png",
        title="Too-early and no-target false alarms (lower is better)",
        color_map="Reds",
        annotation=lambda row: str(int(row.early_false_count)),
        vmin=0,
    )
    _save_heatmap(
        ranked,
        value_column="median_success_lead",
        output_path=output_dir / "median_lead_heatmap.png",
        title="Median lead among successful records (target center = 45 s)",
        color_map="viridis",
        annotation=lambda row: (
            "--"
            if pd.isna(row.median_success_lead)
            else f"{float(row.median_success_lead):.1f}"
        ),
        vmin=30,
        vmax=60,
    )


def save_best_overview(
    recordings: Sequence[RecordingScoreData],
    selected_details: pd.DataFrame,
    *,
    h: float,
    duration_seconds: int,
    output_path: Path,
) -> None:
    plt.rcParams["font.sans-serif"] = ["Microsoft JhengHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    figure, axes = plt.subplots(
        len(recordings),
        1,
        figsize=(14, max(3.0 * len(recordings), 5)),
        squeeze=False,
    )
    details_by_record = selected_details.set_index("record_id")
    for axis, recording in zip(axes[:, 0], recordings):
        detail = details_by_record.loc[recording.record_id]
        axis.plot(recording.seconds, recording.scores, color="#008C95", linewidth=1.2)
        axis.axhline(h, color="#C00000", linestyle="--", label=f"h={h:.2f}")
        alarm = detail["alarm_second"]
        target = recording.target_second
        if target is not None:
            target_start = max(recording.analysis_start_second, target - 60)
            target_end = target - 30
            if target_start <= target_end:
                axis.axvspan(
                    target_start,
                    target_end,
                    color="#D9EAD3",
                    alpha=0.5,
                    label="Target lead 30–60 s",
                )
            axis.axvline(target, color="#C00000", linestyle="-.", label="Behavioral fatigue")
        if pd.notna(alarm):
            axis.axvline(
                float(alarm),
                color="#4472C4",
                linestyle="-.",
                label="Confirmed warning",
            )

        lead = detail["lead_seconds"]
        lead_text = "--" if pd.isna(lead) else f"{int(lead)} s"
        axis.set_title(
            f"{recording.record_id} ({recording.subject_id}) | "
            f"{detail['classification']} | lead={lead_text}",
            loc="left",
            fontsize=10,
        )
        axis.set_ylabel("S")
        axis.grid(True, linestyle="--", alpha=0.25)
        if target is not None:
            default_start = target - 120
            if pd.notna(alarm):
                default_start = min(default_start, int(alarm) - 10)
            axis.set_xlim(max(recording.analysis_start_second, default_start), target)
        axis.legend(loc="upper right", fontsize=7, ncol=3)
    axes[-1, 0].set_xlabel("Second")
    figure.suptitle(
        f"Top-ranked training configuration: h={h:.2f}, duration={duration_seconds} s",
        fontsize=14,
    )
    figure.tight_layout(rect=(0, 0, 1, 0.98))
    figure.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(figure)


def _autosize_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for column_index in range(1, worksheet.max_column + 1):
            width = max(
                (
                    len(str(worksheet.cell(row, column_index).value or ""))
                    for row in range(1, worksheet.max_row + 1)
                ),
                default=0,
            )
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(width + 2, 12), 38
            )
    workbook.save(path)


def write_outputs(
    recordings: Sequence[RecordingScoreData],
    ranked: pd.DataFrame,
    details: pd.DataFrame,
    *,
    manifest_path: Path,
    output_dir: Path,
) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    best = ranked.iloc[0]
    best_h = float(best["h"])
    best_duration = int(best["duration_seconds"])
    selected_details = details[
        np.isclose(details["h"], best_h)
        & details["duration_seconds"].eq(best_duration)
    ].copy()

    data_check = pd.DataFrame(
        [
            {
                "record_id": item.record_id,
                "subject_id": item.subject_id,
                "workbook_path": str(item.workbook_path),
                "analysis_start_second": item.analysis_start_second,
                "personalized_rt_threshold": item.personalized_rt_threshold,
                "source_target_second": item.source_target_second,
                "target_second": item.target_second,
                "score_rows": len(item.seconds),
                "finite_score_rows": int(np.isfinite(item.scores).sum()),
            }
            for item in recordings
        ]
    )
    subject_results = (
        details.assign(success=details["classification"].eq("SUCCESS").astype(int))
        .groupby(["subject_id", "h", "duration_seconds"], as_index=False)
        .agg(
            record_count=("record_id", "nunique"),
            subject_success_rate=("success", "mean"),
            median_lead=("lead_seconds", "median"),
        )
    )

    workbook_path = output_dir / "parameter_tuning_results.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        ranked.to_excel(writer, sheet_name="參數排名", index=False)
        details.to_excel(writer, sheet_name="逐筆結果", index=False)
        subject_results.to_excel(writer, sheet_name="受試者結果", index=False)
        selected_details.to_excel(writer, sheet_name="最佳設定逐筆結果", index=False)
        data_check.to_excel(writer, sheet_name="資料檢查", index=False)
    _autosize_workbook(workbook_path)

    save_heatmaps(ranked, output_dir)
    overview_path = output_dir / "best_configuration_overview.png"
    save_best_overview(
        recordings,
        selected_details,
        h=best_h,
        duration_seconds=best_duration,
        output_path=overview_path,
    )

    configuration_path = output_dir / "selected_configuration.json"
    configuration = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "manifest": str(manifest_path.resolve()),
        "records": [item.record_id for item in recordings],
        "subjects": sorted({item.subject_id for item in recordings}),
        "h": best_h,
        "confirmation_seconds": best_duration,
        "early_false_constraint_met": bool(best["early_false_constraint_ok"]),
        "selection_rule": (
            "early/false alarms <= 1; then maximize subject-balanced success; "
            "then minimize misses/late warnings; then lead nearest 45 seconds"
        ),
        "metrics": {
            key: (None if pd.isna(value) or value in (math.inf, -math.inf) else value)
            for key, value in best.drop(labels=["rank"]).to_dict().items()
        },
    }
    configuration_path.write_text(
        json.dumps(configuration, ensure_ascii=False, indent=2, default=float) + "\n",
        encoding="utf-8",
    )
    return {
        "workbook": workbook_path,
        "success_heatmap": output_dir / "success_heatmap.png",
        "too_early_heatmap": output_dir / "too_early_heatmap.png",
        "median_lead_heatmap": output_dir / "median_lead_heatmap.png",
        "overview": overview_path,
        "configuration": configuration_path,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Tune h=0.80..1.00 and 1..5-second confirmation on train_data."
    )
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--results-root", type=Path, default=DEFAULT_RESULTS_ROOT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--expected-records", type=int, default=7)
    parser.add_argument("--h-values", nargs="+", type=float, default=list(H_VALUES))
    parser.add_argument(
        "--duration-values", nargs="+", type=int, default=list(DURATION_VALUES)
    )
    parser.add_argument("--maximum-early-false-count", type=int, default=1)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        recordings = load_training_recordings(
            args.manifest,
            args.results_root,
            expected_records=args.expected_records,
        )
        ranked, details = evaluate_parameter_grid(
            recordings,
            h_values=args.h_values,
            duration_values=args.duration_values,
            maximum_early_false_count=args.maximum_early_false_count,
        )
        outputs = write_outputs(
            recordings,
            ranked,
            details,
            manifest_path=args.manifest,
            output_dir=args.output_dir,
        )
    except DataValidationError as exc:
        print("訓練資料驗證失敗：", file=sys.stderr)
        for error in exc.errors:
            print(f"- {error}", file=sys.stderr)
        raise SystemExit(2) from exc

    best = ranked.iloc[0]
    print(
        "訓練集排名第一設定："
        f"h={float(best['h']):.2f}, "
        f"持續{int(best['duration_seconds'])}秒, "
        f"成功={int(best['success_count'])}/{int(best['evaluable_count'])}, "
        f"太早/假警報={int(best['early_false_count'])}"
    )
    if not bool(best["early_false_constraint_ok"]):
        print(
            "警告：沒有任何組合符合太早/假警報上限；"
            "此設定只是目前搜尋範圍中的相對最佳者，不應直接視為定案。"
        )
    for name, path in outputs.items():
        print(f"{name}: {path}")


if __name__ == "__main__":
    main()


__all__ = [
    "DURATION_VALUES",
    "H_VALUES",
    "DataValidationError",
    "RecordingScoreData",
    "classify_record_result",
    "evaluate_parameter_grid",
    "first_confirmed_alarm",
    "load_recording_score",
    "load_training_recordings",
    "parse_training_manifest",
    "write_outputs",
]
