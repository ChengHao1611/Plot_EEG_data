"""Analyse lane-crossing times from an EDF vehicle-position recording.

The input EDF must contain both a ``vehicle position`` channel and a ``Status``
channel.  The Status codes are interpreted as:

* 251 / 252: deviation onset
* 253: response onset
* 254: response offset

For each complete event, the script looks for the first 60-unit lateral
displacement before Status 253.  If it is not reached, it uses the Status
reaction time (253 minus 251/252) instead.  The latter is also used to create
the requested 60 km/h estimate by scaling the 100 km/h time by 100 / 60.

Example
-------
    python -m tools.fatigue_time_analysis.analyze_lane_crossing_events \
        data/raw_edf/auxiliary/s01_060926_1n_car_position.EDF
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pyedflib
from openpyxl import Workbook
from openpyxl.styles import Font


LANE_WIDTH_UNITS = 60.0
SOURCE_SPEED_KMH = 100.0
TARGET_SPEED_KMH = 60.0


@dataclass
class Marker:
    """One de-duplicated Status marker."""

    code: int
    time_seconds: float


@dataclass
class Event:
    """A matched deviation-onset, response-onset, response-offset triplet."""

    start: Marker
    response_onset: Marker
    response_offset: Marker


def find_channel_index(labels: Iterable[str], expected_label: str) -> int:
    """Return the channel index matched case-insensitively by its label."""
    expected = expected_label.casefold().strip()
    for index, label in enumerate(labels):
        if label.casefold().strip() == expected:
            return index

    for index, label in enumerate(labels):
        if expected in label.casefold():
            return index

    raise ValueError(f"Could not find the {expected_label!r} channel in this EDF.")


def read_channels(edf_path: Path) -> tuple[np.ndarray, float, np.ndarray, float]:
    """Read vehicle-position coordinates and Status codes from an EDF file."""
    reader = pyedflib.EdfReader(str(edf_path))
    try:
        labels = reader.getSignalLabels()
        position_index = find_channel_index(labels, "vehicle position")
        status_index = find_channel_index(labels, "status")

        # The vehicle channel uses a generic EEG calibration in this EDF.  Its
        # decoded values are nevertheless within the documented 0-255 position
        # coordinate range, apart from terminal padding.  Round to coordinates.
        position = np.rint(reader.readSignal(position_index)).astype(float)
        status = np.rint(reader.readSignal(status_index)).astype(int)
        position_rate = float(reader.getSampleFrequency(position_index))
        status_rate = float(reader.getSampleFrequency(status_index))
    finally:
        reader.close()

    if position_rate <= 0 or status_rate <= 0:
        raise ValueError("The vehicle-position and Status sample rates must be positive.")

    return position, position_rate, status, status_rate


def extract_markers(status: np.ndarray, sample_rate: float) -> list[Marker]:
    """Extract one marker for each consecutive run of 251/252/253/254."""
    event_codes = {251, 252, 253, 254}
    markers: list[Marker] = []
    previous_code: int | None = None

    for index, code in enumerate(status):
        code = int(code)
        if code in event_codes and code != previous_code:
            markers.append(Marker(code=code, time_seconds=index / sample_rate))
        previous_code = code

    return markers


def match_events(markers: Iterable[Marker]) -> tuple[list[Event], list[tuple[Marker, str]]]:
    """Match 251/252 -> 253 -> 254 markers and retain incomplete events."""
    complete_events: list[Event] = []
    incomplete_events: list[tuple[Marker, str]] = []
    start: Marker | None = None
    response_onset: Marker | None = None

    for marker in markers:
        if marker.code in {251, 252}:
            if start is not None:
                reason = "A new deviation-onset marker occurred before the previous event ended."
                incomplete_events.append((start, reason))
            start = marker
            response_onset = None

        elif marker.code == 253:
            if start is None:
                incomplete_events.append((marker, "Response onset without a preceding deviation onset."))
            elif response_onset is None:
                response_onset = marker
            else:
                incomplete_events.append((marker, "Repeated response-onset marker in one event."))

        elif marker.code == 254:
            if start is None:
                incomplete_events.append((marker, "Response offset without a preceding deviation onset."))
            elif response_onset is None:
                incomplete_events.append((start, "Response offset occurred before response onset (253)."))
                start = None
            else:
                complete_events.append(
                    Event(
                        start=start,
                        response_onset=response_onset,
                        response_offset=marker,
                    )
                )
                start = None
                response_onset = None

    if start is not None:
        incomplete_events.append((start, "The recording ended before the event was completed."))

    return complete_events, incomplete_events


def position_index_at_time(time_seconds: float, sample_rate: float, sample_count: int) -> int:
    """Return the closest valid position sample index for an event time."""
    index = int(round(time_seconds * sample_rate))
    return min(max(index, 0), sample_count - 1)


def first_lane_crossing(
    position: np.ndarray,
    sample_rate: float,
    start_time: float,
    end_time: float,
) -> tuple[float | None, float, float]:
    """Find the first 60-unit displacement between start and end times.

    Returns ``(crossing_time, start_coordinate, maximum_displacement)``.  The
    crossing time is linearly interpolated between adjacent 500-Hz samples.
    """
    start_index = position_index_at_time(start_time, sample_rate, len(position))
    end_index = position_index_at_time(end_time, sample_rate, len(position))
    if end_index <= start_index:
        return None, float(position[start_index]), 0.0

    values = position[start_index : end_index + 1]
    start_coordinate = float(values[0])
    displacement = values - start_coordinate
    absolute_displacement = np.abs(displacement)
    maximum_displacement = float(np.max(absolute_displacement))
    crossing_offsets = np.flatnonzero(absolute_displacement >= LANE_WIDTH_UNITS)
    if not len(crossing_offsets):
        return None, start_coordinate, maximum_displacement

    crossing_offset = int(crossing_offsets[0])
    crossing_index = start_index + crossing_offset
    current_value = float(position[crossing_index])
    target = start_coordinate + np.sign(current_value - start_coordinate) * LANE_WIDTH_UNITS

    if crossing_offset == 0:
        return crossing_index / sample_rate, start_coordinate, maximum_displacement

    previous_value = float(position[crossing_index - 1])
    change = current_value - previous_value
    fraction = 1.0 if change == 0 else (target - previous_value) / change
    fraction = min(max(fraction, 0.0), 1.0)
    crossing_time = (crossing_index - 1 + fraction) / sample_rate
    return crossing_time, start_coordinate, maximum_displacement


def analyse_events(position: np.ndarray, position_rate: float, events: Iterable[Event]) -> list[dict[str, object]]:
    """Create one Excel-ready result row for each complete Status event."""
    rows: list[dict[str, object]] = []

    for event_number, event in enumerate(events, start=1):
        reaction_time = event.response_onset.time_seconds - event.start.time_seconds
        crossing_time, start_position, maximum_displacement = first_lane_crossing(
            position,
            position_rate,
            event.start.time_seconds,
            event.response_onset.time_seconds,
        )

        if crossing_time is None:
            source = "未跨滿 60 單位；採用 Status 反應時間"
            time_at_100 = reaction_time
            crossing_position: float | None = None
            note = "事件開始至 Status 253 前的最大偏移不足一個車道寬。"
        else:
            source = "跨越 60 單位"
            time_at_100 = crossing_time - event.start.time_seconds
            crossing_position = start_position + np.sign(
                position[position_index_at_time(crossing_time, position_rate, len(position))]
                - start_position
            ) * LANE_WIDTH_UNITS
            note = ""

        time_at_60 = time_at_100 * SOURCE_SPEED_KMH / TARGET_SPEED_KMH
        rows.append(
            {
                "事件編號": event_number,
                "事件開始時間_秒": event.start.time_seconds,
                "開始位置_單位": start_position,
                "跨 60 單位時刻_秒": crossing_time,
                "跨線位置_單位": crossing_position,
                "最大偏移_單位": maximum_displacement,
                "採用時間來源": source,
                "100km_h時間_秒": time_at_100,
                "60km_h推估時間_秒": time_at_60,
                "Status反應時間_秒": reaction_time,
                "備註": note,
            }
        )

    return rows


def style_worksheet(worksheet) -> None:
    """Apply light formatting suitable for the generated report."""
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for column_cells in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(width, 42)


def write_xlsx(
    output_path: Path,
    edf_path: Path,
    rows: list[dict[str, object]],
    incomplete_events: list[tuple[Marker, str]],
) -> None:
    """Write event data, incomplete events, and calculation definitions to Excel."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "事件分析"

    columns = [
        "事件編號",
        "事件開始時間_秒",
        "開始位置_單位",
        "跨 60 單位時刻_秒",
        "跨線位置_單位",
        "最大偏移_單位",
        "採用時間來源",
        "100km_h時間_秒",
        "60km_h推估時間_秒",
        "Status反應時間_秒",
        "備註",
    ]
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row[column] for column in columns])
    style_worksheet(worksheet)

    for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=10):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"

    incomplete_sheet = workbook.create_sheet("未完成事件")
    incomplete_sheet.append(["Status時間_秒", "Status代碼", "原因"])
    for marker, reason in incomplete_events:
        incomplete_sheet.append([marker.time_seconds, marker.code, reason])
    style_worksheet(incomplete_sheet)
    for cell in incomplete_sheet["A"][1:]:
        cell.number_format = "0.000"

    summary_sheet = workbook.create_sheet("說明")
    summary_rows = [
        ("輸入 EDF", str(edf_path.resolve())),
        ("完整事件數", len(rows)),
        ("跨滿一車道事件數", sum(row["採用時間來源"] == "跨越 60 單位" for row in rows)),
        ("未跨滿一車道事件數", sum(row["採用時間來源"] != "跨越 60 單位" for row in rows)),
        ("未完成事件數", len(incomplete_events)),
        ("車道寬度", "60 單位"),
        ("100 km/h 時間", "事件開始至首次偏移 60 單位；未跨滿時改用 Status 反應時間 (253 - 251/252)。"),
        ("60 km/h 推估", "100 km/h 時間 × 100 / 60；為等比例推估，不是重新量測的受試者反應。"),
        ("Status 反應時間", "253 - 251/252。"),
        ("事件配對", "251 或 252 → 253 → 254。"),
    ]
    for summary_row in summary_rows:
        summary_sheet.append(summary_row)
    style_worksheet(summary_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def default_output_path(edf_path: Path) -> Path:
    return edf_path.with_name(f"{edf_path.stem}_lane_crossing_analysis.xlsx")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export EDF vehicle-position lane-crossing event times to XLSX."
    )
    parser.add_argument("edf_path", type=Path, help="EDF containing vehicle position and Status.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output XLSX path (default: beside the EDF file).",
    )
    args = parser.parse_args()

    if not args.edf_path.is_file():
        parser.error(f"EDF file not found: {args.edf_path}")

    output_path = args.output or default_output_path(args.edf_path)
    try:
        position, position_rate, status, status_rate = read_channels(args.edf_path)
        markers = extract_markers(status, status_rate)
        events, incomplete_events = match_events(markers)
        rows = analyse_events(position, position_rate, events)
        write_xlsx(output_path, args.edf_path, rows, incomplete_events)
    except (OSError, ValueError, RuntimeError) as error:
        parser.error(str(error))

    crossed = sum(row["採用時間來源"] == "跨越 60 單位" for row in rows)
    print(f"完整事件：{len(rows)}；跨滿一車道：{crossed}；未完成事件：{len(incomplete_events)}")
    print(f"已輸出：{output_path.resolve()}")


if __name__ == "__main__":
    main()
