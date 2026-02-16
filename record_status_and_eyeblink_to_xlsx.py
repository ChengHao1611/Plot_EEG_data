import pyedflib
import numpy as np
from openpyxl import Workbook  # 新增
import os  # 新增
from openpyxl.utils import get_column_letter

def process_eye_blink_data(ws, base_path):
    """
    處理眼動資料並填入 F 欄
    base_path: EDF 檔案的路徑（不含副檔名）
    """
    # 尋找對應的 .dat 檔案
    dat_dir = os.path.dirname(base_path)
    dat_filename = os.path.basename(base_path) + "_raw_arousal info.dat"
    dat_path = os.path.join(dat_dir, dat_filename)
    
    if not os.path.exists(dat_path):
        print(f"警告：找不到眼動資料檔案 {dat_path}")
        return
    
    # 讀取 .dat 檔案
    with open(dat_path, 'r', encoding='utf-8') as f:
        content = f.read().strip()
    
    if not content:
        print(f"警告：{dat_path} 檔案為空")
        return
    
    # 解析資料：去掉第一個數字，剩下的轉為整數列表
    numbers = [int(x.strip()) for x in content.split(',')]
    if len(numbers) < 2:
        print(f"警告：{dat_path} 資料格式不正確")
        return
    
    # 去掉第一個數字
    time_points = numbers[1:]
    
    # 按30秒區間統計
    # 建立字典：{區間起始秒數: 該區間內的數字個數}
    interval_counts = {}
    for time_point in time_points:
        # 計算該時間點屬於哪個30秒區間
        interval_start = (time_point // 30) * 30
        if interval_start not in interval_counts:
            interval_counts[interval_start] = 0
        interval_counts[interval_start] += 1
    
    # 建立 A 欄的秒數到行號的映射（只記錄完全匹配的行）
    a_column_to_row = {}
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        a_value = ws.cell(row=row, column=1).value
        if a_value is not None:
            sec_value = float(a_value)
            # 只記錄整數秒數（完全匹配）
            if sec_value == int(sec_value):
                a_column_to_row[int(sec_value)] = row
    
    # 填入或插入眼動次數資料
    for interval_start, count in sorted(interval_counts.items()):
        if interval_start in a_column_to_row:
            # 如果該秒數已存在，直接填入 F 欄
            row = a_column_to_row[interval_start]
            ws.cell(row=row, column=6, value=count)
        else:
            # 如果該秒數不存在，需要插入新行
            # 找到應該插入的位置（按秒數排序）
            insert_row = 2
            max_row = ws.max_row
            for row in range(2, max_row + 1):
                a_value = ws.cell(row=row, column=1).value
                if a_value is not None:
                    if float(a_value) > interval_start:
                        insert_row = row
                        break
                insert_row = row + 1
            
            # 插入新行
            ws.insert_rows(insert_row)
            ws.cell(row=insert_row, column=1, value=float(interval_start))
            ws.cell(row=insert_row, column=6, value=count)

def check_status_253(edf_path, tolerance=0.05):
    f = pyedflib.EdfReader(edf_path)

    # --- 新增：準備 xlsx 檔案與表頭 ---
    base_path, _ = os.path.splitext(edf_path)
    xlsx_path = base_path + ".xlsx"   # 與 EDF 同路徑同檔名，副檔名改為 .xlsx

    wb = Workbook()
    ws = wb.active
    ws.title = "result"

    # A1~F1 標題
    ws["A1"] = "秒數"
    ws["B1"] = "事件反應時間"
    ws["C1"] = "α波時間"
    ws["D1"] = "導回車道用時"
    ws["E1"] = "睡著"
    ws["F1"] = "眼動次數"

    next_row = 2  # 下一筆資料要寫入的列數（從第2列開始）

    # ... 下面保持原本程式 ...
    channel_labels = f.getSignalLabels()
    status_index = None
    for i, label in enumerate(channel_labels):
        if 'status' in label.lower():
            status_index = i
            break

    if status_index is None:
        print("未找到 Status 通道，請確認標籤名稱")
        f.close()
        return

    status_signal = f.readSignal(status_index)
    sample_rate = int(f.getSampleFrequency(status_index))
    total_samples = len(status_signal)
    total_seconds = total_samples // sample_rate

    print(f"取樣率: {sample_rate} Hz, 總長度: {total_seconds} 秒")

    stage = 1
    for sec in range(total_seconds):
        start = sec * sample_rate
        end = start + sample_rate
        segment = status_signal[start:end]
        for i in range(len(segment)):
            #print(segment[i])
            if segment[i] > 1:
                if stage == 1:
                    sec_251 = sec + 0.002 * i
                    stage = 2
                elif stage == 2:
                    sec_253 = sec + 0.002 * i
                    stage = 3
                elif stage == 3:
                    sec_254 = sec + 0.002 * i

                    # --- 新增：在 stage==3 時把資料寫入 xlsx ---
                    ws.cell(row=next_row, column=1, value=float(f"{sec_251:.1f}"))                 # 秒數
                    ws.cell(row=next_row, column=2, value=float(f"{sec_253 - sec_251:.1f}"))        # 事件反應時間
                    # C 欄 α波時間：暫時留空
                    ws.cell(row=next_row, column=4, value=float(f"{sec_254 - sec_253:.1f}"))        # 導回車道用時
                    # E 欄 睡著：暫時留空

                    print(
                        f"第{sec_253:.1f}秒的事件反應時間：{sec_253 - sec_251:.1f}秒, "
                        f"導回車道用時：{sec_254 - sec_253:.1f}秒"
                    )

                    next_row += 1
                    stage = 1
                # 若有需要，可把 debug 的 print 打開
                # print(f"秒數 {sec + 0.002 * i} -> Status 平均值: {segment[i]:.2f}, stage={stage}")

    f.close()

    # --- 新增：處理眼動資料並填入 F 欄 ---
    process_eye_blink_data(ws, base_path)
    
    # --- 新增：儲存 xlsx 檔案 ---
    wb.save(xlsx_path)
    #print(f"結果已儲存到: {xlsx_path}")

if __name__ == "__main__":
    # 請將此路徑改為你的EDF檔案路徑
    excel_file = input("請輸入 Excel 檔案路徑: ")
    excel_file = excel_file.replace('"', '') 
    check_status_253(excel_file)